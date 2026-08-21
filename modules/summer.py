import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import hashlib
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from .ui_components import page_header, section_intro


# ── 썸머 기준 ────────────────────────────────────────────────
MONTHLY_TARGET = 500_000
MONTHLY_HANWHA_MIN_PREMIUM = 50_000

READY_BONUS_RATES = [0, 15, 20, 25, 30]

SUMMER_GRADES = [
    ("HWARANG", 15_000_000),
    ("크라운", 10_000_000),
    ("트리플", 8_000_000),
    ("더블", 5_000_000),
    ("일반", 3_000_000),
]

TABLE_SEQ = 0


# ── 기본 유틸 ────────────────────────────────────────────────
def mark(ok: bool) -> str:
    return "✅" if ok else "❌"


def won(x) -> str:
    try:
        return f"{float(x):,.0f} 원"
    except Exception:
        return ""


def signed_won(x) -> str:
    try:
        value = int(float(x))
        return f"{value:+,} 원" if value else "0 원"
    except Exception:
        return ""


def pct(x) -> str:
    try:
        return f"{float(x):,.0f} %"
    except Exception:
        return ""


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def standardize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "계약일" in df.columns and "계약일자" not in df.columns:
        df.rename(columns={"계약일": "계약일자"}, inplace=True)

    if "초회보험료" in df.columns and "보험료" not in df.columns:
        df.rename(columns={"초회보험료": "보험료"}, inplace=True)

    return df


def safe_table_name(base: str) -> str:
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(base))

    if not re.match(r"^[A-Za-z_]", name):
        name = f"tbl_{name}"

    return name[:254]


def safe_filename_part(text: str) -> str:
    """
    파일명에 사용할 수 없는 문자를 제거합니다.
    """
    text = str(text).strip()
    text = re.sub(r'[\\/:*?"<>|]', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text if text else "미지정"


def unique_sheet_name(wb, base, limit=31):
    name = str(base)[:limit] if base else "Sheet"

    if name not in wb.sheetnames:
        return name

    i = 2
    while True:
        suffix = f"_{i}"
        trunc = limit - len(suffix)
        cand = f"{name[:trunc]}{suffix}"

        if cand not in wb.sheetnames:
            return cand

        i += 1


def autosize_columns_full(ws, padding=5):
    for column_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )

        ws.column_dimensions[column_cells[0].column_letter].width = max_len + padding


# ── 보험사 분류 ───────────────────────────────────────────────
def is_hanwha_life_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("한화", na=False)
        & ins.str.contains("생명", na=False)
    )


def is_db_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("DB", case=False, na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_kb_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("KB", case=False, na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_hanwha_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("한화", na=False)
        & (
            ins.str.contains("손", na=False)
            | ins.str.contains("화재", na=False)
            | ins.str.contains("손해", na=False)
        )
        & ~is_hanwha_life_series(ins)
    )


def is_heungkuk_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("흥국", na=False)
        & (
            ins.str.contains("화재", na=False)
            | ins.str.contains("손", na=False)
            | ins.str.contains("손해", na=False)
        )
    )


def is_special_nonlife_series(ins: pd.Series) -> pd.Series:
    """
    썸머 우대 손해보험사:
    흥국화재, KB손해, 한화손해, DB손해
    """
    return (
        is_db_nonlife_series(ins)
        | is_kb_nonlife_series(ins)
        | is_hanwha_nonlife_series(ins)
        | is_heungkuk_nonlife_series(ins)
    )


def is_nonlife_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("손해|손보|화재|해상", regex=True, na=False)
        | is_special_nonlife_series(ins)
    )


def is_life_series(ins: pd.Series) -> pd.Series:
    ins = ins.astype(str).str.strip()

    return (
        ins.str.contains("생명", na=False)
        | ins.str.contains("라이프", na=False)
    )


def is_other_life_series(ins: pd.Series) -> pd.Series:
    return is_life_series(ins) & ~is_hanwha_life_series(ins)


# ── 데이터 준비 ──────────────────────────────────────────────
def load_df(uploaded_file) -> pd.DataFrame:
    df = pd.read_excel(uploaded_file)
    df = normalize_columns(df)
    df = standardize_columns(df)
    return df


def exclude_contracts(df: pd.DataFrame):
    """
    제외 조건:
    - 일시납
    - 연금성 / 저축성
    - 철회 / 해약 / 실효
    """
    excluded_df = pd.DataFrame()

    needed = {"납입방법", "상품군2", "계약상태"}

    if needed.issubset(df.columns):
        tmp = df.copy()

        tmp["납입방법"] = tmp["납입방법"].astype(str).str.strip()
        tmp["상품군2"] = tmp["상품군2"].astype(str).str.strip()
        tmp["계약상태"] = tmp["계약상태"].astype(str).str.strip()

        is_lumpsum = tmp["납입방법"].str.contains("일시납", na=False)
        is_savings = tmp["상품군2"].str.contains("연금성|저축성", regex=True, na=False)
        is_cancelled = tmp["계약상태"].str.contains("철회|해약|실효", regex=True, na=False)

        is_excluded = is_lumpsum | is_savings | is_cancelled

        excluded_df = tmp[is_excluded].copy()
        df_valid = tmp[~is_excluded].copy()

        return df_valid, excluded_df

    return df.copy(), excluded_df


def find_data_issues(df: pd.DataFrame, require_valid_date: bool = True):
    """환산 계산 보류 사유와 쉐어율 조건 확인 사유를 행별로 반환합니다."""
    blocking = pd.Series("", index=df.index, dtype="object")
    condition = pd.Series("", index=df.index, dtype="object")

    def add_issue(target, mask, message):
        mask = pd.Series(mask, index=df.index).fillna(False)
        target.loc[mask] = target.loc[mask].apply(
            lambda current: f"{current} / {message}" if current else message
        )

    def blank_mask(column):
        text = df[column].astype("string").str.strip().str.lower()
        return df[column].isna() | text.isin(["", "nan", "none", "<na>"])

    for column in ["수금자명", "보험사", "납입방법", "상품군2", "계약상태"]:
        add_issue(blocking, blank_mask(column), f"{column} 누락")

    period = pd.to_numeric(df["납입기간"], errors="coerce")
    add_issue(blocking, period.isna() | (period <= 0), "납입기간 확인 필요")

    premium = pd.to_numeric(df["보험료"], errors="coerce")
    add_issue(blocking, premium.isna() | (premium <= 0), "보험료 확인 필요")

    if require_valid_date:
        dates = pd.to_datetime(df["계약일자"], errors="coerce")
        add_issue(blocking, dates.isna(), "계약일자 확인 필요")

    share_text = (
        df["쉐어율"].astype("string").str.replace("%", "", regex=False).str.strip()
    )
    share_numeric = pd.to_numeric(share_text, errors="coerce")
    # 공란은 100% 단독계약으로 기본 적용하며, 화면에서 개별적으로 50%로 바꿀 수 있습니다.
    add_issue(
        condition,
        share_numeric.notna() & ((share_numeric <= 0) | (share_numeric > 100)),
        "쉐어율 확인 필요",
    )

    return blocking, condition, share_numeric


def build_review_display(review_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "원본행", "수금자명", "계약일자", "보험사", "상품명",
        "납입기간", "보험료", "쉐어율", "확인사항", "반영상태",
    ]
    if review_df is None or review_df.empty:
        return pd.DataFrame(columns=columns)

    out = review_df.copy()
    out.rename(columns={"_원본행번호": "원본행"}, inplace=True)
    for column in columns:
        if column not in out.columns:
            out[column] = ""
    return out[columns]


def build_excluded_with_reason(exdf: pd.DataFrame) -> pd.DataFrame:
    base_cols = [
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
        "납입방법",
        "계약상태",
        "제외사유",
    ]

    if exdf is None or exdf.empty:
        return pd.DataFrame(columns=base_cols)

    tmp = standardize_columns(exdf.copy())

    def reason_row(row):
        reasons = []

        if "일시납" in str(row.get("납입방법", "")):
            reasons.append("일시납")

        product_group = str(row.get("상품군2", ""))
        if "연금성" in product_group or "저축성" in product_group:
            reasons.append("연금/저축성")

        status = str(row.get("계약상태", ""))
        if "철회" in status:
            reasons.append("철회")
        if "해약" in status:
            reasons.append("해약")
        if "실효" in status:
            reasons.append("실효")

        return " / ".join(reasons) if reasons else "제외 조건 미상"

    tmp["제외사유"] = tmp.apply(reason_row, axis=1)

    for col in base_cols:
        if col not in tmp.columns:
            tmp[col] = ""

    tmp["계약일자"] = pd.to_datetime(tmp["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")

    tmp["납입기간"] = pd.to_numeric(tmp["납입기간"], errors="coerce").apply(
        lambda x: f"{int(x)}년" if pd.notnull(x) else ""
    )

    tmp["보험료"] = pd.to_numeric(tmp["보험료"], errors="coerce").apply(
        lambda x: won(x) if pd.notnull(x) else ""
    )

    return tmp[base_cols]


def check_required_columns(df: pd.DataFrame):
    required_columns = {
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "보험료",
        "쉐어율",
        "납입방법",
        "상품군2",
        "계약상태",
    }

    return required_columns - set(df.columns)


# ── 썸머 계산 ────────────────────────────────────────────────
def compute_summer(df: pd.DataFrame) -> pd.DataFrame:
    df = standardize_columns(df.copy())

    df["계약일자_raw"] = pd.to_datetime(df["계약일자"], errors="coerce")
    df["계약월"] = df["계약일자_raw"].dt.month

    df["보험료"] = pd.to_numeric(df["보험료"], errors="coerce").fillna(0)
    df["납입기간_num"] = pd.to_numeric(df["납입기간"], errors="coerce").fillna(0).astype(int)

    if "쉐어율" in df.columns:
        df["쉐어율"] = pd.to_numeric(
            df["쉐어율"]
            .astype("string")
            .str.replace("%", "", regex=False)
            .str.strip(),
            errors="coerce",
        )
    else:
        df["쉐어율"] = np.nan

    df["원본보험료"] = df["보험료"]

    # 공란은 기본 100%이며 화면에서 해당 행만 50%로 변경할 수 있습니다.
    df["쉐어율미입력"] = df["쉐어율"].isna()
    if "_공란적용쉐어율" not in df.columns:
        df["_공란적용쉐어율"] = 100.0
    df["_공란적용쉐어율"] = pd.to_numeric(
        df["_공란적용쉐어율"], errors="coerce"
    ).fillna(100.0)
    df["원본계산쉐어율"] = np.where(
        df["쉐어율미입력"], df["_공란적용쉐어율"], df["쉐어율"]
    )
    df["원본계산쉐어율"] = pd.to_numeric(df["원본계산쉐어율"], errors="coerce")

    valid_share = df["원본계산쉐어율"].between(1, 100, inclusive="both")
    is_shared = valid_share & (df["원본계산쉐어율"] < 100)
    df["적용쉐어율"] = np.where(is_shared, 50.0, np.where(valid_share, 100.0, df["원본계산쉐어율"]))
    df["쉐어건수"] = np.where(is_shared, 0.5, np.where(valid_share, 1.0, 0.0))

    # 원본 보험료는 원래 쉐어율이 이미 반영된 금액입니다.
    # 모든 공동계약을 50%로 통일하고 최종 원 미만 금액은 반올림 없이 버립니다.
    df["전체보험료역산"] = np.where(
        is_shared,
        df["원본보험료"] * 100 / df["원본계산쉐어율"],
        df["원본보험료"],
    )
    adjusted = np.where(
        is_shared,
        df["원본보험료"] * 50 / df["원본계산쉐어율"],
        df["원본보험료"],
    )
    df["실적보험료"] = np.floor(adjusted).astype(float)
    df["조정차액"] = df["실적보험료"] - df["원본보험료"]

    ins = df["보험사"].astype(str).str.strip()
    term = df["납입기간_num"]

    is_hanwha_life = is_hanwha_life_series(ins)
    is_special_nonlife = is_special_nonlife_series(ins)
    is_nonlife = is_nonlife_series(ins)
    is_other_nonlife = is_nonlife & ~is_special_nonlife
    is_other_life = is_other_life_series(ins)

    product_name = df.get("상품명", pd.Series("", index=df.index)).fillna("").astype(str)
    product_group = df.get("상품군2", pd.Series("", index=df.index)).fillna("").astype(str)
    df["치아보험자동판정"] = product_name.str.contains("치아", na=False) | product_group.str.contains("치아", na=False)
    if "_치아보험예외적용" not in df.columns:
        df["_치아보험예외적용"] = df["치아보험자동판정"]
    df["치아보험예외적용"] = df["_치아보험예외적용"].fillna(False).astype(bool)
    long_term_rule = (term > 10) | df["치아보험예외적용"]

    # ✅ 썸머 환산율
    # 손해보험
    # - 10년납 초과: 흥국/한화/KB/DB 250%, 이외 손해/화재 100%
    # - 10년납 이하: 흥국/한화/KB/DB 100%, 이외 손해/화재 50%
    #
    # 생명보험
    # - 10년납 초과: 한화생명 150%, 이외 생명보험 100%
    # - 10년납 이하: 한화생명 100%, 이외 생명보험 50%
    df["썸머율"] = np.select(
        [
            is_special_nonlife & long_term_rule,
            is_special_nonlife & ~long_term_rule,
            is_other_nonlife & long_term_rule,
            is_other_nonlife & ~long_term_rule,
            is_hanwha_life & long_term_rule,
            is_hanwha_life & ~long_term_rule,
            is_other_life & long_term_rule,
            is_other_life & ~long_term_rule,
        ],
        [
            250,
            100,
            100,
            50,
            150,
            100,
            100,
            50,
        ],
        default=0,
    ).astype(int)

    df["썸머환산금액"] = df["실적보험료"] * df["썸머율"] / 100

    def application_label(row):
        labels = []
        if pd.isna(row["쉐어율"]):
            labels.append(f"쉐어율 공란 → {row['적용쉐어율']:.0f}% {'기본' if row['적용쉐어율'] == 100 else '수동'} 적용")
        elif row["쉐어율"] < 100:
            if row["쉐어율"] == 50:
                labels.append("쉐어 50% 적용")
            else:
                labels.append(f"쉐어 조정 적용 {row['쉐어율']:g}% → 50%")
        if row["치아보험예외적용"]:
            labels.append("치아보험 예외 적용")
        return " · ".join(labels)

    df["적용 구분"] = df.apply(application_label, axis=1)

    return df


def check_monthly_requirements(dfin: pd.DataFrame):
    """
    월별 조건:
    1. 한화생명 월 환산업적 합계 5만원 이상
    2. 전체 월 환산업적 50만원 이상
    """
    if dfin.empty:
        return {
            "환산금액": 0,
            "한화생명5만": False,
            "환산50만": False,
            "월달성": False,
        }

    summer_sum = dfin["썸머환산금액"].sum()
    amount_ok = summer_sum >= MONTHLY_TARGET

    # 한화생명 계약의 썸머 환산업적을 월 단위로 합산하여
    # 합계가 5만원 이상인지 판정합니다.
    hanwha_mask = is_hanwha_life_series(dfin["보험사"])
    hanwha_summer_sum = pd.to_numeric(
        dfin.loc[hanwha_mask, "썸머환산금액"], errors="coerce"
    ).fillna(0).sum()
    hanwha_ok = hanwha_summer_sum >= MONTHLY_HANWHA_MIN_PREMIUM

    total_ok = amount_ok and hanwha_ok

    return {
        "환산금액": summer_sum,
        "한화생명5만": hanwha_ok,
        "환산50만": amount_ok,
        "월달성": total_ok,
    }


def get_summer_grade(total_amount: float):
    """
    7월 + 8월 합산 환산업적 기준 등급 산정.
    가장 높은 등급부터 체크.
    """
    for grade, target in SUMMER_GRADES:
        if total_amount >= target:
            return grade, target

    return "미달성", 0


def get_next_grade_gap(total_amount: float):
    ascending = [
        ("일반", 3_000_000),
        ("더블", 5_000_000),
        ("트리플", 8_000_000),
        ("크라운", 10_000_000),
        ("HWARANG", 15_000_000),
    ]

    for grade, target in ascending:
        if total_amount < target:
            return grade, target, target - total_amount

    return None, None, 0


def check_final_summer_requirements(
    july_df: pd.DataFrame,
    august_df: pd.DataFrame,
    ready_bonus_rate: float = 0,
):
    """
    1. 월별 필수조건은 보너스 전 금액 기준으로 판단
    2. 등급 판정은 레디포썸머 보너스 반영 후 금액 기준으로 판단
    """
    july_req = check_monthly_requirements(july_df)
    august_req = check_monthly_requirements(august_df)

    base_total_amount = july_req["환산금액"] + august_req["환산금액"]
    bonus_amount = base_total_amount * ready_bonus_rate / 100
    final_total_amount = base_total_amount + bonus_amount

    amount_grade, grade_target = get_summer_grade(final_total_amount)
    next_grade, next_target, next_gap = get_next_grade_gap(final_total_amount)

    monthly_all_ok = july_req["월달성"] and august_req["월달성"]

    # 금액 기준 등급과 최종 인정 등급을 분리
    if monthly_all_ok:
        final_grade = amount_grade
    else:
        final_grade = "필수조건 미충족"

    return {
        "7월": july_req,
        "8월": august_req,
        "기본합산환산금액": base_total_amount,
        "레디포썸머보너스율": ready_bonus_rate,
        "레디포썸머보너스금액": bonus_amount,
        "합산환산금액": final_total_amount,
        "월별필수조건": monthly_all_ok,
        "금액기준등급": amount_grade,
        "최종인정등급": final_grade,
        "달성기준금액": grade_target,
        "다음등급": next_grade,
        "다음등급기준": next_target,
        "다음등급부족금액": next_gap,
    }


# ── 화면 표시 ────────────────────────────────────────────────
def to_styled(dfin: pd.DataFrame) -> pd.DataFrame:
    df = dfin.copy()

    if df.empty:
        return pd.DataFrame(columns=[
            "계약월",
            "수금자명",
            "계약일자",
            "보험사",
            "상품명",
            "납입기간",
            "원본 보험료",
            "원본 쉐어율",
            "적용 쉐어율",
            "전체 보험료 역산",
            "실적보험료",
            "조정 차액",
            "인정 건수",
            "썸머율",
            "썸머환산금액",
            "적용 구분",
        ])

    df["계약일자"] = pd.to_datetime(df["계약일자"], errors="coerce").dt.strftime("%Y-%m-%d")

    df["납입기간"] = pd.to_numeric(df["납입기간"], errors="coerce").apply(
        lambda x: f"{int(x)}년" if pd.notnull(x) else ""
    )

    df["원본보험료"] = df["원본보험료"].map(won)
    df["쉐어율"] = df["쉐어율"].apply(lambda x: pct(x) if pd.notnull(x) else "공란")
    df["적용쉐어율"] = df["적용쉐어율"].apply(lambda x: pct(x) if pd.notnull(x) else "확인 필요")
    df["전체보험료역산"] = np.floor(pd.to_numeric(df["전체보험료역산"], errors="coerce")).map(won)
    df["실적보험료"] = df["실적보험료"].map(won)
    df["조정차액"] = df["조정차액"].map(signed_won)
    df["쉐어건수"] = df["쉐어건수"].apply(lambda x: f"{x:g}건")
    df["썸머율"] = df["썸머율"].map(pct)
    df["썸머환산금액"] = df["썸머환산금액"].map(won)

    df.rename(columns={
        "원본보험료": "원본 보험료",
        "쉐어율": "원본 쉐어율",
        "적용쉐어율": "적용 쉐어율",
        "전체보험료역산": "전체 보험료 역산",
        "조정차액": "조정 차액",
        "쉐어건수": "인정 건수",
    }, inplace=True)

    cols = [
        "계약월",
        "수금자명",
        "계약일자",
        "보험사",
        "상품명",
        "납입기간",
        "원본 보험료",
        "원본 쉐어율",
        "적용 쉐어율",
        "전체 보험료 역산",
        "실적보험료",
        "조정 차액",
        "인정 건수",
        "썸머율",
        "썸머환산금액",
        "적용 구분",
    ]

    return df[[c for c in cols if c in df.columns]]


def style_detail_table(dfin: pd.DataFrame):
    display = to_styled(dfin)
    highlight_cols = {
        "원본 쉐어율", "적용 쉐어율", "전체 보험료 역산", "실적보험료",
        "조정 차액", "인정 건수", "썸머율", "적용 구분",
    }

    def color_row(row):
        label = str(row.get("적용 구분", ""))
        has_share = "쉐어" in label
        has_dental = "치아보험" in label
        color = ""
        if has_share and has_dental:
            color = "background-color: #eee3ff"
        elif has_dental:
            color = "background-color: #e3f2fd"
        elif has_share:
            color = "background-color: #fff4cc"
        return [color if col in highlight_cols else "" for col in row.index]

    try:
        return display.style.apply(color_row, axis=1)
    except ImportError:
        # 배포 환경에 스타일 선택 의존성이 없더라도 계산과 표 표시는 유지합니다.
        return display


def adjustment_summary(dfin: pd.DataFrame) -> dict:
    if dfin is None or dfin.empty:
        return {"원본": 0, "조정": 0, "차액": 0, "증가": 0, "감소": 0, "쉐어건수": 0}
    diff = pd.to_numeric(dfin["조정차액"], errors="coerce").fillna(0)
    return {
        "원본": dfin["원본보험료"].sum(),
        "조정": dfin["실적보험료"].sum(),
        "차액": diff.sum(),
        "증가": diff[diff > 0].sum(),
        "감소": diff[diff < 0].sum(),
        "쉐어건수": int((dfin["적용쉐어율"] < 100).sum()),
    }


def render_adjustment_summary(dfin: pd.DataFrame, title="쉐어 조정 요약"):
    summary = adjustment_summary(dfin)
    st.markdown(f"#### {title}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("원본 보험료 합계", won(summary["원본"]))
    c2.metric("50% 조정 보험료 합계", won(summary["조정"]), signed_won(summary["차액"]))
    c3.metric("증가액 / 감소액", f"+{summary['증가']:,.0f} / {summary['감소']:,.0f} 원")
    c4.metric("쉐어 적용 계약", f"{summary['쉐어건수']:,}건")


def money_box(title, value, color="#ff9800"):
    return f"""
    <div style='border: 2px solid {color}; border-radius: 10px; padding: 18px; background-color: #fff8e1; margin-bottom: 12px;'>
        <h4 style='color:{color}; margin:0 0 8px 0;'>{title}</h4>
        <p style='font-size:20px; font-weight:bold; margin:0;'>{value:,.0f} 원</p>
    </div>
    """


def bonus_box(base_amount, bonus_rate, bonus_amount, final_amount):
    return f"""
    <div style='border: 2px solid #6f42c1; border-radius: 10px; padding: 18px; background-color: #f3ecff; margin-bottom: 12px;'>
        <h4 style='color:#6f42c1; margin:0 0 10px 0;'>🎁 레디포썸머 보너스 반영</h4>
        <p style='margin:4px 0;'><strong>기본 썸머 환산업적:</strong> {base_amount:,.0f} 원</p>
        <p style='margin:4px 0;'><strong>보너스율:</strong> {bonus_rate:.0f} %</p>
        <p style='margin:4px 0;'><strong>보너스 가산금액:</strong> {bonus_amount:,.0f} 원</p>
        <p style='font-size:20px; font-weight:bold; margin:10px 0 0 0; color:#6f42c1;'>
            보너스 반영 최종 환산업적: {final_amount:,.0f} 원
        </p>
    </div>
    """


def grade_box(final_grade, amount_grade, base_amount, bonus_rate, bonus_amount, final_amount, monthly_ok):
    if final_grade == "필수조건 미충족":
        color = "#b80000"
        bg = "#fdecea"
    elif final_grade == "미달성":
        color = "#b80000"
        bg = "#fdecea"
    elif final_grade in ["일반", "더블"]:
        color = "#0c6b2c"
        bg = "#e6f4ea"
    elif final_grade in ["트리플", "크라운"]:
        color = "#7a4b00"
        bg = "#fff4d6"
    else:
        color = "#4b0082"
        bg = "#f0e6ff"

    monthly_text = "월별 필수조건 충족" if monthly_ok else "월별 필수조건 미충족"

    return f"""
    <div style='border: 2px solid {color}; border-radius: 12px; padding: 20px; background-color: {bg}; margin-bottom: 16px;'>
        <h3 style='color:{color}; margin:0 0 10px 0;'>최종 인정 등급: {final_grade}</h3>
        <p style='margin:4px 0;'><strong>기본 썸머 환산업적:</strong> {base_amount:,.0f} 원</p>
        <p style='margin:4px 0;'><strong>레디포썸머 보너스율:</strong> {bonus_rate:.0f} %</p>
        <p style='margin:4px 0;'><strong>레디포썸머 보너스금액:</strong> {bonus_amount:,.0f} 원</p>
        <p style='font-size:20px; font-weight:bold; margin:8px 0;'>보너스 반영 최종 환산업적: {final_amount:,.0f} 원</p>
        <p style='font-weight:bold; margin:4px 0;'>금액 기준 등급: {amount_grade}</p>
        <p style='font-weight:bold; margin:4px 0;'>월별 필수조건: {monthly_text}</p>
    </div>
    """


def gap_box(title, amount):
    if amount > 0:
        color = "#e6f4ea"
        txt = "#0c6b2c"
        sym = f"+{amount:,.0f} 원 초과"
    elif amount < 0:
        color = "#fdecea"
        txt = "#b80000"
        sym = f"{amount:,.0f} 원 부족"
    else:
        color = "#f3f3f3"
        txt = "#000000"
        sym = "기준 달성"

    return f"""
    <div style='border: 1px solid {txt}; border-radius: 8px; background-color: {color}; padding: 12px; margin: 10px 0;'>
        <strong style='color:{txt};'>{title}: {sym}</strong>
    </div>
    """


def req_box(title, ok):
    color = "#e6f4ea" if ok else "#fdecea"
    txt = "#0c6b2c" if ok else "#b80000"
    mark_txt = "✅ 충족" if ok else "❌ 미충족"

    return f"""
    <div style='border: 1px solid {txt}; border-radius: 8px; background-color: {color}; padding: 12px; margin: 10px 0;'>
        <strong style='color:{txt};'>{title}: {mark_txt}</strong>
    </div>
    """


def make_collector_summary(july_df: pd.DataFrame, august_df: pd.DataFrame) -> pd.DataFrame:
    """
    수금자별 요약은 기본 환산업적 기준으로 표시.
    레디포썸머 보너스는 선택 수금자 화면에서 직접 선택 후 별도 반영.
    """
    all_df = pd.concat([july_df, august_df], ignore_index=True)

    rows = []

    if all_df.empty:
        return pd.DataFrame(columns=[
            "수금자명",
            "7월건수",
            "7월쉐어미입력",
            "7월환산",
            "7월한화5만",
            "7월50만",
            "7월달성",
            "8월건수",
            "8월쉐어미입력",
            "8월환산",
            "8월한화5만",
            "8월50만",
            "8월달성",
            "기본합산환산",
            "월별필수조건",
            "기본금액등급",
        ])

    for collector, sub in all_df.groupby("수금자명", dropna=False):
        july_sub = sub[sub["계약월"] == 7].copy()
        august_sub = sub[sub["계약월"] == 8].copy()

        result = check_final_summer_requirements(
            july_sub,
            august_sub,
            ready_bonus_rate=0,
        )

        rows.append({
            "수금자명": collector,
            "7월건수": july_sub["쉐어건수"].sum(min_count=1),
            "7월쉐어미입력": int(july_sub["쉐어율"].isna().sum()),
            "7월환산": result["7월"]["환산금액"],
            "7월한화5만": mark(result["7월"]["한화생명5만"]),
            "7월50만": mark(result["7월"]["환산50만"]),
            "7월달성": mark(result["7월"]["월달성"]),
            "8월건수": august_sub["쉐어건수"].sum(min_count=1),
            "8월쉐어미입력": int(august_sub["쉐어율"].isna().sum()),
            "8월환산": result["8월"]["환산금액"],
            "8월한화5만": mark(result["8월"]["한화생명5만"]),
            "8월50만": mark(result["8월"]["환산50만"]),
            "8월달성": mark(result["8월"]["월달성"]),
            "기본합산환산": result["기본합산환산금액"],
            "월별필수조건": mark(result["월별필수조건"]),
            "기본금액등급": result["금액기준등급"],
        })

    summary = pd.DataFrame(rows)
    return summary


def format_summary_for_display(summary: pd.DataFrame) -> pd.DataFrame:
    df = summary.copy()

    for month in ["7월", "8월"]:
        count_col = f"{month}건수"
        missing_col = f"{month}쉐어미입력"

        if count_col in df.columns:
            def format_count(row):
                value = row.get(count_col)
                missing = int(row.get(missing_col, 0) or 0)
                base = (
                    f"{float(value):.2f}".rstrip("0").rstrip(".")
                    if pd.notnull(value)
                    else "0"
                )
                return f"{base}건 (공란 기본 100% {missing}건)" if missing else f"{base}건"

            df[count_col] = df.apply(format_count, axis=1)

        if missing_col in df.columns:
            df.drop(columns=[missing_col], inplace=True)

    for col in ["7월환산", "8월환산", "기본합산환산"]:
        if col in df.columns:
            df[col] = df[col].map(won)

    return df


# ── 선택 수금자 필터 ─────────────────────────────────────────
def filter_by_collector(df: pd.DataFrame, selected_collector: str) -> pd.DataFrame:
    if selected_collector == "전체":
        return df.copy()

    return df[df["수금자명"].astype(str) == selected_collector].copy()


def filter_excluded_by_collector(excluded_disp: pd.DataFrame, selected_collector: str) -> pd.DataFrame:
    """
    엑셀 다운로드 시 제외계약도 선택한 수금자 기준으로 필터링합니다.
    """
    if excluded_disp is None or excluded_disp.empty:
        return pd.DataFrame()

    if selected_collector == "전체":
        return excluded_disp.copy()

    if "수금자명" not in excluded_disp.columns:
        return pd.DataFrame()

    return excluded_disp[
        excluded_disp["수금자명"].astype(str) == str(selected_collector)
    ].copy()


# ── 엑셀 출력 ────────────────────────────────────────────────
def write_table(ws, df_for_sheet: pd.DataFrame, start_row: int = 1, name_suffix: str = "A"):
    global TABLE_SEQ

    r_idx = start_row

    for r_idx, row in enumerate(dataframe_to_rows(df_for_sheet, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    end_col_letter = ws.cell(row=start_row, column=max(df_for_sheet.shape[1], 1)).column_letter
    last_row = r_idx if df_for_sheet.shape[0] > 0 else start_row

    TABLE_SEQ += 1
    display_name = safe_table_name(f"tbl_{ws.title}_{name_suffix}_{TABLE_SEQ}")

    table = Table(displayName=display_name, ref=f"A{start_row}:{end_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    # 화면과 동일하게 조정·예외 적용 셀을 강조합니다.
    if "적용 구분" in df_for_sheet.columns:
        headers = {cell.value: cell.column for cell in ws[start_row]}
        target_headers = [
            "원본 쉐어율", "적용 쉐어율", "전체 보험료 역산", "실적보험료",
            "조정 차액", "인정 건수", "썸머율", "적용 구분",
        ]
        for row_num in range(start_row + 1, last_row + 1):
            label = str(ws.cell(row=row_num, column=headers["적용 구분"]).value or "")
            has_share = "쉐어" in label
            has_dental = "치아보험" in label
            fill_color = None
            if has_share and has_dental:
                fill_color = "EEE3FF"
            elif has_dental:
                fill_color = "E3F2FD"
            elif has_share:
                fill_color = "FFF4CC"
            if fill_color:
                for header in target_headers:
                    if header in headers:
                        ws.cell(row=row_num, column=headers[header]).fill = PatternFill("solid", fgColor=fill_color)

    autosize_columns_full(ws, padding=5)

    return last_row


def write_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_final_result_block(ws, row, result):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    fill = PatternFill("solid", fgColor="F2F2F2")

    rows = [
        ["7월 환산업적", won(result["7월"]["환산금액"])],
        ["7월 한화생명 환산업적 합계 5만원 이상", mark(result["7월"]["한화생명5만"])],
        ["7월 환산업적 50만원 이상", mark(result["7월"]["환산50만"])],
        ["7월 조건 달성", mark(result["7월"]["월달성"])],
        ["8월 환산업적", won(result["8월"]["환산금액"])],
        ["8월 한화생명 환산업적 합계 5만원 이상", mark(result["8월"]["한화생명5만"])],
        ["8월 환산업적 50만원 이상", mark(result["8월"]["환산50만"])],
        ["8월 조건 달성", mark(result["8월"]["월달성"])],
        ["기본 7월+8월 합산 환산업적", won(result["기본합산환산금액"])],
        ["레디포썸머 보너스율", f"{result['레디포썸머보너스율']:.0f} %"],
        ["레디포썸머 보너스금액", won(result["레디포썸머보너스금액"])],
        ["보너스 반영 최종 환산업적", won(result["합산환산금액"])],
        ["월별 필수조건", mark(result["월별필수조건"])],
        ["금액 기준 등급", result["금액기준등급"]],
        ["최종 인정 등급", result["최종인정등급"]],
    ]

    if result["다음등급"]:
        rows.append([
            f"다음 등급({result['다음등급']})까지 부족금액",
            won(result["다음등급부족금액"]),
        ])
    else:
        rows.append(["최고 등급 달성", "HWARANG"])

    for i, row_data in enumerate(rows, start=row):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            if j == 1:
                cell.fill = fill
                cell.font = Font(bold=True)

    autosize_columns_full(ws, padding=5)

    return row + len(rows)


def build_workbook(
    df_all: pd.DataFrame,
    july_df: pd.DataFrame,
    august_df: pd.DataFrame,
    other_month_df: pd.DataFrame,
    summary: pd.DataFrame,
    result: dict,
    excluded_disp: pd.DataFrame,
    selected_collector: str = "전체",
    review_disp: pd.DataFrame | None = None,
):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "요약"

    write_title(ws_summary, 1, f"썸머 최종 결과 - {selected_collector}")
    next_row = write_final_result_block(ws_summary, 2, result)

    write_title(ws_summary, next_row + 2, "수금자별 요약")
    next_row = write_table(
        ws_summary,
        format_summary_for_display(summary),
        start_row=next_row + 3,
        name_suffix="SUMMARY",
    )

    write_title(ws_summary, next_row + 2, "상세 내역")
    next_row = write_table(
        ws_summary,
        to_styled(df_all),
        start_row=next_row + 3,
        name_suffix="DETAIL",
    )

    ws_july = wb.create_sheet("7월")
    write_title(ws_july, 1, f"7월 썸머 환산 결과 - {selected_collector}")
    write_table(ws_july, to_styled(july_df), start_row=2, name_suffix="JULY_DETAIL")

    ws_august = wb.create_sheet("8월")
    write_title(ws_august, 1, f"8월 썸머 환산 결과 - {selected_collector}")
    write_table(ws_august, to_styled(august_df), start_row=2, name_suffix="AUGUST_DETAIL")

    if not other_month_df.empty:
        ws_other = wb.create_sheet("7월8월외")
        write_title(ws_other, 1, f"7월/8월 외 계약 - {selected_collector}")
        write_table(ws_other, to_styled(other_month_df), start_row=2, name_suffix="OTHER_MONTH")

    if excluded_disp is not None and not excluded_disp.empty:
        ws_ex = wb.create_sheet("제외계약")
        write_title(ws_ex, 1, f"제외 계약 - {selected_collector}")
        write_table(ws_ex, excluded_disp, start_row=2, name_suffix="EXCLUDED")

    if review_disp is not None and not review_disp.empty:
        ws_review = wb.create_sheet("확인필요계약")
        write_title(ws_review, 1, f"입력값 확인이 필요한 계약 - {selected_collector}")
        write_table(ws_review, review_disp, start_row=2, name_suffix="REVIEW")

    return wb


# ── 탭 렌더링 함수 ───────────────────────────────────────────
def render_result_tabs(summary_df, july_df, august_df, other_month_df):
    tab1, tab2, tab3, tab4 = st.tabs(["🧮 수금자별 요약", "7월 상세", "8월 상세", "7월/8월 외"])

    with tab1:
        st.dataframe(format_summary_for_display(summary_df), use_container_width=True)

    with tab2:
        if july_df.empty:
            st.info("7월 계약이 없습니다.")
        else:
            st.dataframe(style_detail_table(july_df), use_container_width=True)

    with tab3:
        if august_df.empty:
            st.info("8월 계약이 없습니다.")
        else:
            st.dataframe(style_detail_table(august_df), use_container_width=True)

    with tab4:
        if other_month_df.empty:
            st.info("7월/8월 외 계약이 없습니다.")
        else:
            st.dataframe(style_detail_table(other_month_df), use_container_width=True)


# ── 메인 실행 ────────────────────────────────────────────────
def run():
    page_header("실적 관리", "썸머 계산기", "계약일 기준으로 7월과 8월을 분리해 썸머 업적과 최종 등급을 계산합니다.", "SU")

    with st.expander("월별 필수조건·보너스·등급·환산율 보기"):
        st.header("🧭 사용 방법")
        st.markdown(
            """
            **🖥️ 한화라이프랩 전산**  
            **- 📂 계약관리**  
            **- 📑 보유계약 장기**  
            **- ⏱️ 기간 설정**  
            **- 💾 엑셀 다운로드 후 파일 첨부**
            """
        )

        st.divider()

        st.subheader("🌞 월별 필수조건")
        st.markdown(
            f"""
            - 7월: 한화생명 환산업적 합계 **{MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상**
            - 7월: 환산업적 **{MONTHLY_TARGET:,.0f}원 이상**
            - 8월: 한화생명 환산업적 합계 **{MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상**
            - 8월: 환산업적 **{MONTHLY_TARGET:,.0f}원 이상**
            """
        )

        st.subheader("🎁 레디포썸머 보너스")
        st.markdown(
            """
            - 수금자 선택 후 보너스율 직접 선택
            - 선택 가능: 0%, 15%, 20%, 25%, 30%
            - 등급 판정은 보너스 반영 후 금액 기준
            - 월별 필수조건은 보너스 전 기준으로 판단
            """
        )

        st.subheader("🏆 최종 합산 등급")
        st.markdown(
            """
            - 일반: 300만원 이상
            - 더블: 500만원 이상
            - 트리플: 800만원 이상
            - 크라운: 1,000만원 이상
            - HWARANG: 1,500만원 이상
            """
        )

        st.subheader("📌 환산율")
        st.markdown(
            """
            **손해보험**
            - 10년납 초과: 흥국/한화/KB/DB 250%
            - 10년납 초과: 이외 손해/화재 100%
            - 10년납 이하: 흥국/한화/KB/DB 100%
            - 10년납 이하: 이외 손해/화재 50%

            **생명보험**
            - 10년납 초과: 한화생명 150%
            - 10년납 초과: 이외 생명보험 100%
            - 10년납 이하: 한화생명 100%
            - 10년납 이하: 이외 생명보험 50%

            **예외 및 쉐어 기준**
            - 상품명 또는 상품군에 `치아`가 포함되면 보험사와 관계없이 10년납 초과 구간 적용
            - 공동계약은 원래 쉐어율과 관계없이 50% 보험료·0.5건으로 통일
            - 조정 실적보험료의 원 미만 금액은 반올림하지 않고 버림
            """
        )

        st.markdown(
            """
            **🚫 제외 기준**  
            - 일시납  
            - 연금성 / 저축성  
            - 철회 / 해약 / 실효
            """
        )

    section_intro("입력", "계약자료 불러오기", "7월과 8월 계약이 포함된 보유계약 엑셀 파일을 등록해 주세요.")
    uploaded_file = st.file_uploader(
        "📂 썸머 계산용 Excel 파일 업로드 (.xlsx)",
        type=["xlsx"],
        key="summer_one_file",
    )

    if uploaded_file is None:
        st.info("📤 7월과 8월 계약이 포함된 Excel 파일(.xlsx)을 업로드해주세요.")
        return

    base_filename = os.path.splitext(uploaded_file.name)[0]

    file_bytes = uploaded_file.getvalue()

    try:
        raw = load_df(BytesIO(file_bytes)).copy()
    except Exception as e:
        st.error(f"❌ 엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")
        return

    missing = check_required_columns(raw)

    if missing:
        st.error(
            "❌ 업로드된 파일에 다음 항목이 필요합니다:\n\n"
            + ", ".join(sorted(missing))
        )
        return

    raw["_원본행번호"] = raw.index + 2
    candidate_df, excluded_df = exclude_contracts(raw)
    blocking_issues, condition_issues, _ = find_data_issues(candidate_df)
    initial_review_mask = blocking_issues.ne("") | condition_issues.ne("")
    initial_review = candidate_df[initial_review_mask].copy()

    if not initial_review.empty:
        initial_review["확인사항"] = blocking_issues.loc[initial_review.index]
        condition_only = condition_issues.loc[initial_review.index]
        initial_review["확인사항"] = initial_review.apply(
            lambda row: " / ".join(
                part for part in [row["확인사항"], condition_only.loc[row.name]] if part
            ),
            axis=1,
        )
        editor_columns = [
            "_원본행번호", "수금자명", "계약일자", "보험사", "상품명",
            "납입기간", "보험료", "쉐어율", "납입방법", "상품군2",
            "계약상태", "확인사항",
        ]

        st.warning(
            f"입력값을 확인해야 하는 계약이 {len(initial_review):,}건 있습니다. "
            "표에서 직접 수정한 뒤 적용할 수 있습니다."
        )
        with st.expander("📝 확인 필요 계약 수정", expanded=True):
            with st.form("summer_review_editor_form"):
                edited_review = st.data_editor(
                    initial_review[editor_columns].reset_index(drop=True),
                    use_container_width=True,
                    hide_index=True,
                    disabled=["_원본행번호", "상품명", "확인사항"],
                    key=f"summer_review_{hashlib.sha256(file_bytes).hexdigest()[:16]}",
                )
                corrections_submitted = st.form_submit_button(
                    "수정값 적용", type="primary", use_container_width=True
                )

        editable_columns = [
            "수금자명", "계약일자", "보험사", "납입기간", "보험료",
            "쉐어율", "납입방법", "상품군2", "계약상태",
        ]
        for _, edited_row in edited_review.iterrows():
            row_mask = candidate_df["_원본행번호"] == edited_row["_원본행번호"]
            for column in editable_columns:
                candidate_df.loc[row_mask, column] = edited_row[column]

        if corrections_submitted:
            st.success("입력한 수정값을 다시 검증하여 반영했습니다.")

    candidate_df, newly_excluded_df = exclude_contracts(candidate_df)
    if not newly_excluded_df.empty:
        excluded_df = pd.concat([excluded_df, newly_excluded_df]).sort_index()

    blocking_issues, condition_issues, share_numeric = find_data_issues(candidate_df)
    blocked_df = candidate_df[blocking_issues.ne("")].copy()
    if not blocked_df.empty:
        blocked_df["확인사항"] = blocking_issues.loc[blocked_df.index]
        blocked_df["반영상태"] = "계산 보류"

    condition_mask = blocking_issues.eq("") & condition_issues.ne("")
    condition_df = candidate_df[condition_mask].copy()
    if not condition_df.empty:
        condition_df["확인사항"] = condition_issues.loc[condition_df.index]
        condition_df["반영상태"] = "금액 반영 · 인정 건수 보류"

    review_df = pd.concat([blocked_df, condition_df]).sort_index()
    review_disp_all = build_review_display(review_df)
    df_valid = candidate_df[blocking_issues.eq("")].copy()
    df_valid.loc[:, "쉐어율"] = share_numeric.loc[df_valid.index]
    excluded_disp = build_excluded_with_reason(excluded_df)

    upload_key = hashlib.sha256(file_bytes).hexdigest()[:16]

    # 쉐어율 공란은 기본 100%로 적용하되 행별로 50%를 선택할 수 있습니다.
    df_valid["_공란적용쉐어율"] = 100.0
    blank_share_mask = df_valid["쉐어율"].isna()
    blank_share_df = df_valid[blank_share_mask].copy()
    if not blank_share_df.empty:
        st.markdown(f"#### 쉐어율 공란 확인 대상 {len(blank_share_df):,}건")
        st.caption("기본값은 100% 단독계약입니다. 필요한 계약만 50% 쉐어계약으로 변경해 주세요.")
        blank_editor = blank_share_df[
            ["_원본행번호", "수금자명", "보험사", "상품명", "보험료"]
        ].copy()
        blank_editor["적용 쉐어율"] = "100%"
        blank_editor = st.data_editor(
            blank_editor.reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
            disabled=["_원본행번호", "수금자명", "보험사", "상품명", "보험료"],
            column_config={
                "적용 쉐어율": st.column_config.SelectboxColumn(
                    "적용 쉐어율", options=["100%", "50%"], required=True
                )
            },
            key=f"summer_blank_share_{upload_key}",
        )
        for _, edited_row in blank_editor.iterrows():
            row_mask = df_valid["_원본행번호"] == edited_row["_원본행번호"]
            df_valid.loc[row_mask, "_공란적용쉐어율"] = 50.0 if edited_row["적용 쉐어율"] == "50%" else 100.0

    # 상품명 또는 상품군2에 '치아'가 있으면 기본 체크하고, 해제 시 즉시 일반 납기 기준으로 계산합니다.
    product_name = df_valid["상품명"].fillna("").astype(str)
    product_group = df_valid["상품군2"].fillna("").astype(str)
    dental_mask = product_name.str.contains("치아", na=False) | product_group.str.contains("치아", na=False)
    df_valid["_치아보험예외적용"] = False
    dental_df = df_valid[dental_mask].copy()
    st.markdown(f"#### 치아보험 확인 대상 {len(dental_df):,}건")
    if dental_df.empty:
        st.info("상품명 또는 상품군에 '치아'가 포함된 계약이 없습니다.")
    else:
        st.caption("체크된 계약은 실제 납입기간과 관계없이 10년납 초과 환산율을 적용합니다.")
        dental_editor = dental_df[
            ["_원본행번호", "수금자명", "보험사", "상품명", "상품군2", "납입기간"]
        ].copy()
        dental_editor["치아보험 예외 적용"] = True
        dental_editor = st.data_editor(
            dental_editor.reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
            disabled=["_원본행번호", "수금자명", "보험사", "상품명", "상품군2", "납입기간"],
            column_config={
                "치아보험 예외 적용": st.column_config.CheckboxColumn(
                    "치아보험 예외 적용", default=True
                )
            },
            key=f"summer_dental_check_{upload_key}",
        )
        for _, edited_row in dental_editor.iterrows():
            row_mask = df_valid["_원본행번호"] == edited_row["_원본행번호"]
            df_valid.loc[row_mask, "_치아보험예외적용"] = bool(edited_row["치아보험 예외 적용"])

    if not blocked_df.empty:
        st.warning(
            f"중요 항목을 확인할 수 없는 계약 {len(blocked_df):,}건은 계산에서 제외했습니다."
        )
    if not condition_df.empty:
        st.info(
            f"범위를 벗어난 쉐어율 계약 {len(condition_df):,}건은 환산금액에는 포함하고 "
            "인정 건수에서는 제외했습니다."
        )
    if not review_disp_all.empty:
        with st.expander("⚠️ 아직 확인이 필요한 계약", expanded=False):
            st.dataframe(review_disp_all, use_container_width=True, hide_index=True)

    if df_valid.empty:
        st.warning("계산에 포함할 수 있는 정상 계약이 없습니다. 확인 필요 계약을 수정해 주세요.")
        return

    df = compute_summer(df_valid)

    july_df = df[df["계약월"] == 7].copy()
    august_df = df[df["계약월"] == 8].copy()
    other_month_df = df[~df["계약월"].isin([7, 8])].copy()

    if july_df.empty:
        st.warning("⚠️ 계약일 기준 7월 계약이 없습니다.")

    if august_df.empty:
        st.warning("⚠️ 계약일 기준 8월 계약이 없습니다.")

    if not other_month_df.empty:
        st.info(
            f"ℹ️ 7월/8월 외 계약 {len(other_month_df)}건이 있습니다. "
            "이 계약들은 썸머 최종 조건 계산에서는 제외하고, 엑셀에는 별도 시트로 저장합니다."
        )

    # 전체 기준 결과: 보너스율 0% 기준
    total_result = check_final_summer_requirements(
        july_df,
        august_df,
        ready_bonus_rate=0,
    )
    total_summary = make_collector_summary(july_df, august_df)

    # 1. 제외 계약 보기 - 기본 펼침
    if excluded_disp is not None and not excluded_disp.empty:
        st.warning(
            f"⚠️ 제외된 계약 {len(excluded_disp)}건이 있습니다. "
            "제외 조건: 일시납 / 연금성·저축성 / 철회·해약·실효"
        )

        with st.expander("🚫 제외된 계약 보기", expanded=True):
            st.dataframe(excluded_disp, use_container_width=True)
    else:
        with st.expander("🚫 제외된 계약 보기", expanded=True):
            st.info("제외된 계약이 없습니다.")

    # 2. 전체 환산 결과
    section_intro("전체 결과", "썸머 환산 결과", "반영 계약과 제외 계약을 포함한 전체 계산 결과입니다.")
    render_adjustment_summary(df, "전체 쉐어 조정 요약")
    render_result_tabs(
        summary_df=total_summary,
        july_df=july_df,
        august_df=august_df,
        other_month_df=other_month_df,
    )

    # 3. 수금자별 결과 확인
    section_intro("상세 결과", "수금자별 결과 확인", "수금자를 선택해 월별 실적과 보너스 적용 결과를 확인해 주세요.")

    collectors = ["전체"] + sorted(df["수금자명"].astype(str).dropna().unique().tolist())

    selected_collector = st.selectbox(
        "👤 확인할 수금자를 선택하세요.",
        collectors,
        index=0,
        key="summer_selected_collector",
    )

    ready_bonus_rate = st.selectbox(
        "🎁 레디포썸머 보너스율을 선택하세요.",
        READY_BONUS_RATES,
        index=0,
        format_func=lambda x: f"{x}%",
        key="summer_ready_bonus_rate",
    )

    selected_df = filter_by_collector(df, selected_collector)
    selected_july_df = filter_by_collector(july_df, selected_collector)
    selected_august_df = filter_by_collector(august_df, selected_collector)
    selected_other_month_df = filter_by_collector(other_month_df, selected_collector)

    selected_summary = make_collector_summary(selected_july_df, selected_august_df)

    selected_result = check_final_summer_requirements(
        selected_july_df,
        selected_august_df,
        ready_bonus_rate=ready_bonus_rate,
    )

    selected_excluded_disp = filter_excluded_by_collector(excluded_disp, selected_collector)
    selected_review_disp = filter_excluded_by_collector(review_disp_all, selected_collector)

    st.markdown(f"### 📌 선택 기준: {selected_collector}")
    st.caption(f"레디포썸머 보너스율: {ready_bonus_rate}%")

    render_adjustment_summary(selected_df, f"{selected_collector} 쉐어 조정 요약")

    render_result_tabs(
        summary_df=selected_summary,
        july_df=selected_july_df,
        august_df=selected_august_df,
        other_month_df=selected_other_month_df,
    )

    # 4. 선택값 기준 월별 필수조건 체크
    st.subheader("✅ 월별 필수조건 체크")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 7월")
        st.markdown(
            money_box("7월 환산업적", selected_result["7월"]["환산금액"]),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"7월 한화생명 환산업적 합계 {MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상",
                selected_result["7월"]["한화생명5만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"7월 환산업적 {MONTHLY_TARGET:,.0f}원 이상",
                selected_result["7월"]["환산50만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box("7월 필수조건 전체", selected_result["7월"]["월달성"]),
            unsafe_allow_html=True,
        )

    with col2:
        st.markdown("### 8월")
        st.markdown(
            money_box("8월 환산업적", selected_result["8월"]["환산금액"]),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"8월 한화생명 환산업적 합계 {MONTHLY_HANWHA_MIN_PREMIUM:,.0f}원 이상",
                selected_result["8월"]["한화생명5만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box(
                f"8월 환산업적 {MONTHLY_TARGET:,.0f}원 이상",
                selected_result["8월"]["환산50만"],
            ),
            unsafe_allow_html=True,
        )
        st.markdown(
            req_box("8월 필수조건 전체", selected_result["8월"]["월달성"]),
            unsafe_allow_html=True,
        )

    st.markdown(
        req_box("7월·8월 월별 필수조건 전체", selected_result["월별필수조건"]),
        unsafe_allow_html=True,
    )

    # 5. 레디포썸머 보너스 반영 결과
    st.subheader("🎁 레디포썸머 보너스 반영 결과")

    st.markdown(
        bonus_box(
            selected_result["기본합산환산금액"],
            selected_result["레디포썸머보너스율"],
            selected_result["레디포썸머보너스금액"],
            selected_result["합산환산금액"],
        ),
        unsafe_allow_html=True,
    )

    # 6. 선택값 기준 썸머 최종 결과
    section_intro("최종 결과", "썸머 최종 등급", "월별 필수조건과 보너스를 모두 반영한 최종 결과입니다.")

    st.markdown(
        grade_box(
            selected_result["최종인정등급"],
            selected_result["금액기준등급"],
            selected_result["기본합산환산금액"],
            selected_result["레디포썸머보너스율"],
            selected_result["레디포썸머보너스금액"],
            selected_result["합산환산금액"],
            selected_result["월별필수조건"],
        ),
        unsafe_allow_html=True,
    )

    if selected_result["다음등급"]:
        st.markdown(
            gap_box(
                f"다음 등급 {selected_result['다음등급']}({selected_result['다음등급기준']:,.0f}원)까지",
                -selected_result["다음등급부족금액"],
            ),
            unsafe_allow_html=True,
        )
    else:
        st.success("🎉 최고 등급 HWARANG 기준을 달성했습니다.")

    # 7. 엑셀 다운로드
    # 선택 기준에 따라 다운로드 데이터 분기
    # - 전체 선택: 전체 다운로드
    # - 수금자 선택: 해당 수금자만 다운로드
    # - 제외계약도 선택 기준에 맞게 필터링
    # - 선택한 보너스율이 엑셀 결과에 반영
    file_collector_name = safe_filename_part(selected_collector)
    download_filename = f"{base_filename}_썸머환산결과_{file_collector_name}.xlsx"

    wb = build_workbook(
        df_all=selected_df,
        july_df=selected_july_df,
        august_df=selected_august_df,
        other_month_df=selected_other_month_df,
        summary=selected_summary,
        result=selected_result,
        excluded_disp=selected_excluded_disp,
        review_disp=selected_review_disp,
        selected_collector=selected_collector,
    )

    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)

    st.download_button(
        label=f"📥 {selected_collector} 썸머 환산 결과 엑셀 다운로드",
        data=excel_output,
        file_name=download_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
