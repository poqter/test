from __future__ import annotations

# 전달용 파일: 월별 수수료표 스마트 연결 흐름 적용본 v14

import hashlib
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from typing import Any

import streamlit as st
from .ui_components import page_header, section_intro
import streamlit.components.v1 as components
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_PAYOUT_RATE = 65.0
FIRST_YEAR_HEADERS = {"1차년계", "1차년도합계", "1차년합계"}
TOTAL_HEADERS = {"총수수료", "총계", "총합계", "총수수료계"}
PRODUCT_HEADERS = {"상품명"}
PRODUCT_FALLBACK_HEADERS = {"구분"}


@dataclass(frozen=True)
class ProductRate:
    key: str
    source_type: str
    insurer: str
    product: str
    conditions: str
    first_year_rate: float
    total_rate: float
    sheet_name: str
    row_number: int

    @property
    def label(self) -> str:
        detail = f" · {self.conditions}" if self.conditions else ""
        return f"{self.product}{detail}"


@dataclass(frozen=True)
class HoldingContract:
    row_key: str
    source_type: str
    insurer_raw: str
    insurer: str
    policy_number: str
    product_raw: str
    customer: str
    collector: str
    premium: int
    payment_years: int | None
    payment_label: str
    contract_date: str
    contract_month: str
    status: str
    share_rate: float


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("計", "계")
    return re.sub(r"[\s\n\r\t:()\[\]·ㆍ_-]+", "", text).lower()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text in {"", "-"}:
            return None
        if text.endswith("%"):
            try:
                return float(text[:-1]) / 100
            except ValueError:
                return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _effective_max_col(ws) -> int:
    """서식만 남아 XFD까지 확장된 시트의 불필요한 탐색을 막습니다."""
    upper = min(ws.max_column, 120)
    last = 1
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), max_col=upper):
        for cell in row:
            if cell.value not in (None, ""):
                last = max(last, cell.column)
    return min(max(last + 4, 20), upper)


def _is_header(value: Any, aliases: set[str]) -> bool:
    normalized = _normalize(value)
    return normalized in aliases


def _header_positions(ws, max_col: int) -> list[tuple[int, int, int, int]]:
    """상품명/1차년계/총수수료 열로 구성된 표 구간을 찾습니다."""
    positions: list[tuple[int, int, int, int]] = []
    for row_no in range(1, ws.max_row + 1):
        product_cols = [
            col for col in range(1, max_col + 1)
            if _is_header(ws.cell(row_no, col).value, PRODUCT_HEADERS)
        ]
        if not product_cols:
            product_cols = [
                col for col in range(1, max_col + 1)
                if _is_header(ws.cell(row_no, col).value, PRODUCT_FALLBACK_HEADERS)
            ]

        first_candidates: list[tuple[int, int]] = []
        total_candidates: list[tuple[int, int]] = []
        for header_row in range(row_no, min(row_no + 4, ws.max_row + 1)):
            for col in range(1, max_col + 1):
                value = ws.cell(header_row, col).value
                if _is_header(value, FIRST_YEAR_HEADERS):
                    first_candidates.append((header_row, col))
                if _is_header(value, TOTAL_HEADERS):
                    total_candidates.append((header_row, col))

        if not product_cols and first_candidates and total_candidates:
            product_cols = [1]
        if not product_cols:
            continue

        for product_col in product_cols:
            first_after = [item for item in first_candidates if item[1] > product_col]
            total_after = [item for item in total_candidates if item[1] > product_col]
            if not first_after or not total_after:
                continue
            first = min(first_after, key=lambda item: item[1])
            total = max(total_after, key=lambda item: item[1])
            if first[1] < total[1]:
                data_start = max(row_no, first[0], total[0]) + 1
                positions.append((data_start, product_col, first[1], total[1]))
                break
    return positions


def _condition_header(ws, header_row: int, col: int) -> str:
    """조건값의 의미를 잃지 않도록 실제 열 제목을 찾습니다."""
    for row_no in range(header_row, max(0, header_row - 3), -1):
        text = _clean_text(ws.cell(row_no, col).value)
        if text and not text.startswith("="):
            return text
    return ""


def _condition_value(header: str, value: Any) -> str:
    number = _number(value)
    if number is not None:
        text = f"{number:g}"
        if any(token in _normalize(header) for token in ("납기", "납입기간")):
            return text if any(unit in text for unit in ("년", "월", "회")) else f"{text}년"
        return text
    return _clean_text(value)


def _condition_text(
    ws,
    row_no: int,
    header_row: int,
    product_col: int,
    first_col: int,
    inherited: dict[int, Any],
) -> str:
    """납기·종형·만기·담보 등 원본 선택정보를 열 제목과 함께 보존합니다."""
    ignored_values = {"-", "상품별상이", "해당없음"}
    ignored_headers = {
        "상품명", "상품코드", "보종코드", "최초보험료", "월납보험료", "보험료", "가입금액"
    }
    parts: list[str] = []
    seen: set[str] = set()
    for col in range(1, first_col):
        if col == product_col:
            continue
        header = _condition_header(ws, header_row, col)
        value = ws.cell(row_no, col).value
        if value not in (None, ""):
            inherited[col] = value
        else:
            value = inherited.get(col)
        if isinstance(value, str) and value.startswith("="):
            continue
        # 병합표의 보조 열은 제목이 비어 있어도 '3년 보증·5년 보증'처럼
        # 선택에 필요한 조건을 담을 수 있습니다. 문자값만 세부조건으로 보존합니다.
        if not header:
            if not isinstance(value, str) or not _clean_text(value):
                continue
            header = f"세부조건{col}"
        normalized_header = _normalize(header)
        if normalized_header in {_normalize(item) for item in ignored_headers}:
            continue
        text = _condition_value(header, value)
        if not text or text in ignored_values:
            continue
        part = f"{header}: {text}"
        normalized_part = _normalize(part)
        if normalized_part in seen:
            continue
        seen.add(normalized_part)
        parts.append(part)
    return " / ".join(parts)


def _source_payout_rate(formula_ws, value_ws, max_col: int) -> float | None:
    for row_no in range(1, min(formula_ws.max_row, 5) + 1):
        for col in range(1, max_col + 1):
            label = _normalize(formula_ws.cell(row_no, col).value)
            if label not in {"지급율", "지급률"}:
                continue
            for offset in range(1, 4):
                rate = _number(value_ws.cell(row_no, col + offset).value)
                if rate is not None:
                    return rate
    return None


def _extract_sheet(
    formula_ws,
    value_ws,
    source_type: str,
) -> tuple[list[ProductRate], list[str]]:
    insurer = formula_ws.title.strip()
    max_col = _effective_max_col(formula_ws)
    tables = _header_positions(formula_ws, max_col)
    results: list[ProductRate] = []
    warnings: list[str] = []
    source_payout = _source_payout_rate(formula_ws, value_ws, max_col)

    if not tables:
        return results, warnings
    if source_payout == 0:
        warnings.append(
            f"{insurer}: 원본 예시표의 지급율이 0%입니다. 해당 시트를 100%로 저장한 뒤 다시 올려 주세요."
        )
        return results, warnings

    for table_index, (data_start, product_col, first_col, total_col) in enumerate(tables):
        next_start = tables[table_index + 1][0] if table_index + 1 < len(tables) else formula_ws.max_row + 1
        end_row = next_start - 2
        current_product = ""
        inherited_conditions: dict[int, Any] = {}
        header_row = data_start - 1

        for row_no in range(data_start, end_row + 1):
            raw_product = formula_ws.cell(row_no, product_col).value
            product_text = _clean_text(raw_product)

            if product_text:
                normalized_product = _normalize(product_text)
                if (
                    normalized_product in PRODUCT_HEADERS
                    or product_text.startswith("■")
                    or "수수료타입변경" in normalized_product
                ):
                    current_product = ""
                    inherited_conditions = {}
                    continue
                if product_text != current_product:
                    inherited_conditions = {}
                current_product = product_text

            if not current_product:
                continue

            first_rate = _number(value_ws.cell(row_no, first_col).value)
            total_rate = _number(value_ws.cell(row_no, total_col).value)
            if first_rate is None or total_rate is None or first_rate < 0 or total_rate < 0:
                continue
            if first_rate == 0 and total_rate == 0:
                continue

            if source_payout not in (None, 0):
                first_rate /= source_payout
                total_rate /= source_payout

            conditions = _condition_text(
                formula_ws, row_no, header_row, product_col, first_col, inherited_conditions
            )
            identity = f"{source_type}|{insurer}|{current_product}|{conditions}|{row_no}"
            key = hashlib.sha1(identity.encode("utf-8")).hexdigest()[:16]
            results.append(
                ProductRate(
                    key=key,
                    source_type=source_type,
                    insurer=insurer,
                    product=current_product,
                    conditions=conditions,
                    first_year_rate=first_rate,
                    total_rate=total_rate,
                    sheet_name=formula_ws.title,
                    row_number=row_no,
                )
            )

    if tables and not results:
        warnings.append(f"{insurer}: 표는 찾았지만 계산된 수수료율을 읽지 못했습니다.")
    return results, warnings


@st.cache_data(show_spinner=False)
def parse_commission_workbook(file_bytes: bytes, source_type: str) -> tuple[list[dict], list[str]]:
    """예시표의 저장된 계산 결과를 읽습니다. 원본 파일은 변경하지 않습니다."""
    formula_book = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
    value_book = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    products: list[dict] = []
    warnings: list[str] = []

    for sheet_name in formula_book.sheetnames:
        if "변경" in sheet_name or sheet_name not in value_book.sheetnames:
            continue
        extracted, sheet_warnings = _extract_sheet(
            formula_book[sheet_name], value_book[sheet_name], source_type
        )
        products.extend(item.__dict__ for item in extracted)
        warnings.extend(sheet_warnings)

    formula_book.close()
    value_book.close()
    return products, warnings


def _to_product_rate(item: dict) -> ProductRate:
    return ProductRate(**item)


def _format_rate(multiplier: float) -> str:
    return f"{multiplier * 100:,.1f}%"


def _format_won(value: float) -> str:
    return f"{round(value):,}원"


OUTPUT_INSURER_PREFIXES = {
    "KB라이프": ("KB라이프생명", "KB라이프", "KB"),
    "KB손보": ("KB손해보험", "KB손보", "KB"),
    "DB생명": ("DB생명보험", "DB생명", "DB"),
    "DB손보": ("DB손해보험", "DB손보", "DB"),
    "신한라이프": ("신한라이프", "신한"),
    "한화생명": ("한화생명", "한화"),
    "한화손보": ("한화손해보험", "한화손보", "한화"),
    "삼성생명": ("삼성생명", "삼성"),
    "삼성화재": ("삼성화재해상보험", "삼성화재", "삼성"),
    "미래에셋": ("미래에셋생명", "미래에셋"),
    "메리츠": ("메리츠화재", "메리츠"),
    "현대해상": ("현대해상화재보험", "현대해상"),
    "라이나생명": ("라이나생명", "라이나"),
    "교보생명": ("교보생명", "교보"),
    "농협생명": ("NH농협생명", "농협생명", "NH"),
    "농협손보": ("NH농협손해보험", "NH농협손보", "농협손해보험", "농협손보", "NH"),
    "흥국생명": ("흥국생명",),
    "흥국화재": ("흥국화재해상보험", "흥국화재"),
    "롯데손보": ("롯데손해보험", "롯데손보", "롯데"),
    "메트라이프": ("메트라이프생명", "메트라이프"),
    "ABL생명": ("ABL생명", "ABL"),
    "KDB생명": ("KDB생명", "KDB"),
    "IBK연금": ("IBK연금보험", "IBK연금", "IBK"),
    # '하나로라이트'처럼 상품명 자체가 하나로 시작할 수 있어 '하나' 단독은 제거하지 않습니다.
    "하나생명": ("하나생명",),
    "하나손보": ("하나손해보험", "하나손보"),
}


def _remove_output_insurer_prefix(product_name: str, insurer: str) -> str:
    standard = _standard_insurer(insurer)
    prefixes = OUTPUT_INSURER_PREFIXES.get(standard, (insurer,))
    result = product_name.strip()
    for prefix in sorted({_clean_text(value) for value in prefixes if value}, key=len, reverse=True):
        match = re.match(rf"^\s*{re.escape(prefix)}(?:\s*[-_:·]?\s*)", result, flags=re.I)
        if match:
            result = result[match.end():].strip()
            break
    return result


def _output_product_name(contract: dict) -> str:
    """화면과 다운로드에 공통으로 사용할 간결한 상품명을 만듭니다."""
    product_name, _ = _product_display_parts(contract.get("product", ""))
    insurer = _clean_text(contract.get("insurer", ""))
    for token in ("(무배당)", "무배당", "(무)", insurer):
        if token:
            product_name = product_name.replace(token, "")
    product_name = _remove_output_insurer_prefix(product_name, insurer)
    product_name = re.sub(
        r"\([^)]*(?:갱신|비갱신|심사|고지|해약|해지|환급|납입면제|세만기|연만기)[^)]*\)",
        "", product_name,
    )
    product_name = re.sub(
        r"(?:해약|해지)환급금[^\s_/)]*|납입면제형|갱신형|비갱신형|"
        r"일반심사형|간편심사형|건강고지형|세만기형?|연만기형?|무해지형|일반해지형",
        "", product_name,
    )
    product_name = re.sub(
        r"간편가입|보험가입금액형|보험료형|납입면제적용형",
        "", product_name,
    )
    product_name = re.sub(r"(?<!\d)\d{1,2}(?:종|형)(?!\d)", "", product_name)
    product_name = re.sub(r"\(+\s*\d+(?:\.\d+)+\s*\)+", "", product_name)
    product_name = re.sub(r"\(\s*\)", "", product_name)
    product_name = product_name.replace(")", "").replace("(", "")
    product_name = re.sub(
        r"(?:20)?(?:2[4-9]|3\d)(?:0[1-9]|1[0-2])\s*$",
        "", product_name,
    )
    product_name = re.sub(r"\((?:\s*|(?:20)?\d{2}[.\-]\d{2})\)", "", product_name)
    product_name = product_name.replace("_", " ")
    product_name = re.sub(r"\s+", " ", product_name).strip(" _·-/")
    return product_name


def _contract_payment_label(contract: dict) -> str:
    label = _clean_text(contract.get("payment_label", ""))
    if label:
        return label
    detail = _clean_text(f"{contract.get('product', '')} {contract.get('conditions', '')}")
    match = re.search(r"(?<!\d)(\d+)\s*년\s*(납|갱신|만기)", detail)
    return f"{match.group(1)}년{match.group(2)}" if match else ""


def _contract_renewal_label(contract: dict) -> str:
    text = _normalize(f"{contract.get('product', '')} {contract.get('conditions', '')}")
    if "비갱신" in text:
        return "비갱신"
    if "갱신" in text:
        return "갱신"
    if "세만기" in text:
        return "세만기"
    if "연만기" in text:
        return "연만기"
    return ""


def _rate_distinguishing_labels(contract: dict) -> list[str]:
    tags = _selection_tags(f"{contract.get('product', '')} {contract.get('conditions', '')}")
    labels: list[str] = []
    for category in ("type", "form", "underwriting", "surrender", "plan"):
        for value in sorted(tags.get(category, set())):
            if value not in labels:
                labels.append(value)
    return labels[:3]


def _compact_product_display(contract: dict, peers: list[dict] | None = None) -> str:
    """첫 줄 상품명, 둘째 줄 납기·갱신 여부와 꼭 필요한 구분값만 표시합니다."""
    product_name = _output_product_name(contract)
    payment = _contract_payment_label(contract)
    renewal = _contract_renewal_label(contract)
    condition_labels = [value for value in (payment, renewal) if value]

    peer_rows = peers or [contract]
    same_basic = [
        peer for peer in peer_rows
        if _output_product_name(peer) == product_name
        and _contract_payment_label(peer) == payment
        and _contract_renewal_label(peer) == renewal
    ]
    rate_pairs = {
        (round(float(peer.get("first_year_rate", 0)), 8), round(float(peer.get("total_rate", 0)), 8))
        for peer in same_basic
    }
    if len(rate_pairs) > 1:
        condition_labels.extend(
            label for label in _rate_distinguishing_labels(contract)
            if label not in condition_labels
        )
    condition_line = " · ".join(condition_labels)
    return f"{product_name}\n{condition_line}" if condition_line else product_name


def _make_excel(
    contracts: list[dict], payout_rate: float, reference_month: str, excluded: list[dict],
    fallback_collectors: list[str] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "수수료 계산"
    total_premium = sum(item["premium"] for item in contracts)
    total_first = sum(item["premium"] * item["first_year_rate"] * payout_rate for item in contracts)
    total_commission = sum(item["premium"] * item["total_rate"] * payout_rate for item in contracts)
    collector_label = _collector_label(contracts, fallback_collectors)
    title = f"{collector_label} 수수료 계산 결과" if collector_label else "수수료 계산 결과"
    ws.append([title])
    ws.append(["수수료표 기준월", reference_month or "확인 필요", "공통 지급율", payout_rate])
    ws.append(["계약 수", len(contracts), "월보험료 합계", total_premium])
    ws.append(["예상 익월수당 합계", round(total_first), "예상 총수당 합계", round(total_commission)])
    ws.append([])
    headers = ["고객명", "증권번호", "보험회사", "상품 및 세부 조건", "월보험료", "모집 정보",
               "익월 수수료율", "총수수료율", "예상 익월수당", "예상 총수당"]
    ws.append(headers)

    for contract in contracts:
        first_rate = contract["first_year_rate"] * payout_rate
        total_rate = contract["total_rate"] * payout_rate
        premium = contract["premium"]
        product_detail = _compact_product_display(contract, contracts)
        share_rate = contract.get("share_rate", 100.0)
        recruiter_type = contract.get("recruiter_type", "")
        recruiting = f"{share_rate:g}%"
        if share_rate < 100 and recruiter_type:
            recruiting += f" · {recruiter_type}"
        ws.append([
            contract.get("customer", ""),
            contract.get("policy_number", ""),
            contract["insurer"],
            product_detail,
            premium,
            recruiting,
            first_rate,
            total_rate,
            round(premium * first_rate),
            round(premium * total_rate),
        ])

    header_fill = PatternFill("solid", fgColor="2563D9")
    ws.merge_cells("A1:J1")
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[6]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["D2"].number_format = "0%"
    for cell in (ws["D3"], ws["A4"], ws["C4"]):
        cell.font = Font(bold=True)
    ws["B3"].number_format = '0"건"'
    for cell in (ws["D3"], ws["B4"], ws["D4"]):
        cell.number_format = '#,##0"원"'
    for row in range(7, ws.max_row + 1):
        ws.cell(row, 5).number_format = '#,##0"원"'
        for col in range(7, 9):
            ws.cell(row, col).number_format = "0.0%"
        for col in range(9, 11):
            ws.cell(row, col).number_format = '#,##0"원"'
        ws.row_dimensions[row].height = 42

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [15, 22, 18, 64, 18, 20, 18, 18, 21, 21]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:J{ws.max_row}"
    ws.row_dimensions[1].height = 28

    review_ws = wb.create_sheet("검토 제외 계약")
    review_headers = ["고객명", "증권번호", "보험회사", "상품명", "계약상태", "제외 사유"]
    review_ws.append(review_headers)
    for item in excluded:
        review_ws.append([
            item.get("customer", ""), item.get("policy_number", ""), item.get("insurer", ""),
            _compact_product_display({"product": item.get("product", ""), "insurer": item.get("insurer", "")}),
            item.get("status", ""), item.get("reason", ""),
        ])
    for cell in review_ws[1]:
        cell.fill = PatternFill("solid", fgColor="64748B")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col, width in enumerate([15, 22, 18, 64, 15, 55], start=1):
        review_ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(2, review_ws.max_row + 1):
        review_ws.row_dimensions[row].height = 36
    for row in review_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    review_ws.freeze_panes = "A2"
    if review_ws.max_row > 1:
        review_ws.auto_filter.ref = review_ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


INSURER_ALIASES = {
    "KB라이프생명": "KB라이프", "KB라이프": "KB라이프",
    "DB손해보험": "DB손보", "DB손보": "DB손보",
    "KB손해보험": "KB손보", "KB손보": "KB손보",
    "메리츠화재": "메리츠", "메리츠": "메리츠",
    "한화손해보험": "한화손보", "한화손보": "한화손보",
    "메트라이프생명": "메트라이프", "메트라이프": "메트라이프",
    "미래에셋생명": "미래에셋", "미래에셋": "미래에셋",
    "DB생명보험": "DB생명", "DB생명": "DB생명",
    "농협생명보험": "농협생명", "NH농협생명": "농협생명",
    "농협손해보험": "농협손보", "NH농협손해보험": "농협손보",
    "NH농협손보": "농협손보", "농협손보": "농협손보",
    "삼성생명보험": "삼성생명", "삼성화재해상보험": "삼성화재",
    "현대해상화재보험": "현대해상", "흥국화재해상보험": "흥국화재",
}


def _standard_insurer(value: Any) -> str:
    text = _clean_text(value)
    return INSURER_ALIASES.get(text, text)


def _source_type_from_insurer(insurer: str, insurer_code: str = "") -> str:
    if "생명" in insurer or "라이프" in insurer or insurer_code.upper().startswith("L"):
        return "생보"
    return "손보"


def _month_from_filename(filename: str) -> str:
    patterns = [r"(20\d{2})[._년\-\s]*(0?[1-9]|1[0-2])\s*월", r"(20\d{2})[._\-](0?[1-9]|1[0-2])"]
    for pattern in patterns:
        match = re.search(pattern, filename)
        if match:
            return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return ""


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = _clean_text(value)
    match = re.search(r"(20\d{2})[./\-년\s]*(\d{1,2})[./\-월\s]*(\d{1,2})?", text)
    if not match:
        return text
    year, month = int(match.group(1)), int(match.group(2))
    day = int(match.group(3) or 1)
    return f"{year:04d}-{month:02d}-{day:02d}"


def _holding_product_name(value: Any) -> str:
    text = _clean_text(value).lower()
    text = re.sub(r"\(\s*\d+\s*\)", "", text)
    replacements = (
        "무배당", "(무)", "_무", "상품개정",
        "해약환급금", "해지환급금", "미지급형", "납입면제형",
    )
    for token in replacements:
        text = text.replace(token, "")
    text = re.sub(r"\(?(?:20\d{2}|2\d)[.\-](?:0?[1-9]|1[0-2])\)?", "", text)
    for token in (
        "간편가입", "간편심사형", "일반심사형", "보험가입금액형", "보험료형",
        "일부지급형", "저해약환급금형", "보증비용부과형", "간편",
    ):
        text = text.replace(token, "")
    text = re.sub(r"(?<!\d)\d{1,2}형(?!\d)", "", text)
    # 보험회사는 별도 항목에서 먼저 일치시키므로 상품명 앞의 브랜드 표기는 비교에서 제외합니다.
    brand_prefixes = (
        "kb라이프생명", "kb라이프", "kb손해보험", "kb손보", "kb",
        "db손해보험", "db손보", "db생명", "db",
        "nh농협생명", "nh농협손해보험", "농협생명", "농협손보", "nh",
        "신한라이프", "신한", "한화생명", "한화손해보험", "한화손보", "한화",
        "삼성생명", "삼성화재", "삼성", "흥국생명", "흥국화재", "흥국",
        "미래에셋생명", "미래에셋", "메트라이프생명", "메트라이프",
        "abl생명", "abl", "ibk연금", "ibk", "kdb생명", "kdb",
        "교보생명", "교보", "라이나생명", "라이나", "카디프생명", "카디프",
        "현대해상", "현대", "메리츠화재", "메리츠", "롯데손보", "롯데",
        "하나손보", "하나생명", "하나", "aig손보", "aig", "mg손보", "mg",
    )
    normalized_prefixes = sorted(
        (re.sub(r"[^0-9a-z가-힣]", "", token.lower()) for token in brand_prefixes),
        key=len,
        reverse=True,
    )
    compact = re.sub(r"[^0-9a-z가-힣]", "", text)
    for prefix in normalized_prefixes:
        if compact.startswith(prefix):
            compact = compact[len(prefix):]
            break
    # 한 셀에 보험사명이 두 번 반복된 원본도 있어, 명확한 회사명만 추가 제거합니다.
    for brand in (
        "한화생명", "한화손해보험", "신한라이프", "미래에셋생명",
        "메트라이프생명", "kb라이프생명", "db손해보험", "kb손해보험",
    ):
        compact = compact.replace(brand, "")
    return compact


def _product_family_name(value: Any) -> str:
    """갱신 여부처럼 별도 확인해야 할 표지만 제외한 기본 상품군 이름입니다."""
    text = _holding_product_name(value)
    for token in ("비갱신형", "갱신형", "세만기형", "연만기형", "비갱신", "갱신", "세만기", "연만기"):
        text = text.replace(token, "")
    return text


def _product_display_parts(product_name: str) -> tuple[str, str]:
    """상품명 셀 뒤에 붙은 납입 안내 문구를 화면용 상품명에서 분리합니다."""
    text = _clean_text(product_name)
    split_match = re.search(
        r"\s+(?=(?:\d+(?:\.\d+)?\s*구좌|\d+\s*년납(?:초과|이상|이하)))",
        text,
    )
    if not split_match:
        return text, ""
    return text[:split_match.start()].strip(), text[split_match.end():].strip()


def _condition_display(product: ProductRate) -> str:
    _, product_note = _product_display_parts(product.product)
    parts = [part for part in (product_note, product.conditions) if part]
    return " / ".join(dict.fromkeys(parts)) or "기본 조건"


def _condition_pairs(product: ProductRate) -> list[tuple[str, str]]:
    """저장된 원본 조건 문구를 화면 표시용 제목·값으로 분리합니다."""
    pairs: list[tuple[str, str]] = []
    for part in re.split(r"\s*/\s*", _condition_display(product)):
        if ":" in part:
            header, value = part.split(":", 1)
            pairs.append((_clean_text(header), _clean_text(value)))
        elif part and part != "기본 조건":
            pairs.append(("세부조건", _clean_text(part)))
    return pairs


def _short_condition_label(product: ProductRate, candidates: list[ProductRate] | None = None) -> str:
    """선택에 필요한 조건만 남기고 반복되는 원본 설명은 숨깁니다."""
    pairs = _condition_pairs(product)
    candidate_pairs = [_condition_pairs(item) for item in (candidates or [product])]
    values_by_header: dict[str, set[str]] = defaultdict(set)
    for item_pairs in candidate_pairs:
        for header, value in item_pairs:
            values_by_header[_normalize(header)].add(_normalize(value))

    always_tokens = ("납기", "납입기간")
    useful_tokens = (
        "구분", "종형", "형구분", "종구분", "만기", "담보", "보험기간",
        "납입주기", "규정", "연령", "기준",
        "특약유형", "보종구분", "상품군", "세부조건",
    )
    hidden_tokens = (
        "랩포탈상품군", "고성과수수료지급대상여부", "장기유지수수료",
        "특정상품보너스", "commissionrate", "retentionbonus",
        "수정율", "수정률", "환산", "환산율",
    )
    selected: list[tuple[int, str]] = []
    for header, value in pairs:
        normalized_header = _normalize(header)
        normalized_value = _normalize(value)
        if any(token in normalized_header for token in hidden_tokens):
            continue
        if normalized_value in {"", "0", "0.0", "y", "단일"}:
            continue
        is_payment = any(token in normalized_header for token in always_tokens)
        differs = len(values_by_header.get(normalized_header, set())) > 1
        if not is_payment and not differs:
            continue
        if not is_payment and not any(token in normalized_header for token in useful_tokens):
            continue
        concise_header = header.replace("*평균 ", "").replace(" (보장+적립)", "")
        if concise_header.startswith("세부조건"):
            concise_header = "세부조건"
        priority = 0 if is_payment else 1
        selected.append((priority, f"{concise_header} {value}"))

    labels: list[str] = []
    for _, label in sorted(selected, key=lambda item: item[0]):
        if label not in labels:
            labels.append(label)
    return " · ".join(labels) or "기본 조건"


def _condition_option_label(
    product: ProductRate,
    payout_rate: float | None = None,
    candidates: list[ProductRate] | None = None,
) -> str:
    """선택 조건과 공통 지급율이 반영된 익월·총 수수료율을 표시합니다."""
    applied_rate = payout_rate
    if applied_rate is None:
        applied_rate = (
            float(st.session_state.get("commission_payout_rate", DEFAULT_PAYOUT_RATE))
            / 100
        )
    return (
        f"{_short_condition_label(product, candidates)} · "
        f"익월 {_format_rate(product.first_year_rate * applied_rate)} · "
        f"총 {_format_rate(product.total_rate * applied_rate)}"
    )


@st.cache_data(show_spinner=False)
def parse_holding_workbook(file_bytes: bytes) -> list[dict]:
    """보유계약 장기 파일을 읽습니다. 잘못된 dimension=A1 파일도 처리합니다."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = {_normalize(cell.value): cell.column for cell in ws[1] if cell.value not in (None, "")}

    def value(row: int, *names: str) -> Any:
        for name in names:
            col = headers.get(_normalize(name))
            if col:
                return ws.cell(row, col).value
        return None

    results: list[dict] = []
    for row in range(2, ws.max_row + 1):
        policy_number = _clean_text(value(row, "증권번호"))
        product = _clean_text(value(row, "상품명"))
        insurer_raw = _clean_text(value(row, "보험사"))
        if not product and not policy_number:
            continue
        insurer = _standard_insurer(insurer_raw)
        date_value = _date_text(value(row, "계약일"))
        payment_year_number = _number(value(row, "납입기간"))
        payment_years = int(payment_year_number) if payment_year_number is not None else None
        payment_unit = _clean_text(value(row, "납입기간구분"))
        payment_label = f"{payment_years}{payment_unit}" if payment_years is not None else ""
        share_number = _number(value(row, "쉐어율"))
        share_rate = float(share_number if share_number is not None else 100.0)
        insurer_code = _clean_text(value(row, "보험사코드"))
        identity = f"{policy_number}|{product}|{date_value}|{row}"
        holding = HoldingContract(
            row_key=hashlib.sha1(identity.encode("utf-8")).hexdigest()[:16],
            source_type=_source_type_from_insurer(insurer, insurer_code),
            insurer_raw=insurer_raw,
            insurer=insurer,
            policy_number=policy_number,
            product_raw=product,
            customer=_clean_text(value(row, "계약자")),
            collector=_clean_text(value(row, "수금자명", "수금자")),
            premium=int(_number(value(row, "계속보험료", "초회보험료")) or 0),
            payment_years=payment_years,
            payment_label=payment_label,
            contract_date=date_value,
            contract_month=date_value[:7] if re.match(r"20\d{2}-\d{2}", date_value) else "",
            status=_clean_text(value(row, "계약상태")) or "확인 필요",
            share_rate=share_rate,
        )
        results.append(holding.__dict__)
    wb.close()
    return results


def _payment_matches(product: ProductRate, years: int | None) -> bool:
    if years is None:
        return True
    condition = _normalize(product.conditions)
    # 납기 구분 자체가 없는 기본·단일 조건은 특정 납기와 충돌하지 않습니다.
    if not condition or not _has_payment_condition(product):
        return True
    if re.search(rf"(?<!\d){years}년(?:납|갱신|만기)", condition):
        return True
    if re.search(rf"(?:납기|납입기간){years}(?:년|년납)?", condition):
        return True
    over = re.search(r"(\d+)년납(?:이상|↑)", condition)
    return bool(over and years >= int(over.group(1)))


def _has_payment_condition(product: ProductRate) -> bool:
    condition = _normalize(product.conditions)
    return bool(
        re.search(r"\d+(?:~\d+)?년(?:납|갱신|만기)", condition)
        or re.search(r"(?:납기|납입기간)\d+(?:년|년납)?", condition)
    )


def _payment_condition_label(product: ProductRate) -> str:
    """불일치 안내에 사용할 수수료표의 납기 문구를 간결하게 반환합니다."""
    condition = _normalize(product.conditions)
    labels: list[str] = []
    for years, suffix in re.findall(r"(?<!\d)(\d+)년(납|갱신|만기)", condition):
        label = f"{int(years)}년{suffix}"
        if label not in labels:
            labels.append(label)
    for years in re.findall(r"(?:납기|납입기간)(\d+)(?:년|년납)?", condition):
        label = f"{int(years)}년납"
        if label not in labels:
            labels.append(label)
    return " / ".join(labels) or "납기 표기 없음"


def _payment_threshold(product: ProductRate) -> int | None:
    condition = _normalize(product.conditions)
    match = re.search(r"(?<!\d)(\d+)년납(?:이상|↑)", condition)
    return int(match.group(1)) if match else None


def _most_specific_payment_candidates(
    products: list[ProductRate], years: int | None
) -> list[ProductRate]:
    """'10년납 이상'처럼 겹치는 구간에서는 계약 납기에 가장 가까운 하한만 남깁니다."""
    if years is None or not products:
        return products
    thresholds = [
        threshold for product in products
        if (threshold := _payment_threshold(product)) is not None and threshold <= years
    ]
    if not thresholds:
        return products
    best_threshold = max(thresholds)
    narrowed = [
        product for product in products
        if _payment_threshold(product) == best_threshold
    ]
    return narrowed or products


def _selection_tags(value: Any) -> dict[str, set[str]]:
    """상품명과 세부조건에서 자동 연결에 신뢰할 수 있는 핵심 표지를 찾습니다."""
    text = _normalize(value)
    tags: dict[str, set[str]] = defaultdict(set)

    # 해지환급금 유형
    if any(token in text for token in ("무해지", "해약환급금미지급", "해지환급금미지급", "환급금이없는")):
        tags["surrender"].add("무해지")
    if "일반해지" in text:
        tags["surrender"].add("일반해지")
    if any(token in text for token in ("유해지", "일부지급", "50%지급", "일부환급")):
        tags["surrender"].add("일부지급")

    # 갱신·만기 유형. '비갱신'을 단순 '갱신'으로 중복 인식하지 않습니다.
    if "비갱신" in text:
        tags["renewal"].add("비갱신")
    elif "갱신" in text:
        tags["renewal"].add("갱신")
    if "세만기" in text:
        tags["maturity"].add("세만기")
    if "연만기" in text:
        tags["maturity"].add("연만기")

    # 심사·고지 유형
    for token, canonical in (
        ("초경증", "초경증"), ("통합간편", "통합간편"),
        ("간편심사", "간편"), ("간편가입", "간편"),
        ("일반심사", "일반심사"), ("건강고지", "건강고지"),
    ):
        if token in text:
            tags["underwriting"].add(canonical)
    if "간편" in text and "underwriting" not in tags:
        tags["underwriting"].add("간편")

    # 종형과 설계 형태
    for match in re.finditer(r"(?<!\d)(\d{1,2})종(?!\d)", text):
        tags["type"].add(f"{int(match.group(1))}종")
    for match in re.finditer(r"(?<!\d)(\d{1,2})형(?!\d)", text):
        tags["form"].add(f"{int(match.group(1))}형")
    for token in ("납입면제형", "보험가입금액형", "보험료형", "기본형", "보장강화형"):
        if token in text:
            tags["plan"].add(token)

    # 보증기간은 상품명의 다른 숫자와 혼동하지 않도록 '보증' 단위만 사용합니다.
    for match in re.finditer(r"(?<!\d)(\d+)년보증", text):
        tags["guarantee"].add(f"{int(match.group(1))}년보증")
    return {category: values for category, values in tags.items() if values}


def _tag_match_summary(holding: dict, product: ProductRate) -> tuple[int, int, list[str]]:
    """일치·충돌 개수와 사용자가 이해할 수 있는 일치 근거를 반환합니다."""
    source_tags = _selection_tags(holding.get("product_raw", ""))
    target_tags = _selection_tags(f"{product.product} {product.conditions}")
    matched = 0
    conflicts = 0
    reasons: list[str] = []
    for category, expected in source_tags.items():
        actual = target_tags.get(category, set())
        if expected & actual:
            matched += 1
            reasons.extend(sorted(expected & actual))
        elif actual:
            conflicts += 1
    return matched, conflicts, reasons


def _filter_by_holding_tags(holding: dict, products: list[ProductRate]) -> list[ProductRate]:
    """보유계약에 명시된 조건과 충돌하는 후보를 제거하되, 근거가 없으면 억지 제거하지 않습니다."""
    filtered = list(products)
    source_tags = _selection_tags(holding.get("product_raw", ""))
    for category, expected in source_tags.items():
        matching = []
        conflicting = []
        unspecified = []
        for product in filtered:
            actual = _selection_tags(f"{product.product} {product.conditions}").get(category, set())
            if expected & actual:
                matching.append(product)
            elif actual:
                conflicting.append(product)
            else:
                unspecified.append(product)
        # 같은 조건을 명시한 후보가 있을 때만 반대 조건을 제거합니다.
        # 조건이 비어 있는 후보는 자동 확정에 사용하지 않도록 일치 후보를 우선합니다.
        if matching:
            filtered = matching
        elif conflicting and not unspecified:
            return []
    return filtered


def _condition_sort_key(product: ProductRate) -> tuple:
    """같은 해지·갱신 유형끼리 묶고 납입기간 순으로 정렬합니다."""
    text = _normalize(f"{product.product} {product.conditions}")

    if "무해지" in text or "해약환급금미지급" in text or "해지환급금미지급" in text:
        surrender_rank = 0
    elif "일반해지" in text or "일반형" in text:
        surrender_rank = 1
    elif "유해지" in text or "일부지급" in text or "50%지급" in text:
        surrender_rank = 2
    else:
        surrender_rank = 3

    if "비갱신" in text or "세만기" in text:
        renewal_rank = 0
    elif "연만기" in text:
        renewal_rank = 1
    elif "갱신" in text:
        renewal_rank = 2
    else:
        renewal_rank = 3

    condition_values = dict((_normalize(h), _normalize(v)) for h, v in _condition_pairs(product))
    category = "".join(
        value for header, value in condition_values.items()
        if any(token in header for token in ("구분", "종형", "형구분", "종구분", "담보"))
    )
    payment_match = re.search(r"(?<!\d)(\d+)(?:년)?(?:납|갱신)", text)
    if not payment_match:
        payment_match = re.search(r"(?:납기|납입기간)(\d+)(?:년)?", text)
    payment_years = int(payment_match.group(1)) if payment_match else 999
    return surrender_rank, renewal_rank, category, payment_years, text, product.row_number


def _sort_condition_candidates(products: list[ProductRate]) -> list[ProductRate]:
    return sorted(products, key=_condition_sort_key)


PRODUCT_CATEGORY_TOKENS = {
    "실손": ("실손", "실비", "의료비"),
    "건강": ("건강보험", "종합건강", "건강보장"),
    "치아": ("치아", "치과"),
    "운전자": ("운전자",),
    "펫": ("펫", "반려", "애견", "애묘"),
    "종신": ("종신",),
    "연금": ("연금",),
    "치매간병": ("치매", "간병", "장기요양", "돌봄"),
    "어린이": ("어린이", "자녀", "키즈", "태아"),
    "화재재물": ("화재", "재물", "사업장", "주택"),
    "암": ("암보험", "암보장", "암치료", "암플랜"),
}


def _product_categories(value: Any) -> set[str]:
    text = _normalize(value)
    return {
        category for category, tokens in PRODUCT_CATEGORY_TOKENS.items()
        if any(token in text for token in tokens)
    }


def _strip_revision_markers(value: Any) -> str:
    """상품 개정월은 약하게 처리하고 3.10.5·3N5·0545 같은 핵심 숫자는 보존합니다."""
    text = _clean_text(value).lower()
    # 괄호 속 YY.MM, YYYY.MM, YYMM.회차 형태는 대부분 개정 표기입니다.
    text = re.sub(r"\(\s*(?:20)?\d{2}[.\-/](?:0?[1-9]|1[0-2])(?:[.\-/]\d+)?\s*\)", " ", text)
    # 상품명 끝 또는 보험/plus 바로 뒤의 2404·2607 형식만 개정월로 봅니다.
    text = re.sub(
        r"(보험|plus)\s*(?:20)?(?:2[4-9]|3\d)(?:0[1-9]|1[0-2])(?=\s|_|$|\()",
        r"\1 ", text, flags=re.I,
    )
    text = re.sub(
        r"(?<!\d)(?:20)?(?:2[4-9]|3\d)(?:0[1-9]|1[0-2])(?=\s|_|$)",
        " ", text,
    )
    return text


def _smart_product_name(value: Any) -> str:
    return _holding_product_name(_strip_revision_markers(value))


def _structural_signatures(value: Any) -> set[str]:
    """개정월과 달리 상품 정체성에 직접 쓰이는 숫자 표지를 추출합니다."""
    text = _strip_revision_markers(value).lower().replace("·", ".")
    signatures: set[str] = set()
    signatures.update(item.replace(".", ".") for item in re.findall(r"(?<!\d)\d{1,2}(?:\.\d{1,2}){1,3}(?!\d)", text))
    signatures.update(item.lower() for item in re.findall(r"(?<![0-9a-z])\d{1,2}n\d{1,2}(?![0-9a-z])", text, re.I))
    # 0545·0550·4565처럼 상품 핵심명 중간에 남는 4자리 숫자만 보존합니다.
    signatures.update(re.findall(r"(?<!\d)(?:0[3-9]\d{2}|[3-9]\d{3})(?!\d)", text))
    return signatures


def _bigrams(value: str) -> set[str]:
    return {value[index:index + 2] for index in range(max(0, len(value) - 1))}


def _name_similarity(source: str, target: str) -> float:
    if not source or not target:
        return 0.0
    sequence = SequenceMatcher(None, source, target).ratio()
    source_pairs, target_pairs = _bigrams(source), _bigrams(target)
    jaccard = (
        len(source_pairs & target_pairs) / len(source_pairs | target_pairs)
        if source_pairs and target_pairs else 0.0
    )
    containment = min(len(source), len(target)) / max(len(source), len(target)) if source in target or target in source else 0.0
    return max(sequence * 0.62 + jaccard * 0.38, containment)


def _hard_product_conflict(source: Any, target: Any) -> bool:
    source_categories = _product_categories(source)
    target_categories = _product_categories(target)
    if source_categories and target_categories and not source_categories.intersection(target_categories):
        return True
    source_signatures = _structural_signatures(source)
    target_signatures = _structural_signatures(target)
    return bool(source_signatures and target_signatures and source_signatures.isdisjoint(target_signatures))


def _rank_products(holding: dict, products: list[ProductRate]) -> list[tuple[float, ProductRate]]:
    source = _smart_product_name(holding["product_raw"])
    ranked: list[tuple[float, ProductRate]] = []
    for product in products:
        if product.source_type != holding["source_type"] or product.insurer != holding["insurer"]:
            continue
        if _hard_product_conflict(holding["product_raw"], product.product):
            continue
        target = _smart_product_name(product.product)
        if not source or not target:
            continue
        score = _name_similarity(source, target)
        if source == target:
            score = 1.0
        elif source in target or target in source:
            score = min(0.98, score + 0.06)
        if _payment_matches(product, holding.get("payment_years")):
            score += 0.025
        elif holding.get("payment_years") is not None and _has_payment_condition(product):
            score -= 0.10
        matched_tags, conflicting_tags, _ = _tag_match_summary(holding, product)
        score += min(matched_tags, 4) * 0.025
        score -= conflicting_tags * 0.12
        ranked.append((score, product))
    ranked.sort(key=lambda item: item[0], reverse=True)
    return ranked


def _ranked_product_groups(
    holding: dict, products: list[ProductRate]
) -> list[tuple[float, str, list[ProductRate]]]:
    groups: dict[str, list[tuple[float, ProductRate]]] = defaultdict(list)
    for score, product in _rank_products(holding, products):
        display_name, _ = _product_display_parts(product.product)
        groups[display_name].append((score, product))
    ranked_groups = []
    for display_name, rows in groups.items():
        best_score = max(score for score, _ in rows)
        unique: dict[tuple, ProductRate] = {}
        for _, product in rows:
            key = (product.conditions, round(product.first_year_rate, 8), round(product.total_rate, 8))
            unique.setdefault(key, product)
        ranked_groups.append((best_score, display_name, _sort_condition_candidates(list(unique.values()))))
    ranked_groups.sort(key=lambda item: (-item[0], _normalize(item[1])))
    return ranked_groups


def _candidate_products(holding: dict, products: list[ProductRate]) -> list[ProductRate]:
    groups = _ranked_product_groups(holding, products)
    if not groups or groups[0][0] < 0.56:
        return []
    candidates = groups[0][2]
    payment_years = holding.get("payment_years")
    payment_filtered = [p for p in candidates if _payment_matches(p, payment_years)]
    # 같은 상품을 찾았다면 납기 불일치만으로 '미연결' 처리하지 않습니다.
    # 일치 납기가 없을 때는 화면에서 사용자가 다른 납기를 명시적으로 펼칩니다.
    candidates = payment_filtered if payment_years is not None and payment_filtered else candidates
    candidates = _most_specific_payment_candidates(candidates, payment_years) if payment_filtered else candidates
    candidates = _filter_by_holding_tags(holding, candidates)
    if not candidates:
        return []
    unique: dict[tuple, ProductRate] = {}
    for product in candidates:
        key = (product.conditions, round(product.first_year_rate, 8), round(product.total_rate, 8))
        unique.setdefault(key, product)
    return _sort_condition_candidates(list(unique.values()))[:12]


def _review_candidate_products(holding: dict, products: list[ProductRate]) -> list[ProductRate]:
    """확인 화면에는 서로 다른 추천 상품을 최대 3개까지만 제공합니다."""
    groups = _ranked_product_groups(holding, products)
    if not groups or groups[0][0] < 0.56:
        return []
    floor = max(0.56, groups[0][0] - 0.16)
    selected_groups = [group for group in groups if group[0] >= floor][:3]
    return [product for _, _, candidates in selected_groups for product in candidates]


def _auto_candidate(holding: dict, products: list[ProductRate]) -> ProductRate | None:
    groups = _ranked_product_groups(holding, products)
    candidates = _candidate_products(holding, products)
    if not groups or not candidates:
        return None
    payment_years = holding.get("payment_years")
    if payment_years is not None and not any(
        _payment_matches(product, payment_years) for product in groups[0][2]
    ):
        return None
    best_candidate_score = groups[0][0]
    next_group_score = groups[1][0] if len(groups) > 1 else 0.0
    matched_tags, conflicting_tags, _ = _tag_match_summary(holding, candidates[0])
    source_name = _smart_product_name(holding["product_raw"])
    target_name = _smart_product_name(candidates[0].product)
    exact_core = source_name == target_name
    clear_margin = best_candidate_score - next_group_score >= 0.10
    if conflicting_tags or not (exact_core or (best_candidate_score >= 0.90 and clear_margin)):
        return None
    rate_pairs = {(round(p.first_year_rate, 8), round(p.total_rate, 8)) for p in candidates}
    if len(candidates) == 1 or (len(rate_pairs) == 1 and matched_tags >= 1):
        return candidates[0]
    return None


@st.cache_data(show_spinner=False)
def _analyze_product_links(
    holdings: list[dict], product_rows: list[dict]
) -> dict[str, dict[str, Any]]:
    """업로드 직후 한 번만 전체 추천을 계산하고 선택 조작 시에는 결과를 재사용합니다."""
    products = [_to_product_rate(row) for row in product_rows]
    products_by_insurer: dict[tuple[str, str], list[ProductRate]] = defaultdict(list)
    for product in products:
        products_by_insurer[(product.source_type, product.insurer)].append(product)
    decisions: dict[str, dict[str, Any]] = {}
    for holding in holdings:
        insurer_products = products_by_insurer.get(
            (holding.get("source_type", ""), holding.get("insurer", "")), []
        )
        candidates = _candidate_products(holding, insurer_products)
        review = _review_candidate_products(holding, insurer_products)
        automatic = _auto_candidate(holding, insurer_products)
        decisions[holding["row_key"]] = {
            "candidate_keys": [product.key for product in candidates],
            "review_keys": [product.key for product in review],
            "auto_key": automatic.key if automatic else "",
        }
    return decisions


def _initialize_state() -> None:
    st.session_state.setdefault("commission_contracts", [])
    st.session_state.setdefault("commission_payout_rate", DEFAULT_PAYOUT_RATE)
    st.session_state.setdefault("commission_edit_index", None)
    st.session_state.setdefault("commission_edit_request", None)
    st.session_state.setdefault("commission_ratebook_signature", "")
    st.session_state.setdefault("commission_import_collectors", [])
    st.session_state.setdefault("commission_import_contract_months", [])


def _reconnect_contract_rates(contracts: list[dict], products: list[ProductRate]) -> tuple[int, int]:
    """새 수수료표에서 기존 상품·세부 조건이 같은 계약의 요율을 다시 연결합니다."""
    updated_count = 0
    unresolved_count = 0
    for contract in contracts:
        same_product = [
            product for product in products
            if product.source_type == contract.get("source_type")
            and product.insurer == contract.get("insurer")
            and _holding_product_name(product.product) == _holding_product_name(contract.get("product", ""))
        ]
        same_condition = [
            product for product in same_product
            if _normalize(_condition_display(product)) == _normalize(contract.get("conditions", "") or "기본 조건")
        ]
        candidates = same_condition or same_product
        unique_rates = {
            (round(product.first_year_rate, 8), round(product.total_rate, 8)) for product in candidates
        }
        if not candidates or len(unique_rates) != 1:
            contract["rate_recheck_required"] = True
            unresolved_count += 1
            continue
        selected = candidates[0]
        contract.update({
            "product": selected.product, "conditions": selected.conditions,
            "first_year_rate": selected.first_year_rate, "total_rate": selected.total_rate,
            "sheet_name": selected.sheet_name, "row_number": selected.row_number,
            "rate_recheck_required": False,
        })
        updated_count += 1
    return updated_count, unresolved_count


def _contract_data(holding: dict, product: ProductRate, recruiter_type: str = "") -> dict:
    return {
        "customer": holding.get("customer", ""),
        "collector": holding.get("collector", ""),
        "policy_number": holding.get("policy_number", ""),
        "insurer": product.insurer,
        "product": product.product,
        "conditions": product.conditions,
        "premium": int(holding.get("premium", 0)),
        "payment_label": holding.get("payment_label", ""),
        "share_rate": float(holding.get("share_rate", 100.0)),
        "recruiter_type": recruiter_type,
        "contract_date": holding.get("contract_date", ""),
        "status": holding.get("status", ""),
        "source_type": product.source_type,
        "first_year_rate": product.first_year_rate,
        "total_rate": product.total_rate,
        "sheet_name": product.sheet_name,
        "row_number": product.row_number,
    }


def _collector_label(
    contracts: list[dict], fallback_collectors: list[str] | None = None
) -> str:
    collectors: list[str] = []
    for contract in contracts:
        name = _clean_text(contract.get("collector", ""))
        if name and name not in collectors:
            collectors.append(name)
    if not collectors:
        for name in fallback_collectors or []:
            cleaned = _clean_text(name)
            if cleaned and cleaned not in collectors:
                collectors.append(cleaned)
    if not collectors:
        return ""
    label = collectors[0]
    if len(collectors) > 1:
        label += f" 외 {len(collectors) - 1}명"
    return re.sub(r'[\\/:*?"<>|]+', "", label).strip(" ._")


def _commission_download_filename(
    contracts: list[dict],
    fallback_collectors: list[str] | None = None,
    fallback_months: list[str] | None = None,
) -> str:
    """수금자명과 계약일의 연월을 사용해 Windows에서도 안전한 파일명을 만듭니다."""
    collector_label = _collector_label(contracts, fallback_collectors)
    months = sorted({
        match.group(1)
        for contract in contracts
        if (match := re.match(r"(20\d{2}-\d{2})", _clean_text(contract.get("contract_date", ""))))
    })
    if not months:
        for month in fallback_months or []:
            cleaned = _clean_text(month)
            if re.fullmatch(r"20\d{2}-\d{2}", cleaned) and cleaned not in months:
                months.append(cleaned)
    months.sort()
    if collector_label:
        base_name = f"{collector_label}_수수료 계산 결과"
    else:
        base_name = "수수료 계산 결과"
    if len(months) == 1:
        year, month = months[0].split("-")
        base_name += f"_{year}년 {month}월"
    elif len(months) > 1:
        start_year, start_month = months[0].split("-")
        end_year, end_month = months[-1].split("-")
        base_name += f"_{start_year}년 {start_month}월-{end_year}년 {end_month}월"
    return f"{base_name}.xlsx"


def _holding_caption(holding: dict) -> str:
    policy = holding.get("policy_number") or "증권번호 없음"
    payment = f" · {holding['payment_label']}" if holding.get("payment_label") else ""
    return f"증권번호 {policy} · 월보험료 {_format_won(holding['premium'])}{payment}"


def _markdown_text(value: Any) -> str:
    """마스킹 이름의 ** 등이 Markdown 문법으로 해석되지 않도록 처리합니다."""
    return str(value or "").replace("\\", "\\\\").replace("*", "\\*").replace("_", "\\_")


def _product_groups(products: list[ProductRate]) -> dict[str, list[ProductRate]]:
    groups: dict[str, list[ProductRate]] = defaultdict(list)
    for product in products:
        display_name, _ = _product_display_parts(product.product)
        groups[display_name].append(product)
    return {name: _sort_condition_candidates(items) for name, items in groups.items()}


def _direct_product_names(holding: dict, groups: dict[str, list[ProductRate]]) -> list[str]:
    """직접 찾기에서도 관련 상품을 위에 두되 같은 보험사의 전체 상품을 검색할 수 있습니다."""
    source_name = _smart_product_name(holding.get("product_raw", ""))
    source_categories = _product_categories(holding.get("product_raw", ""))

    def sort_key(name: str) -> tuple:
        target_categories = _product_categories(name)
        category_rank = 0 if source_categories and source_categories & target_categories else 1
        conflict_rank = 1 if _hard_product_conflict(holding.get("product_raw", ""), name) else 0
        similarity = _name_similarity(source_name, _smart_product_name(name))
        return conflict_rank, category_rank, -similarity, _normalize(name)

    return sorted(groups, key=sort_key)


def _render_smart_product_picker(
    holding: dict,
    recommended: list[ProductRate],
    insurer_products: list[ProductRate],
    payout_rate: float,
    key_prefix: str,
) -> ProductRate | None:
    """추천은 최대 3개만, 직접 찾기는 동일 보험사의 전체 상품을 검색하도록 분리합니다."""
    recommended_groups = _product_groups(recommended)
    direct_groups = _product_groups(insurer_products)
    modes = ["추천 상품"] if recommended_groups else []
    modes.append("직접 찾기")
    mode = st.radio(
        "상품 연결 방법",
        modes,
        horizontal=True,
        key=f"{key_prefix}_mode",
        help="추천이 맞지 않으면 직접 찾기에서 같은 보험사의 전체 상품을 검색할 수 있습니다.",
    )
    groups = recommended_groups if mode == "추천 상품" else direct_groups
    if mode == "추천 상품":
        product_names = list(groups)[:3]
        label = f"추천 상품 · {len(product_names)}개"
        placeholder = "가장 적합한 상품을 선택해 주세요."
    else:
        product_names = _direct_product_names(holding, groups)
        label = f"{holding.get('insurer') or '보험사'} 상품 직접 찾기"
        placeholder = "상품명을 입력하거나 목록에서 선택해 주세요."

    selected_name = st.selectbox(
        label,
        product_names,
        index=0 if mode == "추천 상품" and len(product_names) == 1 else None,
        placeholder=placeholder,
        key=f"{key_prefix}_{'recommended' if mode == '추천 상품' else 'direct'}_product",
    ) if product_names else None
    all_condition_candidates = groups.get(selected_name, []) if selected_name else []
    payment_years = holding.get("payment_years")
    matched_payment_candidates = [
        product for product in all_condition_candidates
        if _payment_matches(product, payment_years)
    ] if payment_years is not None else all_condition_candidates
    condition_candidates = matched_payment_candidates
    condition_key = hashlib.sha1(str(selected_name).encode("utf-8")).hexdigest()[:10]
    if selected_name and payment_years is not None and not matched_payment_candidates:
        if len(all_condition_candidates) == 1:
            only_condition = all_condition_candidates[0]
            st.warning(
                "⚠️ 납입기간 불일치\n\n"
                f"원본 {payment_years}년납 · "
                f"수수료표 {_payment_condition_label(only_condition)}\n\n"
                "적용할 조건을 직접 확인해 주세요."
            )
            return st.selectbox(
                "납입기간 및 세부 조건 · 직접 확인",
                all_condition_candidates,
                index=None,
                placeholder="불일치 내용을 확인한 후 적용할 조건을 선택해 주세요.",
                format_func=lambda product: _condition_option_label(
                    product, payout_rate, all_condition_candidates
                ),
                key=f"{key_prefix}_condition_mismatch_{condition_key}",
            )
        st.warning(
            "⚠️ 납입기간 확인 필요\n\n"
            f"원본 {payment_years}년납과 일치하는 조건이 없습니다. "
            "다른 납기 조건을 확인해 주세요."
        )
        show_other_payment = st.checkbox(
            "다른 납기 조건 보기",
            value=False,
            key=f"{key_prefix}_other_payment_{condition_key}",
        )
        condition_candidates = all_condition_candidates if show_other_payment else []
        if not show_other_payment:
            st.selectbox(
                "납입기간 및 세부 조건",
                [f"{payment_years}년납 조건 없음 · 다른 납기 조건 보기를 선택해 주세요."],
                disabled=True,
                key=f"{key_prefix}_condition_payment_wait_{condition_key}",
            )
            return None
    if condition_candidates:
        label_prefix = f"원본 {payment_years}년납과 일치" if payment_years is not None and matched_payment_candidates else "납입기간 및 세부 조건"
        return st.selectbox(
            f"{label_prefix} · {len(condition_candidates)}개",
            condition_candidates,
            index=0 if len(condition_candidates) == 1 else None,
            placeholder="납입기간과 세부 조건을 선택해 주세요.",
            format_func=lambda product: _condition_option_label(
                product, payout_rate, condition_candidates
            ),
            key=f"{key_prefix}_condition_{condition_key}",
        )
    st.selectbox(
        "납입기간 및 세부 조건",
        ["상품을 선택하면 해당 상품의 조건만 표시됩니다."],
        disabled=True,
        key=f"{key_prefix}_condition_wait_{mode}",
    )
    return None


def _render_manual_entry(all_products: list[ProductRate]) -> None:
    with st.expander("계약 직접 추가", expanded=False):
        if not all_products:
            st.info("생보 또는 손보 수수료 예시표를 먼저 올려 주세요.")
            return
        source_options = [source for source in ("생보", "손보") if any(p.source_type == source for p in all_products)]
        source_type = st.radio("보험 구분", source_options, horizontal=True, key="manual_source_type")
        insurers = sorted({p.insurer for p in all_products if p.source_type == source_type})
        insurer = st.selectbox("보험회사", insurers, index=None, key="manual_insurer")
        products = [p for p in all_products if p.source_type == source_type and p.insurer == insurer]
        product_groups: dict[str, list[ProductRate]] = defaultdict(list)
        for product in products:
            display_name, _ = _product_display_parts(product.product)
            product_groups[display_name].append(product)
        product_names = sorted(product_groups)
        product_name = st.selectbox("상품", product_names, index=None, key="manual_product")
        candidates = _sort_condition_candidates(product_groups.get(product_name, []))
        if candidates:
            selected = st.selectbox(
                "납입기간 및 세부 조건", candidates, index=None,
                format_func=lambda p: _condition_option_label(p, candidates=candidates),
                key="manual_condition",
                placeholder="납입기간과 세부 조건을 선택해 주세요.",
            )
        else:
            selected = None
            st.selectbox(
                "납입기간 및 세부 조건",
                ["먼저 상품을 선택해 주세요."],
                disabled=True,
                key="manual_condition_disabled",
            )
        col1, col2 = st.columns(2)
        customer = col1.text_input("고객명", key="manual_customer")
        policy = col2.text_input("증권번호", key="manual_policy")
        premium = st.number_input("월보험료", min_value=0, step=1000, value=0, format="%d", key="manual_premium")
        if st.button("직접 입력 계약 추가", type="primary", use_container_width=True):
            if selected is None or premium <= 0:
                st.warning("보험회사·상품·세부 조건과 월보험료를 확인해 주세요.")
            else:
                holding = {
                    "customer": customer.strip(), "policy_number": policy.strip(), "premium": int(premium),
                    "share_rate": 100.0, "contract_date": "", "status": "직접 등록",
                }
                st.session_state["commission_contracts"].append(_contract_data(holding, selected))
                st.rerun()


def _render_contract_editor(all_products: list[ProductRate]) -> None:
    edit_index = st.session_state.get("commission_edit_index")
    contracts = st.session_state["commission_contracts"]
    if not isinstance(edit_index, int) or not (0 <= edit_index < len(contracts)):
        return

    contract = contracts[edit_index]
    st.markdown(
        '<div id="commission-edit-anchor" style="scroll-margin-top:5rem;"></div>',
        unsafe_allow_html=True,
    )
    components.html(
        """
        <script>
        setTimeout(function () {
            const target = window.parent.document.getElementById('commission-edit-anchor');
            if (target) {
                target.scrollIntoView({ behavior: 'smooth', block: 'start' });
            }
        }, 180);
        </script>
        """,
        height=0,
    )
    st.info(f"{edit_index + 1}번 계약을 수정하고 있습니다.")
    with st.container(border=True):
        customer_col, policy_col = st.columns(2)
        customer = customer_col.text_input(
            "고객명", value=contract.get("customer", ""), key=f"edit_customer_{edit_index}"
        )
        policy_number = policy_col.text_input(
            "증권번호", value=contract.get("policy_number", ""), key=f"edit_policy_{edit_index}"
        )
        premium_col, share_col = st.columns(2)
        premium = premium_col.number_input(
            "월보험료", min_value=0, step=1000, value=int(contract.get("premium", 0)),
            format="%d", key=f"edit_premium_{edit_index}",
        )
        share_rate = share_col.number_input(
            "쉐어율 (%)", min_value=0.0, max_value=100.0,
            value=float(contract.get("share_rate", 100.0)), step=1.0, format="%.0f",
            key=f"edit_share_{edit_index}",
        )

        source_options = [source for source in ("생보", "손보") if any(p.source_type == source for p in all_products)]
        current_source = contract.get("source_type")
        source_index = source_options.index(current_source) if current_source in source_options else 0
        selected_source = st.radio(
            "보험 구분", source_options, index=source_index, horizontal=True,
            key=f"edit_source_{edit_index}",
        )
        insurer_options = sorted({p.insurer for p in all_products if p.source_type == selected_source})
        current_insurer = contract.get("insurer")
        insurer_index = insurer_options.index(current_insurer) if current_insurer in insurer_options else None
        selected_insurer = st.selectbox(
            "보험회사", insurer_options, index=insurer_index,
            placeholder="보험회사를 선택해 주세요.", key=f"edit_insurer_{edit_index}",
        )

        insurer_products = [
            product for product in all_products
            if product.source_type == selected_source and product.insurer == selected_insurer
        ]
        product_groups: dict[str, list[ProductRate]] = defaultdict(list)
        for product in insurer_products:
            display_name, _ = _product_display_parts(product.product)
            product_groups[display_name].append(product)
        product_names = sorted(product_groups)
        current_product_name, _ = _product_display_parts(contract.get("product", ""))
        product_index = product_names.index(current_product_name) if current_product_name in product_names else None
        selected_product_name = st.selectbox(
            "상품", product_names, index=product_index,
            placeholder="상품을 선택해 주세요.", key=f"edit_product_{edit_index}_{selected_insurer}",
        ) if product_names else None

        matching = _sort_condition_candidates(product_groups.get(selected_product_name, []))
        current_position = next(
            (
                index for index, product in enumerate(matching)
                if product.sheet_name == contract.get("sheet_name") and product.row_number == contract.get("row_number")
            ),
            None,
        )
        if matching:
            selected_product = st.selectbox(
                "납입기간 및 세부 조건", matching,
                index=current_position,
                placeholder="납입기간과 세부 조건을 선택해 주세요.",
                format_func=lambda p: _condition_option_label(p, candidates=matching),
                key=f"edit_condition_{edit_index}_{hashlib.sha1(str(selected_product_name).encode()).hexdigest()[:8]}",
            )
        else:
            selected_product = None
            st.selectbox(
                "납입기간 및 세부 조건",
                ["먼저 상품을 선택해 주세요."],
                disabled=True,
                key=f"edit_condition_disabled_{edit_index}_{selected_insurer}",
            )

        recruiter_type = contract.get("recruiter_type", "")
        if share_rate < 100:
            recruiter_options = ["주모집", "공동모집"]
            recruiter_index = recruiter_options.index(recruiter_type) if recruiter_type in recruiter_options else None
            recruiter_type = st.selectbox(
                "모집 형태", recruiter_options, index=recruiter_index,
                placeholder="모집 형태를 선택해 주세요.", key=f"edit_recruiter_{edit_index}",
            ) or ""
        else:
            recruiter_type = ""

        save_col, cancel_col = st.columns([3, 1])
        if save_col.button("수정 완료", type="primary", use_container_width=True, key=f"save_edit_{edit_index}"):
            if premium <= 0:
                st.warning("월보험료를 확인해 주세요.")
            elif share_rate < 100 and not recruiter_type:
                st.warning("모집 형태를 선택해 주세요.")
            else:
                updated = dict(contract)
                updated.update({
                    "customer": customer.strip(), "policy_number": policy_number.strip(),
                    "premium": int(premium), "share_rate": float(share_rate),
                    "recruiter_type": recruiter_type,
                })
                if selected_product is not None:
                    updated.update({
                        "insurer": selected_product.insurer,
                        "product": selected_product.product,
                        "conditions": selected_product.conditions,
                        "source_type": selected_product.source_type,
                        "first_year_rate": selected_product.first_year_rate,
                        "total_rate": selected_product.total_rate,
                        "sheet_name": selected_product.sheet_name,
                        "row_number": selected_product.row_number,
                        "rate_recheck_required": False,
                    })
                contracts[edit_index] = updated
                st.session_state["commission_edit_index"] = None
                st.rerun()
        if cancel_col.button("취소", use_container_width=True, key=f"cancel_edit_{edit_index}"):
            st.session_state["commission_edit_index"] = None
            st.rerun()


def run() -> None:
    _initialize_state()

    page_header(
        "실적 관리",
        "수수료 계산기",
        "수수료 예시표와 보유계약 장기 파일을 연결해 계약별 예상 수당을 계산합니다.",
        "CC",
    )
    section_intro("입력", "수수료 자료 불러오기", "생보·손보 수수료 예시표를 먼저 등록해 주세요.")

    with st.expander("① 수수료 예시표 불러오기", expanded=True):
        life_file = st.file_uploader(
            "생보 수수료 예시표", type=["xlsx"], key="commission_life_file"
        )
        nonlife_file = st.file_uploader(
            "손보 수수료 예시표", type=["xlsx"], key="commission_nonlife_file"
        )

    all_products: list[ProductRate] = []
    parse_warnings: list[str] = []
    reference_months: dict[str, str] = {}
    ratebook_hash = hashlib.sha256()
    for uploaded, source_type in ((life_file, "생보"), (nonlife_file, "손보")):
        if uploaded is None:
            continue
        try:
            uploaded_bytes = uploaded.getvalue()
            ratebook_hash.update(source_type.encode("utf-8"))
            ratebook_hash.update(uploaded_bytes)
            parsed, warnings = parse_commission_workbook(uploaded_bytes, source_type)
            all_products.extend(_to_product_rate(item) for item in parsed)
            parse_warnings.extend(warnings)
            reference_months[source_type] = _month_from_filename(uploaded.name)
        except Exception as exc:
            st.error(f"{source_type} 예시표를 읽지 못했습니다: {exc}")

    if all_products:
        insurer_count = len({product.insurer for product in all_products})
        month_text = " · ".join(
            f"{source} {month.replace('-', '년 ')}월" if month else f"{source} 기준월 확인 필요"
            for source, month in reference_months.items()
        )
        st.success(f"{month_text} · 보험회사 {insurer_count}개 · 수수료 조건 {len(all_products):,}개")
    else:
        st.info("생보 또는 손보 수수료 예시표를 올리면 상품을 선택할 수 있습니다.")

    for warning in parse_warnings:
        st.warning(warning)

    current_ratebook_signature = ratebook_hash.hexdigest() if all_products else ""
    saved_ratebook_signature = st.session_state.get("commission_ratebook_signature", "")
    contracts = st.session_state["commission_contracts"]
    if current_ratebook_signature and not saved_ratebook_signature:
        st.session_state["commission_ratebook_signature"] = current_ratebook_signature
    elif (
        current_ratebook_signature and saved_ratebook_signature
        and current_ratebook_signature != saved_ratebook_signature and contracts
    ):
        st.warning(
            "수수료 예시표가 변경되었습니다. 기존 수당 계산 대상 계약의 요율을 "
            "새 수수료표 기준으로 재검증해야 합니다."
        )
        reconnect_col, clear_col = st.columns(2)
        if reconnect_col.button("새 수수료표로 다시 연결", type="primary", use_container_width=True):
            updated, unresolved = _reconnect_contract_rates(contracts, all_products)
            st.session_state["commission_ratebook_signature"] = current_ratebook_signature
            st.session_state["commission_edit_index"] = None
            st.toast(f"{updated}건 재연결 · {unresolved}건 직접 확인 필요")
            st.rerun()
        if clear_col.button("기존 계약 초기화", use_container_width=True):
            st.session_state["commission_contracts"] = []
            st.session_state["commission_ratebook_signature"] = current_ratebook_signature
            st.session_state["commission_edit_index"] = None
            st.rerun()
        st.stop()
    elif current_ratebook_signature and current_ratebook_signature != saved_ratebook_signature:
        st.session_state["commission_ratebook_signature"] = current_ratebook_signature

    section_intro("입력", "지급율 및 보유계약 불러오기", "공통 지급율을 확인하고 보유계약관리 장기 엑셀을 등록해 주세요.")
    payout_rate_percent = st.number_input(
        "공통 지급율 (%)",
        min_value=0.0,
        max_value=100.0,
        value=float(round(st.session_state["commission_payout_rate"])),
        step=1.0,
        format="%.0f",
        help="변경한 지급율은 현재 수당 계산 대상 계약에 일괄 적용됩니다.",
    )
    st.session_state["commission_payout_rate"] = payout_rate_percent
    payout_rate = payout_rate_percent / 100
    holding_file = st.file_uploader(
        "보유계약관리 장기 엑셀", type=["xlsx"], key="commission_holding_file",
        help="계약상태가 정상이고 수수료표 기준월과 같은 계약을 우선 분석합니다.",
    )

    # 선택할 때마다 전체 16,000여 조건을 다시 비교하지 않도록 보험사별로 미리 나눕니다.
    products_by_insurer: dict[tuple[str, str], list[ProductRate]] = defaultdict(list)
    for product in all_products:
        products_by_insurer[(product.source_type, product.insurer)].append(product)

    review_records: list[dict] = []
    if holding_file is not None and all_products:
        try:
            holdings = parse_holding_workbook(holding_file.getvalue())
        except Exception as exc:
            holdings = []
            st.error(f"보유계약 파일을 읽지 못했습니다: {exc}")

        st.session_state["commission_import_collectors"] = list(dict.fromkeys(
            _clean_text(holding.get("collector", ""))
            for holding in holdings if _clean_text(holding.get("collector", ""))
        ))
        st.session_state["commission_import_contract_months"] = sorted({
            _clean_text(holding.get("contract_month", ""))
            for holding in holdings
            if re.fullmatch(r"20\d{2}-\d{2}", _clean_text(holding.get("contract_month", "")))
        })

        product_by_key = {product.key: product for product in all_products}
        product_rows = [product.__dict__ for product in all_products]
        link_decisions = _analyze_product_links(holdings, product_rows)
        registered_policies = {c.get("policy_number") for c in st.session_state["commission_contracts"] if c.get("policy_number")}
        automatic: list[tuple[dict, ProductRate]] = []
        needs_review: list[tuple[dict, list[ProductRate], str]] = []
        excluded: list[tuple[dict, str]] = []
        unmatched: list[tuple[dict, str]] = []
        already_registered = 0

        for holding in holdings:
            insurer_products = products_by_insurer.get(
                (holding.get("source_type", ""), holding.get("insurer", "")), []
            )
            ref_month = reference_months.get(holding["source_type"], "")
            if holding.get("policy_number") and holding["policy_number"] in registered_policies:
                already_registered += 1
                continue
            if holding.get("status") != "정상":
                excluded.append((holding, f"계약상태가 {holding.get('status') or '확인 필요'}이므로 기본 제외"))
                continue
            if ref_month and holding.get("contract_month") and holding["contract_month"] != ref_month:
                excluded.append((holding, f"계약월 {holding['contract_month']} / 수수료표 기준월 {ref_month}"))
                continue
            decision = link_decisions.get(holding["row_key"], {})
            candidates = [
                product_by_key[key] for key in decision.get("candidate_keys", [])
                if key in product_by_key
            ]
            if not candidates:
                unmatched.append((holding, "수수료표에서 일치하는 상품을 찾지 못함"))
                continue
            auto = product_by_key.get(decision.get("auto_key", ""))
            if auto is not None and holding.get("share_rate", 100.0) >= 100:
                automatic.append((holding, auto))
            else:
                reason_parts = []
                if auto is None:
                    reason_parts.append("세부 조건 확인")
                if holding.get("share_rate", 100.0) < 100:
                    reason_parts.append("모집 형태 확인")
                review_candidates = [
                    product_by_key[key] for key in decision.get("review_keys", [])
                    if key in product_by_key
                ]
                needs_review.append((holding, review_candidates, " · ".join(reason_parts)))

        section_intro("연결 결과", "자동 연결 및 확인 필요 계약", "자동 연결 결과를 검토하고 필요한 계약만 조건을 다시 확인해 주세요.")
        metric_cols = st.columns(4)
        metric_cols[0].metric("전체", f"{len(holdings)}건")
        metric_cols[1].metric("자동 연결", f"{len(automatic)}건")
        metric_cols[2].metric("확인 필요", f"{len(needs_review)}건")
        metric_cols[3].metric("미연결·제외", f"{len(unmatched) + len(excluded)}건")
        if already_registered:
            st.caption(f"이미 등록된 증권번호 {already_registered}건은 중복 분석에서 제외했습니다.")

        pending: list[dict] = []
        with st.expander(f"자동 연결 완료 {len(automatic)}건", expanded=True):
            if not automatic:
                st.caption("자동 연결된 계약이 없습니다.")
            for holding, product in automatic:
                col1, col2 = st.columns([0.08, 0.92])
                selected = col1.checkbox("선택", value=True, key=f"auto_{holding['row_key']}", label_visibility="collapsed")
                selected_product = product
                with col2:
                    customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                    st.markdown(f"**{customer_name} · {product.insurer}**")
                    st.caption(_holding_caption(holding))
                    st.write(f"{product.product} · {product.conditions or '기본 조건'}")
                    reason = "상품명 일치"
                    if holding.get("payment_label"):
                        reason += f" · {holding['payment_label']} 조건 일치"
                    _, _, tag_reasons = _tag_match_summary(holding, product)
                    if tag_reasons:
                        reason += " · " + " · ".join(dict.fromkeys(tag_reasons))
                    st.caption(f"자동 연결 근거: {reason}")
                    verify_auto = st.checkbox(
                        "상품·납기 다시 확인",
                        value=False,
                        key=f"auto_verify_{holding['row_key']}",
                    )
                    if verify_auto:
                        insurer_products = products_by_insurer.get(
                            (holding.get("source_type", ""), holding.get("insurer", "")), []
                        )
                        decision = link_decisions.get(holding["row_key"], {})
                        recommended = [
                            product_by_key[key] for key in decision.get("review_keys", [])
                            if key in product_by_key
                        ]
                        selected_product = _render_smart_product_picker(
                            holding,
                            recommended or [product],
                            insurer_products,
                            payout_rate,
                            key_prefix=f"auto_change_{holding['row_key']}",
                        )
                if selected and selected_product is not None:
                    pending.append(_contract_data(holding, selected_product))
                elif selected and verify_auto:
                    st.caption("변경할 상품과 원본 납기에 맞는 조건을 선택해 주세요.")

        with st.expander(f"확인 필요 {len(needs_review)}건", expanded=bool(needs_review)):
            if not needs_review:
                st.caption("확인이 필요한 계약이 없습니다.")
            for holding, candidates, reason in needs_review:
                with st.container(border=True):
                    customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                    st.markdown(f"**{customer_name} · {holding['insurer']}**")
                    st.caption(f"{_holding_caption(holding)} · {reason}")
                    st.write(f"보유계약 상품: {holding['product_raw']}")
                    insurer_products = products_by_insurer.get(
                        (holding.get("source_type", ""), holding.get("insurer", "")), []
                    )
                    selected_product = _render_smart_product_picker(
                        holding, candidates, insurer_products, payout_rate,
                        key_prefix=f"review_{holding['row_key']}",
                    )
                recruiter_type = ""
                if holding.get("share_rate", 100.0) < 100:
                    recruiter_type = st.selectbox(
                        f"모집 형태 · 엑셀 쉐어율 {holding['share_rate']:g}%",
                        ["주모집", "공동모집"], index=None, key=f"recruiter_{holding['row_key']}",
                        placeholder="모집 형태를 선택해 주세요.",
                    ) or ""
                include = st.checkbox("이 계약 등록", value=True, key=f"review_include_{holding['row_key']}")
                ready = include and selected_product is not None and (
                    holding.get("share_rate", 100.0) >= 100 or recruiter_type
                )
                if ready:
                    pending.append(_contract_data(holding, selected_product, recruiter_type))
                    st.success("등록 준비 완료")
                elif not include:
                    review_records.append({**holding, "product": holding["product_raw"], "reason": "사용자가 등록 대상에서 제외"})
                else:
                    st.caption("상품·납입기간·모집 형태 중 필요한 항목을 선택해 주세요.")
                st.write("")

        if excluded:
            with st.expander(f"기준월·계약상태·중복으로 제외 {len(excluded)}건", expanded=False):
                st.info("기본적으로 제외됩니다. 필요한 경우에만 계약을 펼쳐 포함해 주세요.")
                for holding, reason in excluded:
                    customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                    st.markdown(f"**{customer_name} · {holding['insurer']}**")
                    st.caption(f"{_holding_caption(holding)} · {reason}")
                    include = st.checkbox("이번 계산에 포함", value=False, key=f"excluded_include_{holding['row_key']}")
                    if include:
                        insurer_products = products_by_insurer.get(
                            (holding.get("source_type", ""), holding.get("insurer", "")), []
                        )
                        candidates = _candidate_products(holding, insurer_products)
                        selected_product = st.selectbox(
                            "적용할 상품 및 조건", candidates, index=None, key=f"excluded_product_{holding['row_key']}",
                            format_func=lambda p: (
                                f"{p.product} · "
                                f"{_condition_option_label(p, payout_rate, candidates)}"
                            ),
                        ) if candidates else None
                        confirmed = st.checkbox(
                            "제외 사유를 확인했으며 이번 계산에 포함합니다.",
                            key=f"excluded_confirm_{holding['row_key']}",
                        )
                        if selected_product is not None and confirmed:
                            pending.append(_contract_data(holding, selected_product))
                        else:
                            review_records.append({**holding, "product": holding["product_raw"], "reason": reason})
                    else:
                        review_records.append({**holding, "product": holding["product_raw"], "reason": reason})
                    st.divider()

        if unmatched:
            with st.expander(f"연결되지 않은 계약 {len(unmatched)}건", expanded=bool(unmatched)):
                for holding, reason in unmatched:
                    customer_name = _markdown_text(holding.get("customer") or "고객명 없음")
                    st.markdown(f"**{customer_name} · {holding['insurer']}**")
                    st.caption(_holding_caption(holding))
                    st.write(f"{holding['product_raw']} · {reason}")
                    with st.container(border=True):
                        insurer_products = products_by_insurer.get(
                            (holding.get("source_type", ""), holding.get("insurer", "")), []
                        )
                        if not insurer_products:
                            source_products = [
                                product for product in all_products
                                if product.source_type == holding.get("source_type")
                            ]
                            insurer_options = sorted({product.insurer for product in source_products})
                            direct_insurer = st.selectbox(
                                "보험회사를 찾지 못했습니다 · 직접 선택",
                                insurer_options,
                                index=None,
                                placeholder="보험회사를 선택해 주세요.",
                                key=f"unmatched_insurer_{holding['row_key']}",
                            )
                            insurer_products = [
                                product for product in source_products
                                if product.insurer == direct_insurer
                            ]
                        direct_product = _render_smart_product_picker(
                            holding, [], insurer_products, payout_rate,
                            key_prefix=f"unmatched_{holding['row_key']}",
                        )
                        if direct_product is not None:
                            pending.append(_contract_data(holding, direct_product))
                            st.success("직접 연결 준비 완료")
                        else:
                            review_records.append({
                                **holding, "product": holding["product_raw"], "reason": reason
                            })
                    st.divider()

        if pending:
            if st.button(f"선택한 계약 {len(pending)}건 등록", type="primary", use_container_width=True):
                existing = {c.get("policy_number") for c in st.session_state["commission_contracts"] if c.get("policy_number")}
                added = 0
                for contract in pending:
                    if contract.get("policy_number") and contract["policy_number"] in existing:
                        continue
                    st.session_state["commission_contracts"].append(contract)
                    if contract.get("policy_number"):
                        existing.add(contract["policy_number"])
                    added += 1
                st.toast(f"계약 {added}건을 등록했습니다.")
                st.rerun()
        elif holdings:
            st.info("현재 등록할 수 있는 계약이 없습니다. 확인 필요 계약의 조건을 선택해 주세요.")

    section_intro("직접 입력", "계약 직접 추가", "파일에 없는 계약은 보험회사와 상품 조건을 직접 선택해 추가할 수 있습니다.")
    _render_manual_entry(all_products)

    contracts = st.session_state["commission_contracts"]
    section_intro("계산 결과", "수당 계산 대상 계약", "등록된 계약과 예상 익월수당·총수당을 확인해 주세요.")
    if not contracts:
        st.info("추가된 계약이 없습니다.")
        return

    _render_contract_editor(all_products)

    calculation_contracts = [contract for contract in contracts if not contract.get("rate_recheck_required")]
    recheck_count = len(contracts) - len(calculation_contracts)
    if recheck_count:
        st.warning(
            f"새 수수료표에서 요율을 확정하지 못한 계약 {recheck_count}건은 합계와 다운로드에서 제외했습니다. "
            "해당 계약의 수정 버튼을 눌러 상품과 세부 조건을 다시 선택해 주세요."
        )

    total_premium = sum(contract["premium"] for contract in calculation_contracts)
    total_first = sum(
        contract["premium"] * contract["first_year_rate"] * payout_rate
        for contract in calculation_contracts
    )
    total_commission = sum(
        contract["premium"] * contract["total_rate"] * payout_rate
        for contract in calculation_contracts
    )
    metric_cols = st.columns(3)
    metric_cols[0].metric("월보험료 합계", _format_won(total_premium))
    metric_cols[1].metric("예상 익월수당", _format_won(total_first))
    metric_cols[2].metric("예상 총수당", _format_won(total_commission))

    header_columns = st.columns([3.6, 1, 1.15, 1.15, 1.25, 1.25, 1.05])
    for column, label in zip(
        header_columns, ("계약 정보", "월보험료", "익월 수수료율", "총 수수료율", "예상 익월수당", "예상 총수당", "관리"),
    ):
        column.caption(label)

    for index, contract in enumerate(contracts):
        first_rate = contract["first_year_rate"] * payout_rate
        total_rate = contract["total_rate"] * payout_rate
        expected_first = contract["premium"] * first_rate
        expected_total = contract["premium"] * total_rate
        product_detail = _compact_product_display(contract, contracts)

        row_columns = st.columns([3.6, 1, 1.15, 1.15, 1.25, 1.25, 1.05])
        with row_columns[0]:
            customer_name = _markdown_text(contract.get("customer") or "고객명 없음")
            st.markdown(f"**{index + 1}. {customer_name}** · {contract['insurer']}")
            policy = contract.get("policy_number") or "증권번호 없음"
            recruiting = ""
            if contract.get("share_rate", 100) < 100:
                recruiting = f" · {contract['share_rate']:g}% · {contract.get('recruiter_type') or '모집 형태 확인'}"
            st.caption(f"증권번호 {policy}{recruiting}")
            st.markdown(_markdown_text(product_detail).replace("\n", "  \n"))
            if contract.get("rate_recheck_required"):
                st.warning("요율 재확인 필요")
        row_columns[1].write(_format_won(contract["premium"]))
        row_columns[2].write(_format_rate(first_rate))
        row_columns[3].write(_format_rate(total_rate))
        row_columns[4].write(_format_won(expected_first))
        row_columns[5].write(_format_won(expected_total))
        with row_columns[6]:
            edit_col, delete_col = st.columns(2)
            if edit_col.button("수정", key=f"edit_commission_{index}", help="이 계약 수정"):
                st.session_state["commission_edit_index"] = index
                st.rerun()
            if delete_col.button("✕", key=f"delete_commission_{index}", help="이 계약 삭제"):
                contracts.pop(index)
                current_edit = st.session_state.get("commission_edit_index")
                if current_edit == index:
                    st.session_state["commission_edit_index"] = None
                elif isinstance(current_edit, int) and current_edit > index:
                    st.session_state["commission_edit_index"] = current_edit - 1
                st.rerun()

        if index < len(contracts) - 1:
            st.markdown(
                '<hr style="margin:.25rem 0 .45rem;border:0;border-top:1px solid rgba(128,128,128,.18);">',
                unsafe_allow_html=True,
            )

    section_intro("다운로드", "계산 결과 내려받기", "확인된 계약과 수수료 계산 결과를 엑셀로 저장합니다.")
    clear_col, download_col = st.columns([1, 2])
    with clear_col:
        if st.button("전체 계약 지우기", use_container_width=True):
            st.session_state["commission_contracts"] = []
            st.session_state["commission_edit_index"] = None
            st.rerun()
    with download_col:
        months = sorted({month for month in reference_months.values() if month})
        reference_month = ", ".join(months)
        excel_bytes = _make_excel(
            calculation_contracts,
            payout_rate,
            reference_month,
            review_records,
            st.session_state.get("commission_import_collectors", []),
        )
        st.download_button(
            "엑셀 다운로드",
            data=excel_bytes,
            file_name=_commission_download_filename(
                calculation_contracts,
                st.session_state.get("commission_import_collectors", []),
                st.session_state.get("commission_import_contract_months", []),
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
