from __future__ import annotations

import re
import html
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

try:
    from .ui_components import page_header, section_intro
except ImportError:  # 단독 실행·테스트용
    def page_header(_section: str, title: str, description: str, _code: str) -> None:
        st.title(title)
        st.caption(description)


APP_TITLE = "보험 리모델링 비교 제안서"

NAVY = "17365D"
NAVY2 = "245889"
GOLD = "C9A24D"
GOLD_LIGHT = "FFF4D6"
BLUE_LIGHT = "EAF2F8"
GREEN_LIGHT = "E8F4F2"
SOFT = "F7F9FC"
WHITE = "FFFFFF"
INK = "25364A"
MUTED = "667085"
GREEN = "24745A"
RED = "C00000"
LINE = "B8C2CF"
THIN = Side(style="thin", color=LINE)

CONTRACT_ACTIONS = ["유지", "감액", "해지", "변경", "검토"]
ACTION_HELP = {
    "유지": "예: 현재 계약과 보장을 그대로 유지",
    "감액": "예: 고객센터를 통해 불필요한 특약 감액 요청",
    "해지": "예: 고객센터 상담원 연결 후 계약 해지 요청",
    "변경": "예: 갱신형 특약 또는 보장금액 변경 요청",
    "검토": "예: 보험회사에 계약조건 확인 후 처리 방향 결정",
}
ACTION_STYLE = {
    "유지": ("E8F4F2", "24745A"),
    "감액": ("FFF1D6", "9A6700"),
    "해지": ("FDECEC", "B42318"),
    "변경": ("EAF2FF", "1769DC"),
    "검토": ("F2ECFF", "6941C6"),
}


@dataclass
class NewPlan:
    name: str = ""
    monthly: int = 0
    years: int = 20
    custom_months: int = 0

    @property
    def months(self) -> int:
        return self.custom_months if self.custom_months > 0 else self.years * 12

    @property
    def fixed_total(self) -> int:
        return self.monthly * self.months


@dataclass
class ExistingContract:
    company: str = ""
    product: str = ""
    action: str = "유지"
    detail: str = ""


@dataclass
class Person:
    name: str
    old_monthly: int
    old_total: int
    retained_monthly: int
    retained_total: int
    plans: list[NewPlan] = field(default_factory=list)
    coverage: str = ""
    contracts: list[ExistingContract] = field(default_factory=list)

    @property
    def new_plan_monthly(self) -> int:
        return sum(p.monthly for p in self.plans)

    @property
    def after_monthly(self) -> int:
        return self.retained_monthly + self.new_plan_monthly

    @property
    def after_total(self) -> int:
        return self.retained_total + sum(p.fixed_total for p in self.plans)

    @property
    def monthly_change(self) -> int:
        return self.after_monthly - self.old_monthly

    @property
    def total_change(self) -> int:
        return self.after_total - self.old_total

def money(value: object) -> int:
    text = re.sub(r"[^0-9-]", "", str(value or ""))
    try:
        return int(text) if text not in {"", "-"} else 0
    except ValueError:
        return 0


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def safe_filename(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", value).strip() or "보험리모델링_비교안"


def won(value: int) -> str:
    return f"{int(value):,}원"


def rate(old: int, new: int) -> float | None:
    return (old - new) / old if old > new and old else None


def change_amount(old: int, new: int) -> str:
    if old > new:
        return f"{old-new:,}원 절감"
    if old < new:
        return f"{new-old:,}원 증가"
    return "변동 없음"


def change_rate(old: int, new: int) -> str:
    r = rate(old, new)
    return f"{r:.1%} 감소" if r is not None else ""


def combined(people: list[Person]) -> dict[str, int]:
    return {
        "old_monthly": sum(p.old_monthly for p in people),
        "after_monthly": sum(p.after_monthly for p in people),
        "old_total": sum(p.old_total for p in people),
        "after_total": sum(p.after_total for p in people),
    }


def _state_count(key: str, default: int) -> int:
    st.session_state.setdefault(key, default)
    return int(st.session_state[key])


def _money_input(label: str, key: str, help_text: str | None = None) -> int:
    raw = st.text_input(label, key=key, placeholder="예: 694,580", help=help_text)
    value = money(raw)
    if raw:
        st.caption(f"{value:,}원")
    return value


def render_plan_inputs(person_no: int) -> list[NewPlan]:
    count_key = f"rm_plan_count_{person_no}"
    count = _state_count(count_key, 4)
    plans: list[NewPlan] = []
    for i in range(count):
        title = clean(st.session_state.get(f"rm_plan_name_{person_no}_{i}")) or f"신규 보험 {i+1}"
        with st.expander(title, expanded=i < 2):
            c1, c2 = st.columns([2.2, 1])
            with c1:
                name = st.text_input("보험 또는 보장 구성명", key=f"rm_plan_name_{person_no}_{i}", placeholder="예: 암·뇌·심장 진단비")
            with c2:
                premium = st.text_input("월 보험료", key=f"rm_plan_premium_{person_no}_{i}", placeholder="예: 128,589")
            c3, c4 = st.columns([1.3, 1])
            years = 20
            months = 0
            with c3:
                years = st.selectbox("납입기간", [5, 10, 15, 20, 25, 30], index=3, format_func=lambda x: f"{x}년", key=f"rm_plan_years_{person_no}_{i}")
            with c4:
                custom = st.checkbox("개월 수 직접 입력", key=f"rm_plan_custom_on_{person_no}_{i}")
            if custom:
                months = int(st.number_input("직접 입력할 납입 개월 수", min_value=1, max_value=1200, value=240, step=1, key=f"rm_plan_months_{person_no}_{i}"))
            plan = NewPlan(clean(name), money(premium), int(years), int(months))
            plans.append(plan)
            if plan.name and plan.monthly:
                st.caption(f"납입 예정 총액: {plan.fixed_total:,}원")
    a, b = st.columns(2)
    with a:
        if st.button("＋ 신규 보험 추가", key=f"rm_plan_add_{person_no}", use_container_width=True):
            st.session_state[count_key] = min(12, count + 1)
            st.rerun()
    with b:
        if st.button("－ 마지막 신규 보험 삭제", key=f"rm_plan_remove_{person_no}", disabled=count <= 1, use_container_width=True):
            st.session_state[count_key] = max(1, count - 1)
            st.rerun()
    return [p for p in plans if p.name or p.monthly]


def render_contract_inputs(person_no: int) -> list[ExistingContract]:
    count_key = f"rm_contract_count_{person_no}"
    count = _state_count(count_key, 2)
    result: list[ExistingContract] = []
    for i in range(count):
        with st.expander(f"기존 계약 {i+1}", expanded=i == 0):
            a, b = st.columns(2)
            with a:
                company = st.text_input("보험회사", key=f"rm_contract_company_{person_no}_{i}")
                product = st.text_input("상품명", key=f"rm_contract_product_{person_no}_{i}")
            with b:
                action_key = f"rm_contract_action_{person_no}_{i}"
                if st.session_state.get(action_key) not in CONTRACT_ACTIONS:
                    st.session_state[action_key] = "변경"
                action = st.selectbox("처리 방향", CONTRACT_ACTIONS, key=action_key)
                detail = st.text_input(
                    "변경 내용",
                    key=f"rm_contract_detail_{person_no}_{i}",
                    placeholder=ACTION_HELP.get(action, "처리 내용을 구체적으로 입력해 주세요."),
                )
            result.append(ExistingContract(clean(company), clean(product), action, clean(detail)))
    a, b = st.columns(2)
    with a:
        if st.button("＋ 기존 계약 추가", key=f"rm_contract_add_{person_no}", use_container_width=True):
            st.session_state[count_key] = min(12, count + 1)
            st.rerun()
    with b:
        if st.button("－ 마지막 기존 계약 삭제", key=f"rm_contract_remove_{person_no}", disabled=count <= 1, use_container_width=True):
            st.session_state[count_key] = max(1, count - 1)
            st.rerun()
    return [c for c in result if c.company or c.product or c.detail]


def render_person_inputs(person_no: int) -> Person:
    st.subheader(f"고객 {person_no}")
    name = st.text_input("고객명", key=f"rm_name_{person_no}", placeholder="예: 홍길동")
    a, b = st.columns(2)
    with a:
        old_monthly = _money_input("기존 월 보험료", f"rm_old_monthly_{person_no}")
        old_total = _money_input("기존 납입 예정 총액", f"rm_old_total_{person_no}")
    with b:
        retained_monthly = _money_input("유지하는 기존 보험료", f"rm_retained_monthly_{person_no}")
        retained_total = _money_input("유지 보험의 남은 확정 납입 예정 총액", f"rm_retained_total_{person_no}", "사용자가 확인한 합계 금액을 직접 입력합니다.")
    st.markdown("#### 새롭게 가입하는 보험")
    plans = render_plan_inputs(person_no)
    coverage = st.text_area("새롭게 확보되는 핵심 보장", key=f"rm_coverage_{person_no}", placeholder="예: 암·뇌·심장 진단비 보완 · 주요 치료비 강화", height=75)
    st.markdown("#### 기존 계약별 처리 계획")
    contracts = render_contract_inputs(person_no)
    person = Person(clean(name), old_monthly, old_total, retained_monthly, retained_total, plans, clean(coverage), contracts)
    st.info(
        f"신규 보험료 합계 {person.new_plan_monthly:,}원  ·  "
        f"변경 후 월 보험료 {person.after_monthly:,}원  ·  "
        f"변경 후 납입 예정 총액 {person.after_total:,}원"
    )
    return person


# ---------------- Excel ----------------
def _border() -> Border:
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _merge(ws, address: str, value: object, *, fill: str | None = None, color: str = INK,
           size: float = 10, bold: bool = False, border: bool = True) -> None:
    ws.merge_cells(address)
    cell = ws[address.split(":")[0]]
    cell.value = value
    cell.font = Font(name="맑은 고딕", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if border:
        for row in ws[address]:
            for item in row:
                item.border = _border()


def _excel_setup(ws, last_col: str, last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    # 파일을 처음 열었을 때 비교표가 조금 더 크게 보이도록 설정합니다.
    # 화면 확대 비율은 인쇄 배율에는 영향을 주지 않습니다.
    ws.sheet_view.zoomScale = 110
    ws.sheet_view.zoomScaleNormal = 110
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=.25, right=.25, top=.30, bottom=.30, header=.1, footer=.1)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.print_area = f"A1:{last_col}{last_row}"


def _excel_top(ws, people: list[Person], title: str) -> None:
    totals = combined(people)
    _merge(ws, "A1:R2", title, color=NAVY, size=19, bold=True, border=False)
    subtitle = "보험료 부담과 필요한 보장을 함께 비교해 오래 유지하기 쉬운 구조로 재구성했습니다."
    _merge(ws, "A3:R3", subtitle, color=MUTED, size=10, bold=True, border=False)
    labels = (["월 보험료 변화", "변경 후 월 보험료", "납입 예정 총액 변화"] if len(people) == 1 else
              ["합산 월 보험료 변화", "변경 후 합산 월 보험료", "납입 예정 총액 변화"])
    for address, label in zip(("A5:F5", "G5:L5", "M5:R5"), labels):
        _merge(ws, address, label, fill=NAVY2, color=WHITE, size=9, bold=True)
    _merge(ws, "A6:D6", change_amount(totals["old_monthly"], totals["after_monthly"]), fill=GOLD_LIGHT, color=NAVY, size=11, bold=True)
    _merge(ws, "E6:F6", change_rate(totals["old_monthly"], totals["after_monthly"]), fill=GOLD_LIGHT, color=RED, size=11, bold=True)
    _merge(ws, "G6:L6", won(totals["after_monthly"]), fill=GOLD_LIGHT, color=NAVY, size=11, bold=True)
    _merge(ws, "M6:P6", change_amount(totals["old_total"], totals["after_total"]), fill=GOLD_LIGHT, color=NAVY, size=11, bold=True)
    _merge(ws, "Q6:R6", change_rate(totals["old_total"], totals["after_total"]), fill=GOLD_LIGHT, color=RED, size=11, bold=True)


def _excel_person_panel(ws, p: Person, left: int, right: int, top: int, max_plans: int) -> int:
    L, R = get_column_letter(left), get_column_letter(right)
    _merge(ws, f"{L}{top}:{R}{top+1}", f"{p.name or 'OOO'}님", fill=NAVY, color=WHITE, size=13, bold=True)
    mid = (left + right) // 2
    spans = [(left, left+1), (left+2, mid), (mid+1, mid+2), (mid+3, right)]
    rows = [
        ("기존 월 보험료", won(p.old_monthly), "리모델링 후", won(p.after_monthly)),
        ("기존 납입 예정 총액", won(p.old_total), "변경 납입 예정 총액", won(p.after_total)),
    ]
    for rr, values in enumerate(rows, top+2):
        for idx, ((a, b), value) in enumerate(zip(spans, values)):
            _merge(ws, f"{get_column_letter(a)}{rr}:{get_column_letter(b)}{rr}", value,
                   fill=BLUE_LIGHT if idx % 2 == 0 else WHITE, color=NAVY2 if idx == 3 else INK,
                   size=7.5 if idx % 2 == 0 else 9.5, bold=True)
    row = top + 5
    _merge(ws, f"{L}{row}:{R}{row}", "새롭게 가입하는 보험", fill=NAVY2, color=WHITE, bold=True)
    text_end = right - 2
    _merge(ws, f"{L}{row+1}:{get_column_letter(text_end)}{row+1}", "보험 또는 보장 구성", fill=NAVY2, color=WHITE, size=8, bold=True)
    _merge(ws, f"{get_column_letter(text_end+1)}{row+1}:{R}{row+1}", "월 보험료", fill=NAVY2, color=WHITE, size=8, bold=True)
    shown = p.plans[:max_plans]
    for idx in range(max_plans):
        plan = shown[idx] if idx < len(shown) else NewPlan()
        rr = row + 2 + idx
        _merge(ws, f"{L}{rr}:{get_column_letter(text_end)}{rr}", plan.name, fill=SOFT, size=8, bold=bool(plan.name))
        _merge(ws, f"{get_column_letter(text_end+1)}{rr}:{R}{rr}", won(plan.monthly) if plan.name else "", fill=WHITE, size=8, bold=bool(plan.name))
    total_row = row + 2 + max_plans
    _merge(ws, f"{L}{total_row}:{get_column_letter(text_end)}{total_row}", "신규 보험료 합계", fill=GREEN_LIGHT, color=NAVY, bold=True)
    _merge(ws, f"{get_column_letter(text_end+1)}{total_row}:{R}{total_row}", won(p.new_plan_monthly), fill=GREEN_LIGHT, color=GREEN, bold=True)
    coverage_row = total_row + 2
    _merge(ws, f"{L}{coverage_row}:{R}{coverage_row}", "새롭게 확보되는 핵심 보장", fill=NAVY, color=WHITE, bold=True)
    _merge(ws, f"{L}{coverage_row+1}:{R}{coverage_row+2}", p.coverage, fill=GREEN_LIGHT, color=NAVY, size=8.5, bold=True)
    return coverage_row + 2


def _excel_new_panel(ws, p: Person, left: int, right: int, top: int, max_plans: int) -> int:
    L, R = get_column_letter(left), get_column_letter(right)
    _merge(ws, f"{L}{top}:{R}{top}", "새롭게 가입하는 보험", fill=NAVY, color=WHITE, bold=True)
    text_end = right - 2
    _merge(ws, f"{L}{top+1}:{get_column_letter(text_end)}{top+1}", "보험 또는 보장 구성", fill=NAVY2, color=WHITE, size=8, bold=True)
    _merge(ws, f"{get_column_letter(text_end+1)}{top+1}:{R}{top+1}", "월 보험료", fill=NAVY2, color=WHITE, size=8, bold=True)
    for idx in range(max_plans):
        plan = p.plans[idx] if idx < len(p.plans) else NewPlan()
        rr = top + 2 + idx
        _merge(ws, f"{L}{rr}:{get_column_letter(text_end)}{rr}", plan.name, fill=SOFT, size=8, bold=bool(plan.name))
        _merge(ws, f"{get_column_letter(text_end+1)}{rr}:{R}{rr}", won(plan.monthly) if plan.name else "", fill=WHITE, size=8, bold=bool(plan.name))
    total_row = top + 2 + max_plans
    _merge(ws, f"{L}{total_row}:{get_column_letter(text_end)}{total_row}", "신규 보험료 합계", fill=GREEN_LIGHT, color=NAVY, bold=True)
    _merge(ws, f"{get_column_letter(text_end+1)}{total_row}:{R}{total_row}", won(p.new_plan_monthly), fill=GREEN_LIGHT, color=GREEN, bold=True)
    coverage_row = total_row + 2
    _merge(ws, f"{L}{coverage_row}:{R}{coverage_row}", "새롭게 확보되는 핵심 보장", fill=NAVY, color=WHITE, bold=True)
    _merge(ws, f"{L}{coverage_row+1}:{R}{coverage_row+2}", p.coverage, fill=GREEN_LIGHT, color=NAVY, size=8.5, bold=True)
    return coverage_row + 2


def _excel_bottom(ws, people: list[Person], top: int) -> int:
    t = combined(people)
    spans = [("A", "C"), ("D", "H"), ("I", "M"), ("N", "R")]
    headers = ["한눈에 보는 비교" if len(people) == 1 else "2인 합산 비교", "기존", "리모델링 후", "변화"]
    for span, value in zip(spans, headers):
        _merge(ws, f"{span[0]}{top}:{span[1]}{top}", value, fill=NAVY, color=WHITE, bold=True)
    records = [
        ("월 보험료", t["old_monthly"], t["after_monthly"]),
        ("연간 보험료", t["old_monthly"]*12, t["after_monthly"]*12),
        ("납입 예정 총액", t["old_total"], t["after_total"]),
    ]
    for rr, (label, old, new) in enumerate(records, top+1):
        _merge(ws, f"A{rr}:C{rr}", label, fill=BLUE_LIGHT, bold=True)
        _merge(ws, f"D{rr}:H{rr}", won(old), fill=WHITE)
        _merge(ws, f"I{rr}:M{rr}", won(new), fill=WHITE)
        _merge(ws, f"N{rr}:P{rr}", change_amount(old, new), fill=GOLD_LIGHT, color=GREEN if old > new else INK, bold=True)
        _merge(ws, f"Q{rr}:R{rr}", change_rate(old, new), fill=GOLD_LIGHT, color=RED, bold=True)
    return top + 3


def _excel_detail(wb: Workbook, people: list[Person]) -> None:
    ws = wb.create_sheet("기존 계약 변경")
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEF", [16, 24, 18, 44, 16, 16]):
        ws.column_dimensions[col].width = width
    _merge(ws, "A1:F2", "기존 계약별 처리 계획", color=NAVY, size=18, bold=True, border=False)
    row = 4
    for p in people:
        _merge(ws, f"A{row}:F{row}", f"{p.name or 'OOO'}님", fill=NAVY, color=WHITE, bold=True)
        row += 1
        for col, text in enumerate(["보험회사", "상품명", "처리 방향", "구체적인 변경 내용"], 1):
            end = col if col < 4 else 6
            _merge(ws, f"{get_column_letter(col)}{row}:{get_column_letter(end)}{row}", text, fill=NAVY2, color=WHITE, bold=True)
            if col == 4:
                break
        row += 1
        records = p.contracts or [ExistingContract(detail="입력된 기존 계약 변경 내용이 없습니다.")]
        for c in records:
            action_fill, action_color = ACTION_STYLE.get(c.action, (WHITE, INK))
            _merge(ws, f"A{row}:A{row}", c.company, fill=WHITE)
            _merge(ws, f"B{row}:B{row}", c.product, fill=WHITE)
            _merge(ws, f"C{row}:C{row}", c.action if c.company or c.product else "", fill=action_fill, color=action_color, bold=True)
            _merge(ws, f"D{row}:F{row}", c.detail, fill=WHITE)
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1
    _merge(ws, f"A{row}:F{row}", "※ 감액·해지는 신규 계약의 승인과 보장 개시를 확인한 후 진행합니다.", fill=GOLD_LIGHT, color=MUTED, size=9)
    _excel_setup(ws, "F", row)


def create_excel(people: list[Person], title: str, consultation_date: date, consultant: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "리모델링 비교안"
    widths = [8,8,8,8,8,8,10,10,4,4,10,10,8,8,8,8,8,8]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _excel_top(ws, people, title)
    if len(people) == 1:
        p = people[0]
        _excel_new_panel(ws, p, 11, 18, 8, 5)
        _merge(ws, "A8:H8", "보험료 비교", fill=NAVY, color=WHITE, bold=True)
        left_rows = [("기존 월 보험료", p.old_monthly), ("유지 보험료", p.retained_monthly), ("신규 보험료", p.new_plan_monthly), ("리모델링 후", p.after_monthly)]
        for rr, (label, value) in enumerate(left_rows, 9):
            _merge(ws, f"A{rr}:D{rr}", label, fill=BLUE_LIGHT, bold=True)
            _merge(ws, f"E{rr}:H{rr}", won(value), fill=GOLD_LIGHT if rr == 12 else WHITE, color=NAVY2, bold=True)
        _merge(ws, "A14:H14", "납입 예정 총액", fill=NAVY, color=WHITE, bold=True)
        for rr, (label, value) in enumerate([("기존", p.old_total), ("리모델링 후", p.after_total), ("변화", abs(p.total_change))], 15):
            _merge(ws, f"A{rr}:D{rr}", label, fill=BLUE_LIGHT, bold=True)
            _merge(ws, f"E{rr}:H{rr}", won(value), fill=GOLD_LIGHT if rr == 17 else WHITE, color=GREEN if rr == 17 and p.total_change < 0 else INK, bold=True)
        bottom = _excel_bottom(ws, people, 21)
        last = 27
    else:
        _excel_person_panel(ws, people[0], 1, 8, 8, 4)
        _excel_person_panel(ws, people[1], 11, 18, 8, 4)
        bottom = _excel_bottom(ws, people, 25)
        last = 31
    note = "※ 신규 보험의 납입 예정 총액은 입력한 현재 월 보험료와 납입기간을 기준으로 계산했습니다."
    _merge(ws, f"A{last-1}:R{last-1}", note, color=MUTED, size=8, border=False)
    _merge(ws, f"A{last}:R{last}", f"상담일 {consultation_date:%Y.%m.%d} · 담당자 {consultant or '-'}", color=MUTED, size=8, border=False)
    _excel_setup(ws, "R", last)
    _excel_detail(wb, people)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def load_example(count: int) -> None:
    examples = [
        ("홍길동", "694580", "144940000", "0", "0", [("암·뇌·심장 진단비", "128589"), ("암 주요치료비", "111255"), ("순환계 주요치료비 및 추가 보장", "112863"), ("운전자보험", "15000")]),
        ("홍길순", "495470", "102950000", "0", "0", [("암·뇌·심장 진단비", "95691"), ("암 주요치료비", "101245"), ("순환계 주요치료비 및 추가 보장", "117274"), ("운전자보험", "15000")]),
    ]
    example_contracts = {
        1: [
            {
                "company": "DB손해보험",
                "product": "건강할때가입하는청춘어람플러스종합보험",
                "action": "해지",
                "detail": "DB손해보험 고객센터 1544-0100 상담원 연결 후 해지 요청",
            },
            {
                "company": "KB손해보험",
                "product": "KBThe좋은닥터플러스건강보험Ⅱ",
                "action": "해지",
                "detail": "KB손해보험 고객센터 1544-0114 상담원 연결 후 해지 요청",
            },
            {
                "company": "흥국화재",
                "product": "흥GoodThe건강한0550종합보험",
                "action": "변경",
                "detail": "흥국화재 고객센터 1688-1688 상담원 연결 후 일상생활 배상책임 특약 삭제 요청",
            },
        ],
        2: [
            {
                "company": "메리츠 화재",
                "product": "The알뜰한 건강보험",
                "action": "유지",
                "detail": "기존 가입 조건과 주요 보장을 확인한 결과 유지하는 방향으로 검토",
            },
            {
                "company": "DB손해보험",
                "product": "참 좋은 운전자 상해보험",
                "action": "해지",
                "detail": "DB손해보험 고객센터 1544-0100 상담원 연결 후 해지 요청",
            },
            {
                "company": "삼성생명",
                "product": "통합유니버설종신보험",
                "action": "유지",
                "detail": "기존 가입 조건과 주요 보장을 확인한 결과 유지하는 방향으로 검토",
            },
        ],
    }
    for idx in range(count):
        no = idx + 1
        name, oldm, oldt, keepm, keept, plans = examples[idx]
        st.session_state[f"rm_name_{no}"] = name
        st.session_state[f"rm_old_monthly_{no}"] = oldm
        st.session_state[f"rm_old_total_{no}"] = oldt
        st.session_state[f"rm_retained_monthly_{no}"] = keepm
        st.session_state[f"rm_retained_total_{no}"] = keept
        st.session_state[f"rm_plan_count_{no}"] = 4
        for i, (pname, premium) in enumerate(plans):
            st.session_state[f"rm_plan_name_{no}_{i}"] = pname
            st.session_state[f"rm_plan_premium_{no}_{i}"] = premium
            st.session_state[f"rm_plan_years_{no}_{i}"] = 20
        st.session_state[f"rm_coverage_{no}"] = "암·뇌·심장 진단비 보완 · 주요 치료비 강화 · 보장 공백 보완"

        contracts = example_contracts.get(no, [])
        st.session_state[f"rm_contract_count_{no}"] = len(contracts)
        for contract_index, contract in enumerate(contracts):
            st.session_state[f"rm_contract_company_{no}_{contract_index}"] = contract["company"]
            st.session_state[f"rm_contract_product_{no}_{contract_index}"] = contract["product"]
            st.session_state[f"rm_contract_action_{no}_{contract_index}"] = contract["action"]
            st.session_state[f"rm_contract_detail_{no}_{contract_index}"] = contract["detail"]


def _preview_text(value: str) -> str:
    return html.escape(value) if value else '<span class="rm-empty">입력 전</span>'


def _preview_person(person: Person) -> None:
    name = html.escape(person.name or "OOO")
    st.markdown(f'<div class="rm-preview-name">{name}님</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    c1.metric("기존 월 보험료", won(person.old_monthly))
    c2.metric("변경 후 월 보험료", won(person.after_monthly))
    c3.metric("월 보험료 변화", change_amount(person.old_monthly, person.after_monthly))

    plan_rows = "".join(
        f'<tr><td>{_preview_text(plan.name)}</td><td>{won(plan.monthly)}</td></tr>'
        for plan in person.plans
    ) or '<tr><td><span class="rm-empty">입력 전</span></td><td><span class="rm-empty">입력 전</span></td></tr>'
    contract_rows = "".join(
        f'<tr><td>{_preview_text(contract.company)}</td><td>{_preview_text(contract.product)}</td>'
        f'<td><span class="rm-action rm-action-{html.escape(contract.action)}">{html.escape(contract.action)}</span></td>'
        f'<td>{_preview_text(contract.detail)}</td></tr>'
        for contract in person.contracts
    ) or '<tr><td colspan="4"><span class="rm-empty">입력 전</span></td></tr>'
    st.markdown(
        f"""
        <div class="rm-preview-grid">
          <section class="rm-preview-card">
            <h4>새롭게 가입하는 보험</h4>
            <table><thead><tr><th>보험 또는 보장 구성</th><th>월 보험료</th></tr></thead>
            <tbody>{plan_rows}</tbody>
            <tfoot><tr><th>신규 보험료 합계</th><th>{won(person.new_plan_monthly)}</th></tr></tfoot></table>
          </section>
          <section class="rm-preview-card">
            <h4>새롭게 확보되는 핵심 보장</h4>
            <div class="rm-coverage">{_preview_text(person.coverage)}</div>
          </section>
        </div>
        <section class="rm-preview-card rm-contract-card">
          <h4>기존 계약별 처리 계획</h4>
          <table><thead><tr><th>보험회사</th><th>상품명</th><th>처리 방향</th><th>변경 내용</th></tr></thead>
          <tbody>{contract_rows}</tbody></table>
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_preview(people: list[Person]) -> None:
    st.markdown(
        """
        <style>
        .rm-preview-name{margin:.15rem 0 .75rem;color:#17365D;font-size:1.18rem;font-weight:800}
        .rm-preview-grid{display:grid;grid-template-columns:1.15fr .85fr;gap:.85rem;margin:1rem 0 .85rem}
        .rm-preview-card{overflow:hidden;border:1px solid #DCE6EE;border-radius:15px;background:#FFF;
            box-shadow:0 10px 28px rgba(37,72,98,.05)}
        .rm-preview-card h4{margin:0;padding:.78rem .95rem;background:#F5F9FC;color:#17365D;font-size:.9rem}
        .rm-preview-card table{width:100%;border-collapse:collapse;font-size:.78rem}
        .rm-preview-card th,.rm-preview-card td{padding:.62rem .72rem;border-top:1px solid #E7EEF3;text-align:center}
        .rm-preview-card thead th{color:#52677A;background:#FBFCFD;font-size:.7rem}
        .rm-preview-card tfoot th{color:#24745A;background:#E8F4F2}
        .rm-coverage{min-height:7.25rem;padding:1rem;color:#25364A;line-height:1.7;white-space:pre-wrap}
        .rm-contract-card{margin-bottom:.8rem}.rm-empty{color:#98A2B3;font-weight:500}
        .rm-action{display:inline-block;padding:.18rem .42rem;border-radius:999px;font-size:.68rem;font-weight:800}
        .rm-action-유지{color:#24745A;background:#E8F4F2}.rm-action-감액{color:#9A6700;background:#FFF1D6}
        .rm-action-해지{color:#B42318;background:#FDECEC}.rm-action-변경{color:#1769DC;background:#EAF2FF}
        .rm-action-검토{color:#6941C6;background:#F2ECFF}
        @media(max-width:760px){.rm-preview-grid{grid-template-columns:1fr}.rm-preview-card table{font-size:.7rem}}
        </style>
        """,
        unsafe_allow_html=True,
    )
    if len(people) == 1:
        _preview_person(people[0])
        return
    tabs = st.tabs(["전체 요약", "고객 1", "고객 2"])
    with tabs[0]:
        total = combined(people)
        c1, c2, c3 = st.columns(3)
        c1.metric("합산 기존 월 보험료", won(total["old_monthly"]))
        c2.metric("변경 후 합산 월 보험료", won(total["after_monthly"]))
        c3.metric("합산 월 보험료 변화", change_amount(total["old_monthly"], total["after_monthly"]))
        st.caption("고객별 상세 내용은 고객 1·고객 2 탭에서 확인할 수 있습니다.")
    with tabs[1]:
        _preview_person(people[0])
    with tabs[2]:
        _preview_person(people[1])


def run() -> None:
    page_header("고객 상담", APP_TITLE, "고객별 기존 계약 처리 계획과 새로운 보장 구성을 작성해 엑셀로 내려받습니다.", "RM")
    section_intro("공통 정보", "상담 기본정보", "대상 인원과 상담 정보를 먼저 확인해 주세요.")
    c1, c2, c3, c4 = st.columns([.85, 1, 1.15, 1])
    with c1:
        count = int(st.selectbox("대상 인원", [1, 2], format_func=lambda x: f"{x}명", key="rm_count"))
    with c2:
        consultation_date = st.date_input("상담일", value=date.today(), key="rm_date")
    with c3:
        consultant = st.text_input("담당자", key="rm_consultant", placeholder="예: 박병선 팀장")
    with c4:
        st.markdown('<div style="height:1.78rem"></div>', unsafe_allow_html=True)
        if st.button("예시 데이터 입력", use_container_width=True):
            load_example(count)
            st.rerun()

    section_intro("상세 입력", "고객별 리모델링 내용", "신규 보험과 기존 계약의 처리 계획을 구체적으로 작성해 주세요.")
    people: list[Person] = []
    tabs = st.tabs([f"고객 {i}" for i in range(1, count+1)])
    for i, tab in enumerate(tabs, 1):
        with tab:
            people.append(render_person_inputs(i))
    if count == 2:
        shared = st.checkbox("두 고객의 핵심 보장을 하나로 묶어 표시", key="rm_shared_coverage")
        if shared:
            shared_text = st.text_area("공통 핵심 보장", key="rm_shared_coverage_text", height=75)
            for p in people:
                p.coverage = clean(shared_text)
    display_names = [re.sub(r"님$", "", clean(p.name)) or "OOO" for p in people]
    default_title = " · ".join(f"{name}님" for name in display_names) + " 보험 리모델링 비교안"
    custom_title = st.text_input("자료 제목 (선택)", key="rm_title", placeholder=default_title)
    effective_title = clean(custom_title) or default_title

    section_intro("분석 결과", "자동 계산 결과", "입력한 보험료와 납입기간을 기준으로 자동 계산됩니다.")
    t = combined(people)
    a, b, c = st.columns(3)
    labels = (["월 보험료 변화", "변경 후 월 보험료", "납입 예정 총액 변화"] if count == 1 else
              ["합산 월 보험료 변화", "변경 후 합산 월 보험료", "납입 예정 총액 변화"])
    a.metric(labels[0], change_amount(t["old_monthly"], t["after_monthly"]), change_rate(t["old_monthly"], t["after_monthly"]) or None)
    b.metric(labels[1], won(t["after_monthly"]))
    c.metric(labels[2], change_amount(t["old_total"], t["after_total"]), change_rate(t["old_total"], t["after_total"]) or None)
    missing = [f"고객 {i+1} 이름" for i, p in enumerate(people) if not p.name]
    if not clean(consultant):
        missing.append("담당자")
    if missing:
        st.warning("미입력 항목: " + ", ".join(missing) + " · 확인용 파일은 그대로 다운로드할 수 있습니다.")

    section_intro("미리보기", "리모델링 비교안 미리보기", "다운로드할 자료의 핵심 내용을 실시간으로 확인할 수 있습니다.")
    render_preview(people)

    excel = create_excel(people, effective_title, consultation_date, clean(consultant))
    filename_people = "_".join(f"{safe_filename(name)}님" for name in display_names)
    base = f"{filename_people}_보험리모델링_비교안_{date.today():%Y%m%d}"
    section_intro("다운로드", "엑셀 다운로드", "미리보기 내용을 확인한 뒤 고객 상담용 엑셀을 내려받아 주세요.")
    st.download_button(
        "엑셀로 다운로드",
        excel,
        f"{base}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )


if __name__ == "__main__":
    run()
