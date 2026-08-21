import hashlib
import base64
import re
from copy import copy
from datetime import datetime
from io import BytesIO

import openpyxl
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .ui_components import page_header


DEFAULT_COVERAGES = [
    "질병사망",
    "재해(상해)사망",
    "질병후유장해3%일경우",
    "재해(상해)장해3%일경우",
    "일반암",
    "유사암",
    "표적항암약물허가치료비",
    "항암방사선약물치료비",
    "뇌혈관",
    "뇌졸중",
    "뇌출혈",
    "허혈성심장질환",
    "급성심근경색증",
    "질병수술",
    "질병종수술",
    "상해수술",
    "상해종수술",
    "뇌혈관질환수술",
    "허혈성심장질환수술",
    "질병입원",
    "상해입원",
    "간병인지원입원일-질병",
    "간병인지원입원일-상해",
    "상해간호간병통합입원일당",
    "질병간호간병통합입원일당",
    "교통사고처리지원금",
    "교통사고처리지원금(6주미만)",
    "변호사선임비용",
    "운전자벌금(대인)",
    "운전자벌금(대물)",
    "자동차사고부상위로금",
    "일상생활배상책임",
    "치아보철치료비",
    "치아보존치료비",
    "골절진단비",
    "질병입원(실손)",
    "질병통원(실손)",
    "상해입원(실손)",
    "상해통원(실손)",
]


DISPLAY_NAMES = {
    "질병사망": "질병 사망",
    "재해(상해)사망": "재해(상해) 사망",
    "질병후유장해3%일경우": "질병 후유장해 3%일 경우",
    "질병후유장해80%이상": "질병 후유장해 80% 이상",
    "재해(상해)장해3%일경우": "상해 후유장해 3%일 경우",
    "재해(상해)장해80%이상": "상해 후유장해 80% 이상",
    "고액암": "고액 암",
    "일반암": "일반 암",
    "이차암(재진단,계속암)": "이차암(재진단·계속암)",
    "유사암": "유사 암",
    "표적항암약물허가치료비": "표적항암약물허가치료비",
    "항암방사선약물치료비": "항암방사선·약물치료비",
    "뇌혈관": "뇌 혈관",
    "뇌졸중": "뇌 졸중",
    "뇌출혈": "뇌 출혈",
    "허혈성심장질환": "허혈성 심장 질환",
    "급성심근경색증": "급성 심근경색증",
    "중증치매": "중증 치매",
    "경증치매": "경증 치매",
    "장기간병요양진단(1급)": "장기요양 1등급",
    "장기간병요양진단(1,2급)": "장기요양 1~2등급",
    "장기간병요양진단(1,2,3급)": "장기요양 1~3등급",
    "장기간병요양진단(1,2,3,4급)": "장기요양 1~4등급",
    "암산정특례": "암 산정특례",
    "뇌혈관산정특례": "뇌혈관 산정특례",
    "심장질환산정특례": "심장질환 산정특례",
    "중증치매산정특례": "중증치매 산정특례",
    "질병수술": "질병 수술",
    "질병종수술": "질병 종 수술(1~5종)",
    "상해수술": "상해 수술",
    "상해종수술": "상해 종 수술(1~5종)",
    "암수술": "암 수술",
    "뇌혈관질환수술": "뇌혈관 질환 수술",
    "허혈성심장질환수술": "허혈성 심장 질환 수술",
    "질병입원": "질병 입원",
    "상해입원": "상해 입원",
    "간병인지원입원일-질병": "간병인 지원(질병)",
    "간병인지원입원일-상해": "간병인 지원(상해)",
    "암입원": "암 입원",
    "상해간호간병통합입원일당": "간호간병통합입원(상해)",
    "질병간호간병통합입원일당": "간호간병통합입원(질병)",
    "질병통원": "질병 통원",
    "암통원": "암 통원",
    "상해통원": "상해 통원",
    "치과통원": "치과 통원",
    "응급실내원비": "응급실 내원비",
    "교통사고처리지원금": "교통사고 처리 지원금",
    "교통사고처리지원금(6주미만)": "교통사고 처리 지원금(6주 미만)",
    "변호사선임비용": "변호사 선임 비용",
    "운전자벌금(대인)": "운전자 벌금(대인)",
    "운전자벌금(대물)": "운전자 벌금(대물)",
    "자동차사고부상위로금": "자동차사고 부상 위로금",
    "일상생활배상책임": "일상생활 배상책임",
    "치아보철치료비": "치아 보철 치료비",
    "치아보존치료비": "치아 보존 치료비",
    "화상진단비": "화상 진단비",
    "골절진단비": "골절 진단비",
    "깁스치료비": "깁스 치료비",
    "통풍진단비": "통풍 진단비",
    "대상포진진단비": "대상포진 진단비",
    "질병입원(실손)": "질병 입원(실손)",
    "질병통원(실손)": "질병 통원(실손)",
    "상해입원(실손)": "상해 입원(실손)",
    "상해통원(실손)": "상해 통원(실손)",
    "반려동물배상책임(대물)": "반려동물 배상책임(대물)",
    "반려동물배상책임(대인)": "반려동물 배상책임(대인)",
    "반려동물수술비(개)": "반려동물 수술비(개)",
    "반려동물입원비(개)": "반려동물 입원비(개)",
    "반려동물통원비(개)": "반려동물 통원비(개)",
}


GROUP_RULES = [
    ("사망", {"질병사망", "재해(상해)사망"}),
    (
        "후유\n장해",
        {
            "질병후유장해3%일경우",
            "질병후유장해80%이상",
            "재해(상해)장해3%일경우",
            "재해(상해)장해80%이상",
        },
    ),
    (
        "암\n보장",
        {
            "고액암",
            "일반암",
            "이차암(재진단,계속암)",
            "유사암",
            "표적항암약물허가치료비",
            "항암방사선약물치료비",
        },
    ),
    ("뇌\n보장", {"뇌혈관", "뇌졸중", "뇌출혈"}),
    ("심장\n보장", {"허혈성심장질환", "급성심근경색증"}),
    (
        "치매·\n장기요양",
        {
            "중증치매",
            "경증치매",
            "장기간병요양진단(1급)",
            "장기간병요양진단(1,2급)",
            "장기간병요양진단(1,2,3급)",
            "장기간병요양진단(1,2,3,4급)",
        },
    ),
    ("산정\n특례", {"암산정특례", "뇌혈관산정특례", "심장질환산정특례", "중증치매산정특례"}),
    (
        "수술",
        {
            "질병수술",
            "질병종수술",
            "상해수술",
            "상해종수술",
            "암수술",
            "뇌혈관질환수술",
            "허혈성심장질환수술",
        },
    ),
    ("입원", {"질병입원", "상해입원", "암입원"}),
    (
        "간호\n간병",
        {
            "간병인지원입원일-질병",
            "간병인지원입원일-상해",
            "상해간호간병통합입원일당",
            "질병간호간병통합입원일당",
        },
    ),
    ("통원·\n응급", {"질병통원", "암통원", "상해통원", "치과통원", "응급실내원비"}),
    (
        "운전자",
        {
            "교통사고처리지원금",
            "교통사고처리지원금(6주미만)",
            "변호사선임비용",
            "운전자벌금(대인)",
            "운전자벌금(대물)",
            "자동차사고부상위로금",
        },
    ),
    ("배상\n책임", {"일상생활배상책임"}),
    ("치아", {"치아보철치료비", "치아보존치료비"}),
    ("생활\n보장", {"화상진단비", "골절진단비", "깁스치료비", "통풍진단비", "대상포진진단비"}),
    ("실손", {"질병입원(실손)", "질병통원(실손)", "상해입원(실손)", "상해통원(실손)"}),
    (
        "반려\n동물",
        {
            "반려동물배상책임(대물)",
            "반려동물배상책임(대인)",
            "반려동물수술비(개)",
            "반려동물입원비(개)",
            "반려동물통원비(개)",
        },
    ),
]


COLORS = {
    "header": "DCE6F2",
    "premium": "FCD5B5",
    "cancer": "EBF1DE",
    "brain": "FDEADA",
    "heart": "E6E0EC",
    "white": "FFFFFF",
    "blue": "0000FF",
    "black": "000000",
    "line": "728197",
    "red": "FF0000",
    "completed": "92D050",
}


LOGO_BASE64 = "iVBORw0KGgoAAAANSUhEUgAABKYAAACTCAIAAAA7hIa8AAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUwAADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAAh1QAAIdUBBJy0nQAAkuRJREFUeF7tfQeAHMW1Lfj5v/+fn7NfclQimAwGGxuDccYJbMDYxsbGxsCzUQCtVlkCZaGEUCJIqyyhhAJotcoBSQhJSCjnnHPY3Ukd95+q6u5tTeiumemZnd254/Eyu6rprjoVuk7de8+9uqam5qoG9kKTrr7qajSq5mp8wH9E+5y/4EMN+19NTeTy1aGLZtXZmvBF8/KpmsunzcpTV10+cVUsZJ47dPXVV5ns+ya+efVVHzPZVdn3rr76Y/yK7MfV+AOKsP9c9bH/bHb1Jz5z9Wf+5+rPfulj/3Pt1Z/8j6s//d8f+9yXrv7Ep6+66l94lUyUYtVi38blrubfYzUT9aUXIUAIEAKEACFACBAChAAhQAgQAgEjAOrRsF6mWcP/J36AqOHF/mOaStisPm+cPahvXxRbMDg69pnw4IfCL/8g9OKd1e2ury5pXN2mcVWbptV4lzYLtcGv+NwEf7Q/NKlqw97sLyWNQ6X4jPJN8IEXaxpi78aVvDz7Y6dbQr3uDQ9+MDr2WXXZSH3fWrPyjBmtNk3jigqyOopa0osQIAQIAUKAECAECAFCgBAgBAiBgBFg1qaG9mIET2ccT7yqzum7litLhkfG/zMy8OdV7a5jFI4RPEbSbFLHaBuncI1C+KeSppz7oQync+wvNvFjHwTHY/QvxN+cK7K3YIPWz5ImVSXsIlVtmjGuWNIo1OPb0bK/KYtHaPvXmkrErDEsZkqUr6ENQWoPIUAIEAKEACFACBAChAAhUCgINDjKBxalq8bF4/rBdbHFwyJvPBHq+e0QDG6cwoXaNOKWOovvhdlfmoZLYZQDCQR5a+YY9IQRT5jvBMHjdj9B8HCRxlXcuMd/BT8UBsDG1aWsDOiiKCz4pKCC+EuohFsI2zYNdbkt1P8nsdkv6fvXGlXnmd2PXoQAIUAIEAKEACFACBAChAAhQAjkAIFCp3zc5ZG5Prp8H2sdNgUg/HfmumnEQvqWedEZnWOv/CLc8SZBsRj7KmlSWdKsutQxyjXBP9kmPmGvs3gdStoEjxUGWwuVNLL+4uJ4nAdyax77IyN49lvcgt/UJorCEnhlGVSpGTMndrwlOuL3ytI3DEQSWi0VTqmiXXgzSyBZAXMw8umShAAhQAgQAoQAIUAIEAKEQFEgUOiUTzAf5yUYoP1HED2jxjCM4zvUtW9FRv65utPNzFgHKsXYHZiYMMFxmxui75hpDnF6IjzPsrwxjsf/yfbwxHeZwycje+xtOXDy63DfTmbHYyF/4l+ZnZB/Pd234x3KrYtNqjvcFJ3QEkY/OHzadM89/ijYryhmIzWSECAECAFCgBAgBAgBQoAQCByBgqd8lpHLYJonFvvDR7MG3puVp9SP5oQnPBfufW91+6+DfVnulxbf4zTPDr3jfpjMG5MZ3Jj4CiOEgtFxyie8NLl2C+dywlHT/ruw+DHnT/YG6yttym7HTILMPJgu32OXxUWEnye/COOcqNtLd0bKntIOfgilGWHk48ZLMvIFPuzpgoQAIUAIEAKEACFACBAChECxIFAPKJ/LqZMTIDWq7X4vNrtbqM99wlfT0dWEUIrjRcmteYKSOT6WTLUlzIVYqkqbweLHqReT3GQuoG2vDfW+NwyNzV73Qmkz1Ou+UM/vhV66i1kLOTdjyi4WCeReoLh+rVZnelY+1IGRT7iMcj9SUFAmBMo0Y7giaMebomP/Vz+6pcbQuccqIv3IylcsE5LaSQgQAoQAIUAIEAKEACFACASLQMFTPkZ4uMFLixnnDqurxkWG/6668y1CXsXmexavA48CMROqKpYljf0qBFSaVHe5LdLrPmRNiIx4LDajs7JoqLZuOsQ8jZN7jPOHjQtHDSRRqDzNf/IPl8+Yl0+aFw4bZ/br+z5Qty1QV4yKzewSHfVk5JVfhnvfH+p6hytgLy3Wx+giY3dc7kWQPe6P6vIv7f7t2Ds9jAvHEKJo55oItuvpaoQAIUAIEAKEACFACBAChAAh0PARqGPK53ZatPLpMaOeo2DJo/VqapBMLzrrxeoBPw3BOidMc1aKPB6tV9rIkWbhRjzO90RahbbXhAf9UpnRSX1/orZ7uX5ihxm+BBLl6thMDGhm6IJxcpe25z1lZVnsrefDPb5teWba6i9WhJ4wBrKqCg0Yq24uQZdEp1D2lyquLsNiBYc+rG2bX2NoVqpBJ9Ngwx+Z1EJCgBAgBAgBQoAQIAQIAUKAEAgAgTqmfEKNE5xOmPLYT6G+KVQ6Y9X63tWR0U+HOn7dcbAUdI77WFqymTw/npVqrwocr+c90bKnlPmD9N3vGeHLSNoA3lib8DwTipcItKiiiLRjvNQ8f1R9ryw6oUWo932CrfFa8VztLDsf6szSu1fa4jE+4X8s0k/khYcLaLPo9I7mhWPsLjZI5OoZwNinSxAChAAhQAgQAoQAIUAIEAJFgECdUz7LbuVIcYKcGfhftFLbsTT65p9DXW+37GPC6RH+kDxVukh7YEXrQUylww3Vve+Ljv2HvmmufmKnGa3i7EiQRyH8IrgkWFP2L0dIRvBVEFZxWZN5n57arW6YHRnzv9Uv3hVqew2vJ6N8MPRxCRmLpnpTvipHFxT53EEg210bGfZb7eimGh0p5q1mZd8MugIhQAgQAoQAIUAIEAKEACFACDR4BOqe8gnDngU0XC61mLalIvLGH6vbX2eRPUZ7hEyLiHwT6c65w2RJk3Df70dnvqhtLjcjlznxsmyG/IK2ydD6GKxtTBj6uLaKxSw5AxSfDd04fSC2eHh48EPVbZHhnRNUVnM7LbtPXgeuKGMH+wm50Uj3u7VN75p6jDPZIKhrgx/d1EBCgBAgBAgBQoAQIAQIAUKg6BEoAMpnsSTTiFapO5ZGXvtDdacbrXg8TvO4NU8YykRKgybwdQz1uCcMWcvti8zLp01N4QSIJW+wzXq2Yc9yHLUoUkA6KHH+lVameDsjfC0FZJKboQvKlorwiD9AckaQVZHZzzevgzv9g0V0Ydvseoe2agzaW/TjlgAgBAgBQoAQIAQIAUKAECAECAEpBPJO+biFyjHqcVMc+4N+dHN08vMslzqXsqy0tCvB7hjfYyYvll+B8aVwr+9qCwbrR7bAHpja2OUYwerWGmbdHWxW2zyveshDLNODCO3jWd0t31T+IeEtdEdFGce82TTU6WZl1VhTYxnbLW9S+5NUh1MhQoAQIAQIAUKAECAECAFCgBAoJgTqgPI5Kp3cEmdARTP6dufq9jcIX03LAsai9Vh+BStnHT50uCEy/A/KhlmmEhLemwGZ7PLT28wEWKNE1U1zkeABieAZ5eOJ3cEAudyLj92PQcFpMEOp3bXKrJdMnct4cpti3fLa/CBIdyEECAFCgBAgBAgBQoAQIAQIgQwQyDvlE/F1gqxUnVNXjQ29/MNqUCB3egORoBwBbEhWjkQFHW+OvP4Hddtio+oc/2atX2X9YX0iXpG9jXNHlPL+oS63swBFbtJ0rHmpiJ+IAKzkTqEwezK7X7vrlRWjTDXqiHhm0Pf0FUKAECAECAFCgBAgBAgBQoAQaPAI5JvyCeZj6oq6d1V4yG/CpdC0FI6OLK861ylhn7nFD0qV10XefELbtYJF6/HvWVYtoZFik6h60kkwS7I6s4hDJJbfuTw64g+h0mscD09vQx+KMb9WkD1QPi5bCi1Tdc0kS8qFzHz1ZBBQNQkBQoAQIAQIAUKAECAECIE8I5BvyocEDMb5w9EpbUNdv2GrmIjANlAaxvq4ua9pqP3Xw8MeZeoskUph1xO5yEX0muXNaOlz1ge6Y2VSr60qqJ9RdSY2u1u40y1Cz9PjLTiwHQQoKDHP7d75Zv3QhhojkMwTeR54dDtCgBAgBAgBQoAQIAQIAUKAEMgHAgFTPsftspabiVYIT05DUzfPCw/6FQieTWA4mWE2vWbcZZEn3Ot2t7p6vFl9wYpVywcO+bmHi/JxUEwlrK6eACdPxvpKoOxiJWZIpH9ukU+h5sIdXxsjXx8otMg7KNiwBXd+GkR3IQQIAUKAECAECAFCgBAgBAiBwkYgYMpnpcJzJY4T/phw5TTO7I9Obl2N/ArcdTMMxsLfXKAFMi08pK3LHbF3euiXjotwv/pgv8usey27JcfJ0PasDg94gKFhmfusRPOepr9GTPqFq7lEJzSviUWsCEnH/zWzetG3CAFCgBAgBAgBQoAQIAQIAUKgYSEQOOWrhcdRaTGRXX3rgkj/B0KlzYRNT4hzhplpCwywEcvE0O7ayPDfajuWQZLEslaJqLeGqEbp2OKs2EQ4eZ7YFh32GCfDIrLRz9WTp6dnTrBQvml3vfrRHHiKupQ7Gy5ZbljTj1pDCBAChAAhQAgQAoQAIUAI5BqBgCmf7djJ6JolUFl9Llb+cqjjTdVtGlUyQxYoDTKqc2LDPDk59+tyh7poGLKWW0n6LMJiXSHXEOT/+iJpuxOjKKiacfZQ9cCfc4Mnf3unbWBpHrh1lCd4CA/8BYyoyHhhS9rkv010R0KAECAECAFCgBAgBAgBQoAQKEQEAqZ8trAKpzOGph/ZFHn9TzBDMYLHHBEtExb/tRkLXSttigL63vdNXbWj0UQYoHgZDdrKxxsqWsgJoHFiZ2T477iPq3+aviph6OMm01Dba5TyfjWaYgVNNkTTaCHOHqoTIUAIEAKEACFACBAChAAhUPAIZEX5RLidyAVemyLPEl0xtI1zQn3u45yEWaKEsyI37iHHADPxVXe4Pjato1l52qI+CWA1XPfE+JZZEY8Q3zy2NdT92wwoO7QvlZ+nJeJikcMm1S/eqZ87bHdGwY87qiAhQAgQAoQAIUAIEAKEACFACOQFgawoHzdOuV6WuarGrDwTe6d3dXsY98DxWDoBoUUJR0Q4dnJ3xCYgNhpyMCDhHtwRhbWrmF+CMcOoCdvo7vcAjkhDb0nd+Pp5MjmcprEppaah21qpxYwmtZ0QIAQIAUKAECAECAFCgBAgBCwEsqN8wsZnZcsTlirDrDobKfsb5FiEGUqY9USmdTh2MtZX2iw87BF9zyorLTujOdytsYhpn20s5eQXeerXTKpuhyzt3L0TbNnP1VPgHO79Xf3ELls1h4Y4IUAIEAKEACFACBAChAAhQAgQAjVZUT5HrIUbqFg4mnZgbfSVXwllTmSN44wFxI/zFuHeWdo0Or45cu5xfmhLcloWruK19DGqzCQ3+YsDE5n8AidyiH70z9lgUeu2zZRFQ2AmpHFNCBAChAAhQAgQAoQAIUAIEAKEgEAgK8rnaI8wW5+u6HtXheGRKNw4S5oJK5+VaR0Kk9Dn7HQrHD7N6nOWddD2C3VlFyjafuE8zwmMRLa+03vDwx62ObOUoAuMqOGhj5qVZ4vYXFq044caTggQAoQAIUAIEAKEACFACCRHIE3KZ9mhhA8mJyniL4aqvj8x1IPpjnAHTm7Z45+ZTAtMVUi53ukWZUUZ2aA8R6LLzmcYysZZ4Q7X87QNzEDq/WaiOEC+/XXGoY1FHxlJs50QIAQIAUKAECAECAFCgBAgBCwE0qN8nOExjRGb8nG6F61UFrwCZ06b5tWSE0ttsqRxuPs96pZ5YIYEvCQCPLZRj016wWHRPpSP+3+C+EXf7hKnqiN5RypGCBAChAAhQAgQAoQAIUAIEAIND4G0KZ/jhMllQmrMaFXs3V6hDjeLOL0EWsI1J7vcqm0qh5hkw4Mvdy0S9j794Ifh7nfLZeqz5D1DA39hqpHcVYyuTAgQAoQAIUAIEAKEACFACBAC9QiBNCmfCDizjXxmLKTMHxSCOKeVRC4+5IzF9XX7lrq5nGVaF8kY6CWDAIOZpaE31Wj0nd5VbZv5mfiYJir0cpgDbbe79CObZW5CZQgBQoAQIAQIAUKAECAECAFCoMEjkB7lE2SPsRGY+MIXY7O7Vnf4up1mPZnEyIt3alsXmoZ2Zf6+Bo9q1g3kEPOISZMnZ/+WRCwfj5mETbXjTcoHU7KuAV2AECAECAFCgBAgBAgBQoAQIAQaAgLpUT5LVhJURFdj7/YJceuTO4QvzOQ6YWviIi6db4OmC0rWe5xqTZu2ndKVijBJ6/i/Ovnla7/D5W54agruE8u+yZMZJgPISoDByLURnVyCzOw83YUjjZPEoIqOqOLIx97tXVAyOVbmxno/DqgBhAAhkBwBmuM0MlIhkMHYyOArhYC/U+16Wn+2I6Hj+byPJMI875CnfcMM+iiDr6RdrfS/kB7lE/zEjFbH5nQPtbteMBD3m3t4MltTqOMN2ppJzFJVf182LbOYGWdogq1ZOIimJSyRzCOzNk+CIHU2Wbb+Li5tJWVIsYfibI+X0Q+srep4oyV8ymImWdheot2P5UIsbVzZpklk/D/MyOVCAL4wB30hIEN1IAQaHgJ8ffT33xfFZEo2PIioRZII0PCQBMqbaacFoyic1leyrGTxfJ1WvHrd1/KTQr5knQCSHuVj64GmKkuGhjrdXIVUe0mIB2cj7a6PzulRU29FRGw7HCd01g/HOFdrvvPsMGvpFFTPymTBfkFAo/0n2xiYYotkJ6rHf6vORt/4IzPiWVa+pCY+JtcJ8NEvkeGPmZdP18l4cm5qGMaCBQtef/311157DT89XsuWLSucSTJ37lzv2op/Xb58ed3CS3cnBAoHAWf+Hjt2rGvXrs/4vZ599tnFixfbJ2b+FLFwWko1yRiBjz766I033pBZXRVFKZwngnx7E+scCoUwIzDUR40ahYbPmDFj69atFy9eVNVa16eMW5r4RTxzI5EIrn/mzJlDhw6tXr160aJF48aNk8EcZaZPn058T767JUuim9atW4ddkO/gB/7oO8nLUrH8IOCeZZhN//jHP/webs+8+OKLsVgsP9XL4C7pUb4aXVFXjw+3/7pwIEw0NHES2CQyrZ0ZrhQCJBnUqe6/4rLv2azP2p/wNdEwYtU1kYvmpRPGxWPmxaM1CW/j8pka2NmiVTDRCRMfN3i6z9DEIXfKo25m3zMNa8CZhrL0DSaTw2VRqxi1i0vTZ9v9SpiVNdzne8bZg3ULIx7bv/nNb67ir6uvvlp8SPp68skn67aq7rt/97vf9a0wCjz++OMZP6oLp7FUE0IgWAQ2btx43XXXeUx2558GDBgQ7K3pagWOAHZCH//4x2XGRmVlZYG3xbd62PPNmzfv5z//+ac//WnnIYjn4Mc+9rFrrrnmhRde2LFjBxha9g8RYTs6fPjwrFmzunXr9oc//OGGG2745Cc/KYNzXJlbb701+/r4IlNsBdDLXbp0kemO2267befOncWGTz1qb/fu3X1XMMzxr371q1VVVQV7epIe5dM2zg6/dJclEwI3woT84IxvvPqQGbog4d1TuH3NDXM6Uh2Y4Qtm5Snj1G7j4Hp95xJ9wzRt5Uh9xevasmH6sqH6klfYe/EgfTF+1r419nkwymjLh+krXtNXlxkbpuk7FhgH15mndhmXT5ih8zWxEM9akSKWz2LKghQyD0/98MYQsjUIgRYWP3kF6xN/cd5QSTVO7albfEH5Hn74Ye+VTlDBv/71r3Vb1UTK57tAg/IV7JQuHDCpJsWGACw5RPmKrdN92yuIhDzlExum+vu6dOkSrAGf+cxn3M8R97kniF/jxo1x6hGNRrNp5smTJ2fOnImH0de+9rVPfepTuIX36ar3c40oXzZ9keq78pQP+BPly0UXBHVNGcqHKeZQvqDuG+x1vCmf5c0orFHG8e3heOnIRtVthKwIN/rB+tT7e9qBdcKXsUBOjGpD7xzNFAahi2sJfiWOy8D0Lp80j2zUt5Zr66eCsKkLB2jzemnzeuKnXoF3b628N/uVfe7Ff+LXXmp5TxXFKnrqFawkL8z+jl/1cvGX3tbX57+sLR9hfDBe3/yOsf994+x+EzZDy/tTJMGohZ1XlNdPCUUH/QJOm7DvhVnW9TgrX/yvxtE6ztMgQ/nEE6gArXySlC/YqUhXIwTqOwJE+ep7D+au/qB8//Iv/+K7tKJAvaZ8MFG2atVKppn/5//8nzfffFPToGeenjMUyofDYXhswifF1+wgUxNRBpQjd71ftFcmytdgur4YKB+T5hTOmcaJHeFBv4gXa2GWpaZIFM4dDhuH2n1d3TKvBu4KjD/pMJQVQmfbvpEWm7K8LC3Wx6meFjMvHdOPbNA2ztSXDNErXlbn9QF5c5geeJ3G+Bv+wn4y5maTQG1eD/538U+99AX9Ndj3YOVbNFDFtzj3c7gi44Qog3c5foIBMt6oVPTRF/TTVo7Sdiw0T+4wqs6YusBcpDG0fEHxh9jsnkIKtaoNmDZRvpwMLuHY6fsSVj56EQKEgBsBonw0HlIh0OApHzsWN4whQ4b867/+q+8TRBRo1KjRypUr0xozoIirVq361re+5ebP2Rj3nKoS5UurIyQLE+WTBKrwixUH5RPHT0gIPupv1aXxCcHDIB48/TeLLmvbLDqjY42hMzdER46yEGL5GOdz2xx5+BzemsKC8fau1NZO1JcOEwxNregNYsb4GCdy+FUY6Bipm99bAzdbNFBbOkRfM9bcOEPfVgFfTePwBuPsXmasO7vPvHDYuHQcb/PiMePsARN/PLPXAJFDsb2rjC3l+vq3jPfeVJcM1hYN0Be8rFX0rSWW5T20+X315SO09VP0PcvNqnOmGmNenTaY2uZyy3UzSSwfWfmCWTGI8gWDI12lKBEgyleU3S7V6AZP+YAC9Fpuv/12Sb6HYqBqbdq0ASuQQhC+PooybNiwL3/5y3EcjyifJID5L0aUL/+Y5+iOxUD5eMp1NaaU97W9N69MyWCFkLFAsujw3zMhE2aUEkHJrjwFOeoBycs6+ptWrUyj6pyxf5W+drxW0Ydb3npqzPeS+WSC6amW4U4Y6PpoSwdr6yZr2yvMIx+Z5w7WhC6C1nIXUNzeStjAKmKLtPCP1stWgWFCLA7r5N6jGmN0p3frB9eCN+prJ+lLhzBKye4OtslvveBlfc0YY99q4wJQZXY/4/zR6o43M91OZlZNlvjeZfojx07J0RFXjChfZrjRtwgBIECUj4ZBKgSKgfJBJ1OefYmSCMNzC3h6jB+4jPbq1QvuoIIrCmIpfztfIkpWvlxMXqJ8uUC1Tq5ZDJSPaYdoG2aFun6jKpUnIZeIrO58q75zGVOY5F0h0gtYH+ukc1w3Fe6RrCVqxDh/WN84U1s21LDJHnOznMcse8x7E1Y+xrX6asuG6Gsnm3tWmOePmOFLMHJaTbOiAa08Cw65E9TOMh7yNAxX3J/D4hIBvfI3+MFCJyZyCeKfjOCtn6avGAEboGX9m99HWzxQ/2CccWSjcXJXtP8DjOyV+PA9+N8S5cts3BHlyww3+hYhQJSPxoAHAg2e8mED8P777/syqziqBikX2O58Rw6YQ6dOnT7xiU/EXZ8ony90dVuAKF/d4h/g3Rsi5XMYm5WOzoR3YrjnPZZEZDKmUV2C3N+NYzO7XkFqrK8HiHYal7LIlSBeIjWCoRqnd6vr31IX9LfFVIQcC8ge+wCmx/w5Fw3UN84CWTIrT6ebRN4ilmlUM1VRsyZ8yTy919i1VF9VxohoeQ84nTKpmMWvRAf9MoV9j9n9EOOHSD8eXdnIOELyLZl0BlG+TFCj7xACHAGy8tFASIVAg6d8aDii7CQpn7tYKsrnlnVBTj/Y99IieKKw/FfIypeLyUuULxeo1sk1GyLlc9LECbfF8MXomGdCJU1ZtB6sfAkhZCK0LPLqg8a5wzwLX2G8GM0TpjZ4pUaM49v090dz/UwmlyKsZ0JJBb+yD4sGGuunGEc/qlHCNYaWnnhWblrMvUSNGl03L51EtKHx/jgNwqFzX4q+8nNkY2ekLsGxM1zCIip5vsTG1d3vrhdJGsRjjxQ7czOI6KqEQL4RIMqXb8Trz/2KgfLt3r07Xcr3xS9+0dexE8bDZs2aOVf2YHHyBC+xnoLypSsfWn8GYN3UlChf3eCeg7s2QMrHeZ5wV4RUp6YsHxVqfwOjfJzaJfXtrOpwg7pmkgg2c0Wx5QBvuUs6KpfIegcZFX3DdH3BAMdv08mXYCCnQkUfbcmrxtZ5xrmDJqRc7OwIhdAKkafBajHW4FiVeWafvm1e9M0/gdSxxBhX+tk6Gfmq2jRFN4V73WucOSAHWK5KUZKGXCFL1yUEChUBonyF2jN1X68GT/nwoEZ6iSZNmsiwPieH3p///Oek8i0O9QIhfOqpp+D/KXPZpGWg7YlcDr6vb3zjG8T3Ap8nRPkCh7SuLtgAKR8/4rEkSYxjW0Iv3lnNlUI8xELCwx4z1DDrAyGDWQAv2MdYaNz2RUwSE6Is3HXTitNjeRF4ogXkUt+1FMUEybVrbkuN1n0rhMlUQMrNlfyzuvQ1kD2mj5okFTvL3wB7LLxww0MfgXmwbhtBlK9u8ae7EwL5R4AoX/4xry93bPCUDx2h67qjsCLD0P7rv/5r4cKFqXpQ7MU2b94sn49BMEm84AV64403/vGPfywpKenTp89giddbb73Ft0IFsYurL6Pat55E+Xwhqi8FGiTlEzQDLp2XIqOfEo6CzL5nGfoSMoB3vEHdscQ2ixWGlU+LGad26yvfhBqn7cAJN07Y9PBmSi3akqHGtnlm9XkIb9ZaJgXtEynQC2IAurRA7Rx9qJy6bnp1hxuSOXY2QVClTc6bRMf+Q7DZOnwR5atD8OnWhECdIECUr05grxc3LQbKh444ceLEI488IsP3/t//+3/dunWLRqMe3QfC8I9//MOJynNsgx7X/9KXvgTL4erVq48fP46M7ba2eGHsa+rFSA20kkT5AoWzLi/WgCmfoa6bVt3pRpdxzzIrhUUKPmZiYtak6PgWZuSyyx00T/2RQC45TcXN1Zi+fb6K1HnllgIn+2DlQ++NyD3jg/HIun5ljvi4ixXOyphYE1P96N1Q51tZUCWP2XO/0VmVvF/wjs3uZuoqJ7FWPvf8e6sS5cvTZKDbEAIFgwBRvoLpioKrSJFQPuB+9OhRsD7HFTNpfN0nP/lJKHAir7p3P4FAIuu6fITeXXfdhcA/+IIWytl1wQ3DfFeIKF++Ec/Z/Rog5RN+jcbZg+E+P6wuSbDp8YxwTB1EeBW++A1953Kotth8K2dIX3lhOyWe+Kvl9Mi4zYVj2geTINPC8y4wgRaeYY8LtODD8uFwVeUxe/X3Zaqb3kU+DNELV74t/1vWNe2vV1aMEijZgZl1YLokyld/xxnVnBDIDAGifJnhVgzfKh7Kh96srq4uKyv79re//alPfcptkQMPRLDfb3/72/Xr18MLlO1gPB/Oixcv/p//+R9fm6Gw/v3sZz9DYsBiGEv1qI1E+epRZ3lXtQFSPtZgw4jNH8hcBIVK55Vv/IVb+ZriZ2TkkyyhnEW73Ln4ctzFNsW0KQ1Yp25cOKIj254dtsddOkXYHsx9fXQIciJNfIHEGmYGD+dwsPJVd74Fefnioit5Fg2WoSEMMU8I6iwdVhOtssitRf0yu2vm3yLKlzl29E1CoH4iQJSvfvZbPmpdVJRPAArXyuXLl48ePXogf7366qtz5szZsmUL/C3diCdlfeKP4I2QXfGgfI4B8DOf+czGjRudS5GVLx9jWuIeRPkkQKofRRom5dNP7Ih0v7uSh/AlVYbkroONQm2v1XcurdUWYRka8mU/E/exFC0NeGma+1YjpR5C9VgOBhA8FrkHW19vHV6dC/oZu5aZmnCXr7V71Y8hFl9LU31/YnX764T3ZoJjJ0/PAIfPLrdqb3dAXgrz0lErLWFdLP9E+ernGKNaEwKZI0CUL3PsGvo3i4ryeT9yJR/IcPvELjMV33PInvjwzDPPNPQRVC/bR5SvXnZbsko3QMpnKqHo251Cpc0cIZAkDoRw+CxtHB7yiFF11iZfee5Tx8xnmrqCtHVCmZO/bWVOIdS5+BXj4DpTi/Nrzxc1DRQVznANddmblmNtSVwgn6Do6JomkYE/BQHWoE26dKh+aL3ly8pyFeb1RZQvr3DTzQiBAkCAKF8BdEKBVqE4KZ8Hu+Py6F67ETxDW7du7Uv5UAD6nPPmzSvQji/uahHlazD931Aon73mgFTohzZUd72d+W1y+cfwlbyCkw1OLdpdoywZhsR9ddSXwpkU/pyacWCNtmgAF2thadb1ip5IuMc+iwTrRzeBFEqYH69Yea1cE9yC6Ze0wXJrvQIHu3aWWdHPWV8eQ2QajFW8Ypv44qx8cMRtjAyK+Fdl4RDz9B5j5UgEMYIMG7uXmUrETveQv0waRPmcnoV7D57H4lVeXu58zt2HRYsWISnwnj17QqGQ/ADLc8mzZ8/u3bt3+/bt8jjMnz8fkuX41pkzZ3JRW2cHBgmEiooK34rt2rWLrUXBhcoiBAht9L0vCiTN5ZUNJvA3S3VfQOGMWziPedw6EMp3+PDhnTt3IopJ1EemI8SoOHXqVDYIpPVd6GpAXj+fk3rJkiVo4759+y5fvuxR1QBHY1qA+BYuBsonwMfaKzOFMdMx3z1wk6d8X/va1zD1fLtAsoAvF/UdfvBc3b9/P5aUBQsWBP7swxKEiYBZEOfFGjfyC2QiFBvlS2sKyEwTjzLu3dR7772HeXfgwIFIJJJ0fGY/HiQp33/8x3/MmDEjy6Ylfh3POOzoLly4ID+Lk5a8CrYjEeQGRhGd9AJIHahduLRpdUl8vm9hX2JUsNd9xtHNUEyRvHfgxditEb93cL0OT07LsgezXi/4c6rMn7OntmyoeXq3cGuUvrtT2BY9EWkbkl3hihSEIrWDTREd51H+wZ1rQboiyQoy8qmEY1PbVQlrXoJjJ/7OYi/bXqcd3shuG6k0NkxXF/RlAY0fTjXDF3hl7M1p7vuOKJ/Tjch39K/5ff3f//t//+3f/u0Tn/jEv//7vyP1EzTfWrZsibAQMBlnyRArYPbroMywxl2OHTu2dOnS4cOHP/vsszfccMPnPvc5VA+VxCstbKBsLpqGK6BdcGd64403VqxYAQKZWJOMW3frrbfK1Or555+Xab58GTy0EJMjbo1O9KhDqgeb/L3cJQHU008/LdPkxx57DFM71V0yoHwQqd+6devkyZMxRG+55ZZPf/rTYmDIgOBUWAwJMSquu+46pK4eOnQoOBKOBuIIajbbWXerx40b99nPflYGsaDKiEmNFyb1f//3f99zzz2Q7x8/fjzk+M+fP59Zv+fzW8VA+QSeSIUn0+kQ7dyxY0fGlE9Y/4RXJ1bFa665ButqNq9HH31Ucjwkrqs4bcEZTY8ePX79619/5StfQdMwUFErDFrxkgFEsox4BGAWQNjmxz/+MTbi77zzDp4vgZ+CSaLhXazYKJ9Ao02bNpK9mU0x97hyL49f/OIXv//977dr1w7L44cffoiDlcQRm8HeQJLyYUpm06jE74pmund03/zmN1u1aiV2dInH395Nu8o2VBn6ie3hF+/idjzYi3gahsRMALD+wXWw7Cm4C6bDpgKZO9ZFWHvA907u0BYPBt8D01NYCB8T5+S/9tQW9DdO7DDAahDmJ3PnK2x1giY68YK4jI6EB8w9Ek3WYvyn/UYiBPyr0Cy1TILsfpZt0EqQ4FxNpiopy7CdCpIljnoSvI5TvkRlHcbGI4N+aV4+bSVwV0L6jvk8P2Evbe144/IJp8vsBmZVJe8vE+Vz8Jk0aZKv5FquCwhJN+wU77333i5dusA8ha12BqueR6cnXVVhZly7dm3z5s2hHo7tuLNHcX+I++wNhZOZygllgQje5z//eeyAsQhu27bNSUWVzeDGzkmmR9CubO6S+F1Yt7Cyy9w6WMqHmvz1r3/1VokQgD/88MPZUz5casCAAZcuXYI5GnvEZs2auZNNu6shL0+fWBLjDZKJ6COMQPdTP5BhP2bMGGw63dtumV7LvkxcBBdww6S+//77sduG6cOd5C2QZgY4vIuH8uHwQqajsY0D5fPoJm8rn1g25SeIb5VwziXZ3c65CURHDx482Lt3b5y+4azKWckTaxVgPZ2GONfEnhjPF6z/MIMUWoKK4qR86Avf8RZggVTjDX/HAcSPfvSj/v3741DA48klM/IlKV+A7Uq8lGip8xOnHt/5znc6duwI07ezJfCnfIy0aGr0nZ5Vpc2YEakUeb0bxWlCCo7B0rK3aayufYuTijqx8jHCB/lNbckQBO8xgsc0OZGJgYlzgv7Bz1M/+AFLs85MgVdY41J3qsPyUN4ww5X6ucP6oc36zpXah+9oq6dqS0Yp81/T5o/QKoar/K3MZz/15WO0NTP0TQv0PR8YR3eYl87UaKq4p3AJBUQBkiuj8kx4wAMhWF9L46183DYLln5NrLwfz8jHbs3ubRj67mX6wgFKRU99+YiayyfZaMiLhidRPjHeAPjEiRPTYjUBLhmpHrTYKf7hD3+A141vYiiZpTBpGVCvqVOn/vznPxe7eY9HvuRuwF0s6VfE6dpPfvITuFV4Jzj2bhS67MYbb5TphdxRPm9M8K95pnwOGoFQPlztoYcewlGlk7sskTvFcRvfg4C4Am4AMQJ/+MMfwtheVVWV8XiO+yIoH4yKMoMkP2XQXphW/vSnP8Hoh0ldaHwP6BHlizvLQABeUFa+QMYYKF9aw+b06dN9+/bFiUOemV6qmY4pAMN7ixYt4N0X1DTP/jpE+QIZnN7HkYmnAIlbDhxJwPsD1MjZ86Q12jESZCif7y4lAzQcghfH99yXwjMOqV9mzpzpuzGAYyez/Rin94X7/8SSAAHrY2kA7Px7LoMScx3scKNx8bhwCAsuhiWNmWXGqvU1YzV4b5aD6THKJ/Ix8HR8vfUdC00t5mSkkzJFxkLmpdPmoS362tlqxbDY9G6xSZ2VcaWx0S1jo1rgHcXPMv65jP1qvctaKiNbKKNaKqOax8Y8HxvfNjalqzKjl7p8rLZ9uXn2oBm6xNhXsnC/NFrrFAWDPXOgut11trJOfB52Rvm63qnuXGozOp6EnQVoKvqhtWB9jCGvLjMvHs+PRy5RPqfrHCtfsIeykotg0mKiJjCOwTcMx7TBusSg62FX+dWvfoU9aNK1OPGZLcn6EolB0mMwLO6wHcHnIV1C6zwD0rLypfvk8Jj+biufNya+K3u6iwysfDK0PCjKJ9njksVSnYa6hx9GI7x9wIiyOQ5wUC00yue09Etf+hL8jQtq1ytAKwbKJ5YCSSufQ/lSLSAyVr4ADxPdlM9jTcM/4XmBeYRjFByxSc5QyWIZb4idL+IU6bbbbnv77bezNOmku36mKl9slE+MnDxb+eJmQSqjH8YGjH7wcoIzZAZPbRnKJ7PhSXeQS84dFENYBHY+a9as8dj5wLGT9ZC6ZlJ1aTM7Qiwx07dFMGADjL75RE0MmhDJg9yCmidXXoezS8ExddXYtVibz2x6zLLHWB97Ix8DS9KwYqQZrRaCnjbVsj7xP1numk5Um3nmiL5xvrZgROytzqBzjLyVtYzi56hWjM6xv7RUxpYo07op03vgZ2xih1jZ87GRzSNlLVTO/Tjla6Wwzy0jZfhiC/wr+/voF2Kz+2urphiHNteoPEUEu7kVNilaZ5nb5DghSuk7FodLmKZOctZX0ijy6oNmJGlkv6nvX6MseBnyNvryN4wQ4vpE/7kDDwPuOqJ8iZQv3ameh/JYJvCAXLVqVVDdD7+afv36wZk+D5X3vgWCGEeMGBGXAovNO4kgxrQoX1DQ4ToF69jpQB0U5aur4fGf//mfnTp1wurkPPIzePajpwqW8glg4eeMY5dgj3KyHOfFQPkEROlSvlTA+lK+YCeRpGMnBhWEZ5BQPti7B3U1sT+GhAZivLMcsYF8vdgonwAt/5RPfvzAIPbAAw9AW8jah3M3ZZm+Tovyydcn8JLYfb3yyivYiTm7HXfrrmI7IEOLDP416FxihFjiX9QFgyCVaXEWGZyyLSM2aBZHMk7t0Bb2Yw6c3LjHHTt7GjDu8ZQM5qldNpWxNFVscsX5lfC21GL6mcP6+rnMIjf2Bct8N6aV8lYX9Z2ByuLRxsb5+p41xukDZtV5Q4mYKo/fUxUD4Xwgb/g1FjbPHzeObtW3LlVXvqUww2APdUwJv5RlGFQY/XsuWtYiCgPgpI7a6in60e3s67VOlbxGdtMkQDJi5S8jig/vqoRYPmGeVVeM4khx+17cy9DMQ+u1+X0VmENh66uC4KEh6gJJU+ahKzfoJeppFZGnfPAwhPbGuXPn8LPOX4hJkDmnefzxx+WhKIRYPu9l5ctf/vK0adNSrRFJW5pInPAXBPH/5S9/wdFv4KtYBhfEsx/n6HB1w9CKa4LvEk+ULxXg9Z3yoV145EOp4tChQ77DwGOOFzjlw+D/whe+gPh+dxOyaa/8cpeqJFG+uDkViGNnBgtjqq9IUj4oLWFoSVoeAqye/KVE3WDxQICrCOJ1XtkP43SvQJRPvuPyVhIjBOa+dA/F6gvlA4zYg+EZd+TIkcSdD6N8+vFtoQ5fTxSBTOR71Z1vVrdUCKOblM9kuvMjRXlrf6lEkGGc+XDCq3NeH+7PyVgfzH1MqHPzO1xURrBRTqY4/bH8GxHZp4SNI1vVRW/GJneOgpKVtYrBgje1m/b+dGP/RvPcEUTxIQJOhOLZBNEKynN+dblN2vZDNWpcPGkc36Wtf0d995XY+FJcmVn/hBcoc/6ESbClMqGduvAN48BHNSpSJojqCcOlHNky9PDQR5IGWFrd1PU24zw6OGXwIAyk+rZ5LFU9bKQfTjNhqhVEWvqQI63OlKd8sMbgTLpAXp/61Kc8HKadJamBUT60C3HAcIaRHIxJ95FQ4wAfwD4mbwu3941EP8KXAy6LyJORFusjytfwKJ87OBCjFDH9CEZKa01zFy5wyofuQ3sxqaFjLmx9GUztjMFJ+kWifPWd8mEIwWesadOmBbLCJ62G+/ENX27Y+up25BPlK6jR4g6uuf7665cvXy4/POoR5RM7HwTXIJmQTYksSxCL5VPm9a9sx9Kv+1r5wn1/oB/bKrwC88j6OL2EHMm+ler8vnp5DyRjMLiJDy6dllDn/D4m81e0qRr3orQ1VGqQsQBaLLGZfWCIY36YZc+rcweDoelnDjE1Tsf0xu2AyUy8Nsl1CKXL9Gj9mx3baJw/qm9ZopYPjo2BCZE5ecLWxz80B/fDZ21Of33HSpgKbQylMkkYJ3aGe3wn3KZxJbqJybfEK3bGZr2IIEaH6ia3zMRC+sa3WdoG4LZtQQ0TeslVP8pTvoJaEURl3I+NpNVreJQPzQT3RnIbyX2ee6HEZ5hnf//738epcRROzz7yyCPChil5xkGUr4FRvqRGCSj9JJ4FSI7/Aqd8TnuRsmLdunVxT33JNgZbjChffad8SJgJeUDn+Vg4y7sH90O8OlL41OGRB1G+AhwnzvJ47bXXwt1DcqGrL5TPaR32Yw8++KA42XQ2bFeZlWeirz3OxFpKJSjfkIeNqrMWsZLEKYBinMpVn9NWvqmVw6DHJDrhz8m1OhHOxz7om+ZYRNQyPgoeZZiRKn33anVWXwUEbHQrdXJH7b3JxrFdZrSKiXNayRgYII7zaK310uGCjjXsCsumEwrHSZ9lT7Q2lUa0Sj+2U31vgjKpk6P4wqMEmdFPGVOilg/RD23hrp61lUgNlamunRrqeAPPnJFEWSfc/W5972orMi+FAVZY8wzAuPx1DWGQC/vD1ZNJmwqn1wC66YpL1F/K5z4HSrVaNUjKhwUCpo+jR4+mNRbYoDIMpKKCUnbBLu7o05deeklet4MoX4OkfG7iJ/x+kSUyMdpTZvwXOOVzug/NvO+++7xTfsu0N/syRPnqO+VDEGyqZCoFtfK77fn4DHHgbOz5WY58onwFODbcIwShH5cvJ5XAiO/5+kL53IB//OMfR1ylc96NJl2lH94Q6vEdpOCToXzR8f+0zsn5f7KcDGl9nYWiMfMUC94Tlj3rDfq3eKBxdp9LsEXwLsM4sk2pGBoF2YOFbUIHdd1sA2a9K6rtNCGttsQVTvJdwaLYZvjMAXXlVGU0RF+eA+WDw6cK4jeSyb3ArVR7bxJSLwij6RVmScsv1WKVMAlGp7RlmRgY32uCrImivzj9w7tRbEJzBBl642k7uNYYxzZpC15mmS2WDjEusf29LXeTVof4FK6/lE9mhcqM8smQSZm7+5aJs2n4Gi2dC4L1devWLd1xgOxqhb8VgG4HnNwkDzeI8uWU8sVRL+/xnNOoIYxbxCaJQ1DJsSFmx9ixY528fL7zMfAC8jMat8ZT/9VXX013UgdenihfLihfgLPDieVLOhGQ1gxuIIGP5IznvvwUQOp2JGhJa3YHOPiJ8hXOmElaE+h7T5gwQUbpqs4pX2aTHc+pWbNm1Vr51JVjuXALIxK+jp2xioG27Eet12SA0yPJpXh6Pbhfau/BNsWziie89bWTmdVOcFBeL3hXqsvGx8a2ZsKbk7to62eZ1edNHWnTpWxqWbeI8zd2FZYaEGo35tkj2tKxTMqFiYKyrA9c5BM/n49N6aLvXmNwwiYY3hUMkv/JOHco3Od+liGDC7dwERdIdzapYobZJuF21xvHt0vUmWWZZzeB/+fWcuHeqX8w3lAQ1Kfzigb5IsrnoIksCOPz/kJ6XGSh+cEPfgBpFodqynBOPNcR9essEL5PSmS/vfPOO32X9cTVCpmykZkN7uZPPPEENO6gLO/7+t3vfvezn/0MEqMI0kiLNojq3XzzzQg49G0ROo4oX04pX6qLo0+RWeuOO+74xS9+AVUn3/GAAhg5OH9B+sfbb79dJINO94UcjCI9WpyvsvdqiEmNfeS4cePyNrNBMiFH8eSTT4pJ7W6m71YAu3nvFHBBLv0prkWUL3DK5/T7v/3bvzVr1uzr2b0g755qGGBDLLPfdRtPYEJHdeBXhkR5r7/+OqZJxpPF+WJpaSkmO7Kuu9P/AFXv8Y9/hVKAzLKfi1lQPJTPjTDS3+VtYXRu1LVrV6jHfe9734O8kPxTAMMDO4rz58/79r7MFMB9oQfxz3/+E7Y1meeXZJnnnnsOQgn3338/lHJxhOc75uOaf/fdd4u0PeijqyITW/IoPkSI+VM+fd0UO4wvgZz4ApZRAeFTaZzZq88XfO9KEx9IS0UfY9dSQQxZgzRFP7BBmdETMXuxMa2ZYsrJfSxgj1vdBOXJ9UvcyrKqMe7HvUxjUX3bcmXqS9zPsyVL8TeSGf0YA4S57/3pLI/fFfsOkdGBvdWPZleVItgSZI9naGDknEl0ss+lTaJjnwWLk2kUULLwvHRSWzqEyZzO76vtf98U7p2BvojyJYUzzw8eMJwPP/ywXbt2EK0WgXZJn47uP+JR3atXL0mjh67rffr0kZHodG6BBatRo0adO3dG5DROjt1y+d4DUECHfHTYbSOJfOvWrRGn4ZvnPW7tw+ZDXMe7I4jy5ZnyoR+R07lt27awGCNkCC648ocO6E2Ux6jAiGrTpg3kJdwxpb50CAOyffv2sRgCoTNcAzP+YgYrLu6FSY3wPLQUhmuZnQ0QQBtfe+21DG4X4FeI8gVO+ZwL3nLLLYsXL96d3QszSHR34tTDhhinMDKDDWUw+7AxHTly5K5duxxf+oznSNyTCAwKoQfw14DOvtj7ymx/P/GJT+zduzfAwSx/qeKhfHGDRx6iAEsC7QsXLixduhQJh0G9JM/FMGLnzp3rWw1JyvfVr361qoqZoIJ9iaZt3rwZJyCgtbBd+z7anObj8eo4elwV6vqN6pJG4BJMFyRBFCTuL9rulSKRt81ogm1UkqsxZmlo2pZ3VVucM97Kt6CveW4/K4aVqvqSunx8bEwJUuQhB4Oxd71NZhghtDlYHuosKJ/FMjnhs4Rh4MapLh4F8yOcPKMW8YOt7zlm8ZvTz4BqqAlTpPguZ6n4qGuR1xFsyXoHPpzVJUxopwqOndy9M9Tjbn3/WtkmOXsa9OKuJRZ/XjbMvByvZyh7wdTliPJlj2FQV8BAQhYaZM6VfGbDjCCEnjye0+KfoNqSVgo+KCMjpg57Vl/Slart7u0I9igdOnTA2ifz1BdthwUJjMIXWKJ8eaB8jmsWrHPgA+nG26Si7hcvXoQ1wO2E5vtoRPryjHeEGe9lfQehdwHc96OPPoL5zntSO20HK8jyjll+nShfjigfuhjuEuIUP/tX3HgWv27YsKFx48aSjw+oIoHsxdUk8GmCMxoEkMeZ+1LVsA6PPIqN8mU/AjO+gnuM4fOSJUtuuukmyeURNjScX3vfWp7yVVZWZtyKxC8mzh1kWoevpox3lXMmcs011+C4HBe/igeDNQ6VJBGBTGSAxtHN9j5Q8JkMj0XTgcOAFKe+qown4kti5dMWDqiB+iWi5k7sVuYOhs9kbGwbbcVE8+JJQZ+Ej6Ulxcl+5+6NOX0JUxqnbOKH5W8qpEGjkM2cr0xoz6rKyZ4Q84S5T53e3TiwEW3hNjeLouqHPqzuenuIZWAHzYNvp4jf4+F8bZooc3oyoU7p5tglDcjwaEuHcgmcnvqmWfymQb6I8jloxp1TBolyimslHtOi4L59+6BPJfPYhnMdVM68H9LiX3Hg5LufdhYdmF+QCQcDw42MmCnpYuJs93E1OOLDbChZDRz3QnjD945E+XJK+dydBb9KPL3c8eXuwSAzCBMHD7aDFRUVcHWTHBUoNmjQIHEd37ERd7t0y6c71D3QwK0/+OADcDkZ6z1OsrECZHP3LL9LlC9Yyuce20gn69joMu4mj5EMrwrJ7DswcQi+l7t54VwZR4fQ3pC05DzzzDMZI5PNF4uK8uWu09PqAlENIL969WokqvEO+xT/Cidk31M/econrHzBopF4FoP1HMcrjq3be2sHVyyhZXAVeB2CxISLoK+Vz8r8ZsGf/k4trX4TD2C8z+xTITciQvggOsKFOpGNXYfs5Lxe5ooRoHbavvXqW52ZLMrEDsiQzsLVLM5Vu71M/+a5+gYape1dq0zqzBw77bg+nsGvZWxCB/3QR0xNVKCrxWJzela1BSFnnpz8pyPc0jTU/ZvgwzZOadZWV41tFZDuhPlUm9/PvCBMH4KdBvCqv5RPZpuYlnyLN5rBrgse98IKCLdGRH3IsL7hw4f7VhtOO1hxZK6GMjBHgO8FMLASLoF24TzPSajoa/FDMolQKORdE6J8eaB8mGjwN3ZSCAQ+NjAqJM840FjYScQhaJavvE1np56Y1KmSo8QtZaNGjcqyddl8nShfsJTPfbUArXxJuxjHapLrPNS/fK0l2YyiuO9u375dsmLI0yOv2BxgDYuK8qWLWx5WSwQLeG8JhMABPE2mTp3qXf90KV+6aGRQfv/+/bD1yWxZUQbiDpibV1naLTxOzJfysTTi+bDsXdF2fe8qFnXGKZ8ucspZH/C5t/7+WG1DhTK+LaxkyvQe5ql9zLhn2yC5KkkgFCaD7kj9FVQPK8Hx3agwi+vjUi5WgB8+jG+nblkESRXUm6Xj6/09IdnC0/EJ984mYXzoeJO6eqzlP5pRG41zB43Fg7V5PQCj8dEsnjEiMKzqKeWTmTxYQQKkfO5RkusVEFnIEMgr84yEUIpjSUs1jkHhvNPyOmCCj8G93veCGc8xXBmhSjIenqgSfIGgT0OULw4BJKyXGRhwgHHbaeMuAldDJIKTuQ7KCKm03I15LLHTp0+X1HTBYfCyZcsyHoF19UWghyQTvm0UMxGPfKeeuYM9FRRE+eov5evXr5/MpEZwKeyBGAB5G12w5wsrt+8Lp5MIQ8j/PCXKF7fDSRwbOR0tCGmB377v8EABSBh463YWGuVjNMc0YbvDkZ/3xlX8K5JqIqoFjp0I5BNyIP5WvholJDhB3qRQcC99w3RO+ZB7Xch18kR883orSM9Q3kN5p090TGtmK5s3xDh7mJM9kXBPABKwv2JAS4Zw/awxj+9Wp/cUGduFxU84eSrjSvUDG0ELlfmDOc2DniooH4vlYybZUhhmm0VHP2NUnbNcRtNnagwZSN2sHKkyw2lvfckrBgx9tktq9jxZnvLBYwS78AJ5SSYbyBHlC2h0pbwMVrQ4T5hUS6GI/PFei+EhKViWx0usNThhdR/9BrjEi0mOqiL4UD5Ycfbs2d51ICtfTq184uKQ2UQ8ek7HPFYhZOWSed5jFRowYIAznDKolfeJRoBjPnELBV1E3zZiJkLtVlgyc1QZb9CI8tU7yucMaUnKd/3112/atCmDuZPxVxDU9Oijj/oe1KIAKN+5c9gv5ftFlM8b8VyvRWfOnMGDRpjyEhdJ9x8hs+nt6FEglC8OMdTZw7dfNFk0EwJp77333lXw5wyXgEJAB7KRhJWPCalZmyxOWvLw0pcNZ9a8Cpj4rkzHV95dmdyOSV+Obhlb8LoZAx01uENkbZRhcFaroBvKHE9ZkCEERWPTezApF5aoHW/u6slyS3TUtiys7nE3OJ5jia1kifi4UGfX2/UTu7gWaOYRlbi7vuc9WPnAqJm77NZ5tkNpbQ9n3Gx5ygc+APM0/JIL4SUZEZsl5XMmba7Xu8TuGzhwoO/uEAVglxPfTVVDBF8hM6+zmnhcEynaEViV8UBK9cXEijluq6l2AM7foezvXR+ifLmmfDhbQdaBwEdF4gUx9lL5Pca1EdE+maVlj7tp/ic1vOl8j11Q4N577z116pQ3Nc1djxDlq6eUD0Oib9++Mk8Nh/LlbQoIyidTN7Ly5W5qy1w56WlaHsYJvHn/93//1/dQAEMImrSQ/vJoS4FQPqeGAj2cKfTv3993CggEpkyZIuRbrLcv5TMuQtpRuErmhe2xLaehVfTlxj34H/bkjp29wU9UxvfaKmUto2Ut1UVvGpVnrSFlSWVaCpn5q6fMqK8twwEU5FSkjJ/YgZn4WMq+55iVD59HNo8OeLC6LQheU5Z7nRE/Zuhj5r721ytLobhtWTIzbSMfLmf3q8x8yinfksE18O0UdsOsubIk5cNARKap7G+XHvypS2NX5NAYj2UiS8on7i9Ud0F34XsAfflp06ZhQub6haMs39UBBSBz4g1pdXU18jilupSATvxE0hsnMDrYjo67GsS74Z7nwUKdDoUkKVG+OATy7NiJmNKTJ6GwlfPnCEaFpOvX97///XRVQxNHEazZULRHskpM6pUrV+Z6OovrI4ehzEz8xje+UYcKLkT56h3lE8Mbk/Tll1+W2TTXuZUvsZLOX4jyBbVHyuY6oOhwr8XyiGQDCxcuzM/y6LFRcW/2MEK8vU4KhPIlPjSR2QgJq2TO35GuGZSvmZPd25/yHd3sSD7aOQiyGQD+3zWjldyTE5xE/GSGPnxWprQD2WM0aXq3mvBlwVGEzStTCuRfmVyVMHR92zJ1bBuetuG5WFlLpG2IDv9LqMNNEOpk+jpgekjM0Eak5msaHvM0MssHURkTKewhecoSYJSDTvcyT+2wHXezvbwk5cMwBeXL9mbBff+73/2uDCPKkvJdvnx5xowZ0JjGKgO1SUnNJZmKBVUGO3JvULE4IrO5zD4AcinBahZ7VAw5GGQQAAv1Ps8jK1+urXyIAg1u1npdCWcTWGFkRgVmIs5fxB438Yq+xjGMKJzaNG/e/Mc//jFcaApwUkOVDjut/MCeeBeifPWU8qErQflkZlCeKR+mpKRjJypPlK+uJr64L/wLkDMd6fJw3oqT2QJcHusL5UvsR1Bo7HxkNmNPP/00U+zkUpD+2i0oY+xbDWdAZnrLF7Eyq88ihI+pSlqBfDDx9VCmtI+WtYIdLDahRHvvDVvCJADbVF3NClNT9TUzmEvnm7BbtoiObBHudg/Meng7Kddh60MXQM0FprnAaK0aNlaXgUiD76nQQd30DmxPgVhxifIljiVYALZu3dqxY0coWaeVSVPmcRtsGV/KB3sI5BZ9D5awDL3wwgt5MOYItJGQTQYHbE28FZmJ8uWa8j3xxBP5WWwx9nC0KePbiTG/bds2j1ol+mPjL7DVr1mzBvkhv/CFL8CHWea5KzNEc1GGKF9+hhz8xmW6D+GjO3bs8KgSnqGtW7eWuVSuFTslKZ9IwyNalKM1P84/EPEFyMnuARFZ+fIz5p0ed/c7PkNfB5rMzz77LHYLyBMgM5jrqkz9pXzYjMFFRQY3yPKB8qXh2Kmum2bpheTLmmZWnmHJGCqYfQ9enWp5T3VG59joVswUNqaVOudFfdFg9th1KGjOHYVyM4PQhtCl2Jx+6kgY+lpEh/wB1leWcp1TcZ51HQoujas6fF1bNyMQSmYp3KhRbf00pL5gTrPlPbXVZWboksgAn+WLKF8cgDACICLC0TN0uz7KzNU8l/GlfJBLkZG6wWEehLCyHEvyXx8xYoQvC0UBUG5s0z0uS5Qv15Sva9eu8t2aWUlnd4ggT1+dIdFeeGOm2q0mWvnwF2gDwNsHinCFzPScriTKl9lASvdbRUv5/v3f/x0auenClU15aFf813/9lwzl++lPf0ryLdlALfNdNyHHZ+x5EEf35S9/2VkeC3mdrL+UD/GKv/nNb1LNAjfmLPCVx4nJxvLF5vXjlM/6ITMIsixjVp1lZM9O0qDN6hob/QKzho1uqbzdmalNLuxnVp8RIp1Z85QsK5vV19lsObI9Nql97LW/h7p8ozZNIhPXYYkZQu2/rlQMMJUAkkeJYzgm1oLsfNvnMwsq+B5+Lhpgnt0v2GBWjampIcrnHHaiYzds2HD//feLOVkvlj8ZyifDQoUQYpZjSfLrwBmKIDK1+spXvrJq1SqifG4E8hzLJ0ZFjkwBcT07c+bMz33uc74DAxNTxu/RqTPSoCPliYz90PfW+SlAlE9yJcmyWNFSPgxjLCMeSVyyBNb5ukMtJk6cKDl3EIuRz4SBTlWLR7HT7QSBz9DNQhB1gR9tuwdPPaV8AnaEz4i9pTepvvHGG0H5WEY+wfp83Tujo5/maQ8c4ZCgpnDK61hWPkb5+ujvdlfGIx8D7GCt1KkdmFUK5qn5LxvHtnAlFMvhNOd1yskNeAN0VVs9Ldznx0jBJ9IkivwZzBJb2pSBH4KgULZkrLb6jPSB8i0Q4ZEi+4V5aEMgtyDKJ3DGio/gWhj3fHOnSD668lMsKMoH5U/fjAhBzSesffPmzfPGR6yJRPkSMc8/5cs133Ouv2LFii9+8YsyE0dY+XxfiCCCNQPuwTLXLJwyRPl8ezaQAsVM+WDoQ6IwTJBAkPS4CGY3soBCkSjuIDVxumHBx8MXkfO5rlLS6xcP5XOaD2dOPPSRpNG3awpnbURN6inlE7ALyuf7YpSP6UDC0AfiB9dBv4i+6LBHzRAyKfFDlrxMILPqHM/I10Mt76VObo/sBQh1Uya00ct7wNtTpOkztleYBktcHghXyUuzEm9iKY2qe96ransNI+GW9RXhfCwpX6TzHdrWxdzCipIsnDL7enLJTk3fvpDBCPLM3Dt7mFvn2aw+qzsUM+VzTw6kIP/a177mTMU8OzZkfLugKN+nP/3pXGRoSDU0P/zwQ99VDwWI8hUC5ctqfUnny0FRPmdeL1myxJdDuo+3fQ9fZQZt9vsnonzpjJrMyxYz5ROr6xtvvIF9f+YIJnwzcb+JDGN33XWX5AMOUWQTJkwIsD7ylyoqyodugikVpldvb1v55U6mpOQY8L1UsVA+IQLJBSH9KV+o173a4Y18uHPykQfeZ5gaEjPAgXN6RyHRqYx5QZvzEjdM2ZnZV481Q0zBsv66dgoCZ5w9GB38a25xZRKdYUHF0S8db4kO/ZM6e4AZCwfYTHZTXTG2Vhh2gns4yuofTGTqPFkzymKmfOJ5gNkBBnLTTTdlvCRl/MXsd4dE+XyfECgAeUb5Z79MyZ07dwJ5mVt7J42VuVdcmfxb+TKoZGZfCYryiUkNgQrIVMj0UeBlsvSSIsqX2fhJ91tFS/mcow1QrHvuuWfw4MFIPgShLMhLII4ugxfynbi/hUvBm3rSpEmQ3cdhouT8Qq1QGIqR6fZjIOWLivIBMSy2n/zkJyW7Jsuj8LgNUpb7pWKhfNyy19SiFn5WvnCHG7QNM62McoFMCL+L4BGrLR4Ml87Y2OetNOVTOzjZGoSVT5/fzzi1x7aA+V2xIP+dMSw1Enu7c3W7a9AdQlMHlI/n4rsmMuiR6Kjm6rhSbf96rpaaNSGzSYmhxfQPp+nlgjz3hminsWIEUiFmT56J8sG55Wc/+1m6a1/g5TNbB4nyyXQEUb641RSuVo5AkTeAeYvwFLsQX6OcqK2vYyf0Wh588MF051Rc+G66X5cZijJliPLl5+FftJQvcQePGFoceoL+3RfEC5eCvj9kw9Iym6MwxC3y0/WJdykqynfo0CHJNFcy61XGZTJbYIuG8pU04ryCsws/ygceos5+ydQVmzLkYx7pa8Ypb7WLjmrB1CzHtdbe7cETssMREY6djKXABqive8uWHAmIDuWjZa57GJr6wWSWkqFUiLXAz7YZEq9Xt7suOuDXLB0Fy9LeXK0YbkYqs+djVveBOyphfdVoaKIqLOchQxKQ1hhq9uZbonxvvvkmRNszOMfKbMHKeH1M+kWifDJ4EuUrNsoHwwUGhsyOM9UsznJ2Z/N1onz5eaoXM+VLOj4zHrTuL2Z8kc9+9rPLli3LT9cXM+WDK2+7du1kdLwTn62Zda738Ej3mkVD+VioGEsGICPfwrwN+//ICF9MVKzO3YzS189U4NJZ1gJencjQIFRGWOaG2szsEHHpW3P5pGXoy11VcnRlmNV2LAn1vs+SbLGj+KpLrwm//LPYyOdYxnm8QfzGtDaPIZNPMLSWXSVySVvYj5E9SxOVibiA8mXf0CKnfEePHk2U6JRhEVinICiPpxSyOWfzQhg9vGvSXfWcGhLlk+ksonxFRfmgOZ6BZAvmILZB2U9qyCFgUrtPkWSGqLsMUb7sn2syVyhmyuceb1n6Iac7vJOWx9R77rnnqqurZTouF2WKx8p34MAB2GDFiZh836EwxHWw3/j85z+fzYYHef+wPEIhPN0KOFUtGsrHo8Wg3SKVkJ2xkaba/rVCRSQg6pE40RwpFu7wOKd/bBQS8bWIjW/NVDqtED7hi8jePKivp755jqkJ86OwUQWQaSAXS0DtNS3FFlM/sz884AEAi3DKKtYRLJYPv0Z6/TA28p8KFEo55YOdUxnZUl1SFgzhY/UwzDP7gB6zlJaLTBjM1gdNl+xZZZFTvlmzZnlszpKuiddeey0eToMGDUKgOUTPkDgum9eUKVPKysqQQveRRx6RD3sgyiemZ33JyxdgNDXF8vk6do4ZMwa7CvndDLab11xzzTPPPDNs2LDx48dDxyibGQ29iqlTpyINCTIBYlI7ATPyGyyifLl9oNtXJ8onP0dyVNKZFFjJjxw5kp9+T3qX4qF88NVPqzfB9OCpC8MglkcovixcuDCb5RHe+1geR44cieUR5E0s1HEe9d7VC5zyBfh09h3AaSh2Qv2f2/dY/JivYycnJI1j09rz5Ocs/3kuX4y5GUe3Q6KTJ2Zoob3dJZHvMZYCx06WU26QeWonl5TB/0XAm5DxLMAXS4/AWbNpXDgeeeNPIHjcxMd8a6vg29n2muiY5rE3n4uOBM1rLlgfM3WOaqlM6mhWM9HULFslWLG+tVwwPWY4reitM8fOXtzKl+31i5nyYfwhA6n88nfrrbdiJ4dQdVWFS222yIuBITIX4YU0nTh+a9++PRJGyx+AkZVPpvvq1soX1FARA4Yonzflg17OU089JcmvQPZuu+22GTNmVFZWYlJnuVa7v86ebYYhJnWXLl1wNO5Mat+6EeULsCM8LkWUT2bxzF0ZZyI0adJk9erV+en0VHcpEsqHVU64NXmsQs4/gewhvHPBggWXLl0KPJ8HAA+Hwzt27MByLZZHyVewlC/Yp7PvGPamfA7yPC+fZVmyEsF5s77qNgj8axLp9xPj3JFAEnYnb4m16TXNWEhd9GZsVHPQHhUmvne7axU94lgfZyw98JOxvtWjIUHJE0gYFvXzhapuClh2SGS8iE1pE2p3DWgel2xpjKDKEOL3QG6PblUmd4yNei7GnVrB+hTu2qqMfkHbsTKAWqMKSthY+Sa37AHVHizlPVjfvB41egBJdYqZ8qF3mjVr5rvQiHmISQjRC0HSAujWKy/hXBMrMnLWfeYzn/HdF4pqE+Xz7T4UqFvKF+xoIcrnTfngqn3vvfeKUeE7iVAS4qvOsUtQPRW3RGBzg1Nt4Uwl8yLKF1RHeF+HKJ/MaMxRGcebtHHjxsiQ5GYUuXjC+o6oIqF8EFOFm5LvwijcOCFrt2fPHl/osiwA4vf6668jM7DkSAuW8mVZ+XS/noaVL9TpFubVCaWQEn8rn4j3q25/nbJiFDJwpFstyfJ2oKChH94aG98+ykx8LZVpHTknqfXntLkfD+rjUi5gffrelcjRxzbQ/P/8UsHvpCUbkrIYrxgIbeytUpFsXVj50BGw72krR5tq1AxdUuYNYebNkS24dgtYX3Me0ddSXVpWo0azrAPyMBjHtqoL+2sVfVhgJAAU2C54mSif7xrx+OOPe+CPo33h1ekby/7lL38ZMtZZdqXM14VxAIoykp5pRPl8x0CdUD5nRMHoFOwOhiifN+VDwhWEiwB/3wgl5KTCGbPMrEy3TFyPi1+hKINdlMxwJcqXLuCZlSfKJzMac1fm4x//+N13371582bRfeLkxfmcWZ9m/K0ioXzwWhe511OxPufvOA0/fPhwHJ4BPsvc3Y27dOjQQXKkBUv54qqR8fiR/GIalC9S9hQyfQvi4e/YyZVFmKFvyMNm5ZkcsSnLfqhr6vLxXLykZWxsa30uzFBJAvmYcY8LeDJCCFvf0lfNU7tZNB93C82B4USyCzyLoXbhS7E5ParbX8/JHqfc+Nnxpti7PRGRyE2UhrbqLUH2eCwf53vMw7O5Or2HefFktvVQo8bGmQw9/gZ0HMBe+orXajhnzvJVzFa+Xbt2SYYR/+53vwNQOV0d3IvpyZMn4VAhswIS5ZNBKf9WPufBiVRXWc7QuK8T5fOmfCtXrpQUo/v73//uzpoY4G4maY+HQiHIHsjYHonyBTtlUl2NKJ/M4pmjMvDl++c//wlSAa6Vn+72vos85YO/z9atW3O9XOQIk9GjRzunyd62vq5duwoenqOaxF0ckMLe62t+xGgMivIhIVBVVVWu2xiHniTl+853vnOVunIMFDsZkZOx8pXA/5DZA2GPUjfMEmFz/BSFNTAgBsiuyILcqi8oY0sVTnjUhSP198ciZi9ZLB/TbnH9vae68k0QKsH5kln4cjjUUg1iW+jGhklXlXd7g++FhXYOKDf8Odt/XVn+Okx//CKcsO54TxmDbIStGAhCs1Q4eY55wTi2U9xLNFO2SfZRF75oXDisLejPOF55T70CbrG99XJmL9XXTa4JYq0sZsqHM35B+Xxfr732Wu4WvqRXfvLJJ31rhQJFS/nEWsZc3iVeiNgMtvviUrF7PKhgdAr21jIDA/V5+OGHMbVT3boB5+WDPIDEiLgKypxQeQm2a3yv9qc//UmmbvWF8p04ccK3yYVcoOFRvr59+8oMMKeMzA47rQvKFMaJDFYnmJuCjZ7NcqTFUT4PZEBW4Yma5e3q6uujRo36xCc+IdNNeY6uhEQCsjLKVMyb8mFjAGEYGJB9L4VxiBjFfHYEBGl/+ctf+lYMBWBjuEo/ui3U94fcsZPROb9YPkb2hNZL5JVfGZVnON8Q5CqYFOGOEKi+aQGXLWmujG5l7FhlnNipzn/5SnaXggFCzeXDqTXRSkGJnFdOzxW8O9hiscJ8WXU2NqNTqO11wl7KfpY2Cff5nrZxtsFyoItKG+z/x3crY0ssKx8je7VvZUO57bYqxfdEN9n0EOntY/q6KYjfExKd3EzKaXNFL2PXEkvxNLsxS5RPZgZu2LAhO5jT+zamAE7jZCpWtJSPH7eYjz32mAxKyMotyjs/0+uPK0vjOsuWLXOEXr29goN9cGKT5EsbRH2I8qUaGAIfOH/OnTs3m2GQwXd79eolM1zrivKJCfLiiy9KmklhUK3Dh3UG+Md9peFRvn79+skMMHeZXLM+cX2MqK985StI/92xY0ecNwl394IaPKgMdLMFMt6YQIMXiibZD786uYIk5YPzJ44181lDXddbt24tM3pTUT5nRPXu3VuG8uFeCPzOZxuPHz9+3333ycw4RvnMaCgy7llua/KnfMzEx/w/myGPX6jD9cqS4abG1B35JEvD2uQFB7dawblRnTeUaZaMek6Z8qJx+kBN5LL+wQQu1pLK1lf7d31+H2PzXETEXWHat0L78tkXtXxTbAqNWCg6oVVV++sZkjwfA/hzuPd9+t7VpiOaYtkEYee8pIxvayVpuJLyqRVDOUz29f3NfMLVVfA+3TjyEWL2kJuB6bWUC8fOHjypfR/j2OYU1tH0cCPK5701FPNz06ZN6cGaaWnnKQixeJnlr5gpHzCW9HL8wQ9+gLjNAHcYyKvhtg97LOKTJ0/OdCwk+R6OQn/+85/LDIxnn33W4xCdrHwQCVy3bl2AQ8LvMJE9e/v06SPTd3VI+VBJSLFL+j688cYbotV5gzHAqYRLNRjK5+DvkBbvYYbjKox/5EXI6QsE74knnkDuE2QFmDZtGqbbuXPnko6WAhk/b731lqT9E+t/sEMxb1eTpHxwQUfYS95qJW4kGc6XlPK5hxD6EU4cMivtu+++m8824rGLhLEylK+kpOQqRq82zqlqx4xO/lY+ZHTg2pIsezvMfd2/bRz+yCZ7gvVl/+I2rlP71be6sOg1CLeUDzajyKRpakc/0ub386V8cFYULqDIQGAqoUI48xFGNv3cociovzJNzpImYcaxm1S3bRYZ/hgMmJy+2SSu1iZoRCe057z3ChMfw2RyJ8GyHbdaf9wtWl5jXDquLx/OAyMFf2aGPpHaXl82xLx4TN5R1OOmRPlkloa8UT7niUiUD/2Cg+FVq1Yljl7nPO+ll16S6T7sobds2eI/9eRKwAUInqIyCzfKIF5F7qr+pdDqjRs3IoOcd5NFxXr06OERJEOUr2nTptiD+oMeaAlB+XxHTl1RPtHWOXPmSG6YHnroIQ/n4UCRy8nFGgzlc9BBAILMeoh9p3ii5Y5rxV056Y1yd/fMhguStUgediDfZma3qPNvSVI+yAvnn/IhQ5XM6PWw8gl4kRnVyYbqfcE2bdrkcxC+/fbbwqvW4xEg/glaX1ex6RkLhfv9yJfvccNUY671wmx9VTx1eHT007BcsUnOMAmE8rE4Pn3HchbDhpR0I1toK8Zzr1FwJsV4f4w35QOTsVVeeusL+ho7FoH1ue1hdTI32Fby3JHIkEfA8USmdYAZ7nBDdEIL89IJnkWQo2exZwEjo3Sx6T0toc4rrXyxcSVG+LLVFinULRusqUWN9VM0FrbHMrBz+ZaeenkPFRF9+MvaiXYwYbY4EeWTWWUCj8jy6DaxBpFjpwflczYrSOrqu4HGdbDOTpo0KZvF3f1dSA5Iiuvg1nDkOHv2bLaz1P7+hAkT4LLi22S4UaGkx02J8gkN3myGRLp9inuJUCvf7qsTyudAAR0FyQ0T1N5F3pp6+mowlM/pO1gtfEcXRmCjRo3ef//9nPaaqJLk/JIsltMK4+KLFi36whe+IDND8UAJXJcr160T15ekfMKxM5/9gnu1bdtWZjPmK9+CGOPPfvazMpeCIfrCBWTPzscLnqsvvPCCDN/713/911mzZjHKh3rF3itjRiceWgYjXphLuXCDXoo3K9CIZRfocH1s6QhTV2ulcIWxyuIhUnQkDhhTVbTl43hCAmQmaKHvXGGFC4L5nd6tQ3SEx54xfc5ypjOZQAJdzp8ouW0+PDxddJTXT7iP5qZHnKZbNBjam3veC7/yKx4Gyd5wiw13ujW2cDACDlkKwcRq2JdQ5w5m6fji+B6sfGOe108f9DSr2hzc1RHge9qupXpFX0eoE8F7HEyBYW99z/KgIClyyofZxeXc2clK3Mv9R8xAN+B5WAolnRyK3LETWh3YnqbqQbeXDpSykBc7kI7DWZ1MCg0nZqy8vDyQ2YrK//SnP5Ux8cEtZ8mSJUT5UmGFrhGnAIH0i/xFnn/+eZmNCChfPt0K4uqPOCvkBXXXMxWLwMkCdmlJDX2BTDR5YDMr2WAon9N8zHoZuo6nHo6E3H1UL/ors16W/9bu3btvv/12mRmKGYEHNDbx8hcvkJIO5fM2NMHa6STPyF3NnVGHDxA0lgzT8KV8qLCvL4zo5c997nOzZ892Ghj4LHBfENxSpMfwfcHpev369VcJwUf9/OHI4IfgcMij9fBuykP7fFw9WUkU6/oNddsi0DFBMkTMWBoOh1f2PKtOtDo29SWQPWbgKmthnNxvZdnDfzTF2FYBuREk4oP3Jk/H5xPax8LV1k0yLp+w9EvsSDmLluaE9lmUkltQq5W1U0Kdb0MApJWMAcSv3w+1ncsQr2i3K8ngFwAq5a+C+kYTWJ8y+nnjEDzKUtbeMuoJ06FglTAm7lmBED4WD8lpHkOGxe/1FrRZr+hjVjOfeH7ZbHEpcsrn7cghlkX8/Mc//gE3ucCfkamOQrH8YV3zXRqE/cp7ncKJl8x1Pv3pT0NCLXeLe9yVYTWVqVUqx04+S9jIP3DgwLe+9S3fS6EHEb6CQ9zsF3fEBN55550yR+lOrSDShW9leWu0d968ed4mPqdWyHZ15MgRonweAwNYQacECaDTMkdkMEGc6YkNoqTA7M0337xt27YM7pX9V0Rtv/e97wnofMc51p+ZM2cWiM5+us1veJQPFlqcF/iuhyjw6KOPxmKxwPe46XZBQZWHZP8vfvELGfRQBlk9p0yZkuvVI3B8fK18zpQfMWKE0zo3NwuqSnFj7+DBg758W9QN556+pjls2HzPRsVJ8QMPPIAIebaJD9q05L4gjpvhtuq7nIoCCNe/ePHiVYIBwZVSeW80THa2jCQcOP0z9dkenk2q+//EOHvQ4npcpsQRC0m/I03j0mmWj66spcqS0TU3qi9bbqPCNFd9Tls+wmIpEmoulhlw1Ujz/BFu2BMkV1CabImNu3UJXWua4cvKnO6hzjdXsbR7/F3aLDLyL8axbfBStZBPNSI4a1bLB0ehWTqqVZyhLwrKt3O1ffekrRDNtPRr0L86IiEX9BfBe5BsEXzPYn0iru/9sVZ0YPp9lviNYqZ8IAw47/Rd4jEPsRjt3bs3CLyTXyNuSYUgGJzpfSvmUD6P1aqhUj6BI7jxH//4RxmgUAbemHv27El8SMuv9Vi4oTgneTunGBJwd+vWTRwJy98rbqBAM/a2227zvbV4jEHVE2SmOCnfmjVr3CrkHk9Z4HnmzBnnYZ9x1yTinHgp/GXZsmUeOnLueiIpE7Tdcrfa+F4ZMYe+uxNnKMJLEE4Q3uPN9451UqDhUT7oo0jKO8FPYe3atYlb+QBnQZ30aZY3he1OxmdEzA7EA0O0Fgt7LtaQLBuS6usTJ07E8a7vcwQFsAq505bmqD7OZRGjISkULEP5EJCMx65zZO/RXhSDIoBjsM3F+MeJ2PTp06ERLQM76oM9BmDhjp2MHRimqkbeeEJY9oShTyq6TziCImcDZEjOwtXQoVGM9WXCqcAX929gpq2RPAV5WctEB0zj1E510QB4dQprlfebpZ4TVqzFA429K7ltTRgkWfq5ABl47TLHrq3rBz9kwXvCXRZqN/jZ7U5l4avcy5TDxOknd+yM52zOX5S5Q3gS9nj5ltjoVvqmhZ6Uj9+D92yNoRm7lmoL+iFmT1A+YeIDesKrk/1c0N84ub2WC2c9EYuW8olhILn8YR5iG420KgLvoNaFuOuIX+FDDzuSzOqAMnBg8B4CDZvyoe2QxHS2+N7OKoDrm9/8psiaINODceTw5MmTf/7znyWVLeKMJJ/61KegwYWjO/nx49QQT6PFixfD2cM9JDxairHqqyHegGP5tm/fLunVAzyx6Q+Hw1kvovEXSNxJw0vK9wzb6V8cPOMsQ36oBFh/UXOQAbg8SS5BKAZnQsgFYYIkNjzAugV+qYZH+QARjAmSW2fYAyF0keidKLM2Bt4XBXJBpOr11fd3EwnE/iELXGJcn0MCC6RdTjUwtb/61a/60lpMajzpcFLpLI+5GxVgRGDOSIwuueD8/e9/91209+3bJ3NCKu6I4w+43J86dSoXyxc2jSNHjnT2mc6DO9UTHJsZIZzDFDttRz5T27Mq1O1bPDM7NFrw9nPsRLAfMg2ITH1tr4289gfYtRjHsIL5MrSh6R++A9UWIVOpTu1mq5oAN5H6D36oqrl7ibpALk0fE3QRPKcnBF30j+aYl0+C7gmPx0xIqfdsA5+MVCorx4a7f7u6lNFmBmO7a8MDf2EcWFejxqy9fa1YS9Iq4CqsY7Q5/RjvtVKxu4jf6Jb6R54pXIR3LfikUg2+x1MaMt57RRY+YfFjXrJ9tLUTjQgCC52QzAz7zsGmaCmfQOD73/++5EIDF1D4mmPrloujLzZnDAOJQRH3Ba88ySqhGIioN4Fp8JQPPh54hvki5qyw4E5Q0di/fz/8mlKtEO7HG7ZEeBjgDBJGQnnTR2J9MH5++MMfLly4ULiRyGwFMNIQXoKYb8Tm+TYQBZyzZ4+mifs2YMoHQ8evfvUrGbjEw/6pp54CS/RFTKa/nDLOYQE+oD5wmYZGovdG3D204JWU1u2CKuwMSxg/4eHmO9rdBeAuAU47aNAggIkmF1SW7VT4NCTK5/SdvFwhzoZgoYXZB1zd7YMguToFNeoK6jrAIa3nL9YQjHycJA4cOBAjH2t7HnzFs0EMNbz11ludh4X3OomYXnAhxAgEuzyK+mOYAW2sFch4gUM639VG1Bkv+CD4epJj/WnXrp3k2QeujKczpFzeeecdPOsDCdFEDdE06I1D1jgu8t/7YBphIKJ1sPK5XlpMWTgk3O46kTVOxsonTIK8ZBPQv9iE541LJwQxc1n53J99x5WpLhrJZSp5Hva3+9R+wbIa8osrYW39W4zAgNGJiD5E9yX380yQeFk6RD+yEenI7Ss7ajNSGyZ7YLHRJbxi3fZM48SuyPh/hNpfB8teuE0jZuJrd23s3Z76pZPCoJpg/Ex+U+F9qkxsx3FIVHBpZTDKF+eJaV3KMiDiZuFLiGNUKhC/V2vfs4yi5aB/LB0fVFtgADSPbXLQ8O0hmQJFS/nEORxSBknuDsWKg803iN/w4cPxZAVtcMdoyaCdWAaPW2zr58+fjxO1n/3sZ76OpnHnc2PHjvW+b4OnfGh+ly5dJC1gTjEQv6efflr0I/BHVya+EEKN/VBpaSk2AdgeyY8Tj5KC+KHCEARHxgVxUzgYOz/xAfVB2OGrr74Ko6Lj35v4nHCPBOdfcUSNGAzfodiAKR+el46hQ2Ybgc5CyCgGw+uvv44jcPSFY2HzhTFpAewYMOlwUgthAORev//++xFHKjN4nNoiOi6zWwf4LWjbyAiBJLYL59k4R0MWExyswAKPkVyHL7hDe2DSkCif00xsdsFAJEe+2OzecccdWOVwqgXRIMgR52J/LzMy47wqZL4SeBnUAWu+ZKqGuMEPgoQEsM899xxGPlLD1eGwx609khKJUAjJEYJizZo1w/KInQYi8A8dOuRrYfPuFFDiY8eOIegUyyNYGRStJZ+tosJwj5RMpocgDsgyp7UxwNMTcwcZ4fEMhTIc1nD3AzrpJiHxj6B5GEKiaWn5BKGqWHKxFRQAXkH5GGOIVkfHPhNi5in/WL5ETsgSzQ19xDh32OZDtQxH9ozHNJQ5A7lpi1E+9e1eV/a05Q3Jrh+6YKwaJVifeIPy8fxynq6eCGND8vH5/YwPJpjnD8Lp0aJwwsfT5ZnqNcKsVeQKK6FReUZZPirc7a6qUii1cAmc0mvCI36n7VyK9BL8anYoocSKwtihErZSsSfIt0TLWmmb0YW1p/qWrVZQSuYuquontmorXgcgMOKJfOtXvh2semrrJsMYJFGpNIoULeUTGGFbj5NOmQ2ZuwweCZicEALGAoQw7mxen//853EdWPPlT6ScmuBsDCu794QtBsqHpfmmm26ScdyPewCgH+FyCfzjXvBqw1/wT9isSz6QfB8t7kcsromHATbHibcWf8G5oFupxddhVdwdxRDngEME3/nfgCmfmNRCbF3+BejEYEDXZzmpIcuW8aRGNdD7jgO5bz/mrgCYw49//ON055QbcCxoGOQYyXX4evDBB61tQzK7eoOkfGgvjgycUwb5nT3En0FaMP4xgLN5omX2XUgWuQez7C40BxMAdh64CcThJgmjGP+FMPJxXMg3skkMFU5QcVqNwojCAyuo5REjDXueDJ6tUGuDoVKm29F2cG8nok/mWeAsd84GT+wEZF5OSewMse6lha14dmPYNG/e3GHUV1r5WF8axqnd4QE/DUFxRNrQ53A/8ByWrO+1x/UTO5ixw9FJsVQj/XmFGa1S3+4DygffTpZz/O3ecd1gjzVuMTt/SF0xgtmpLNlJBKfJaHiyhAQshm3JYGPb/JrwBSvvn7yXJ3cxdUY9sjDoB9cjmrEaxj3m7Ir0FU2h0qmU9zMuHreUWkRYnbDzyb3M80fVcW2hXAoxm4Q8DYjlW+CKAbRcMnEDhnm00ti5RF84AM00mDOnFb93JeXrzYIhWYjjIOPcIdk6ydUcpYqc8mGCwTVcZn46ZRI/yKwmSctkcCnnK1jLcPYG3z+ifDg4xLGcb5LTjLspkC8mHT9xNluHucXdUWZ84ivYW0PVU2bqN2zKBzubjF+i95RM1ReBDIZUt0ZHt2jRQqYH81AGW0OwX8mxlztYsrkyWKvH8thQKR/8cuEelsHDJf9j3ulcuDN4kPM8jHb3LeD64Q7yr49T4He/+10q0BA/8qMf/UjQjGwmV7rfdZ8fZXZrfKt///7yg+Ho0aMQH076kE2sfNIqSX5X8mreTxzYJBGCKIg6XnGUTwSA1egH1oZ63eMby5fM8xNSJSx6LdLjHm3r/BomOuQ4M8rRiuoL6ozecOyEYifYjsopX8I3HX9K07xwRFsyWGPCJEKVxJ/yWSWZRyi3EC4abO5YaFaeFsRPQsXTMqSxhhm6cWpPbHLr6o43VZdyN05YR9tfHxn7T/3QBtPUbB9L+2REDgMx+PT9H8bGtI6yZBUJlG/08xpLV+iE3vEPqLoSNQ99qC8fzqAQlk8m05IEE8s6Or+vvn8Vj2z0Z+PyU6LIKZ/YDcCVSz50OPCHYmZrH6qBZyRq7vS1WCYSu74YrHxiGD/xxBOZPYTS/VY25VN1t8cwkBwhOIWFb7CzbfI+CGjYlA8gwND3pS99Kd2ekoQ6rcumdU0oanh4ZKW1sGdfGCcpcHaVSUHpwWDTan5awMoULk7KJ8Z/nFebDFx1WEZQPvlz9uyHt/cVEJUqvPLqdgBn3COgfB5gwnsQxtgMmpbBVzJuQuIXEYUICbS0Bsn7778Px1TfOgTSrmw4LQ7X4o5rExw7rfgwXdu2IPziNyTD+VzFIPrCKF9VaZPwS3dFFw1BTgVO+hg9gXak74wyL51RpvdgSRq4oU+d3sNtT+OzV/AyES7IfuindmrLhwtblr9jJ/f/VBHDVoHC3MmTGQl76cuG6zsWGZdPQRvGv5KM7GnGxRNK+cvhPveLJsOyVw0Nm4E/V9dNRxCdo8jJrXvc+ma5jkrRPhTXN86Ljk4ayNcyNuZ57fDmWidXXFuLmaf3aGsnITCP8dgKBDfaCfeSe7r2Blb6+ik1GlcQlaqULzC1BYrcyieAGDZsmFvY3XeBqKsCzpkTnkaIQnTWPmvEFivlE81HmAEEymX8Y93reyBrPcYDGBecS33V3tI6MvA9YnSeMXBEgQxJonBcqoWgwVM++CVCUAF0RaZ/U53vZrzhy2yAYVJD261whE8wreAEAasjXP6yXPFkeiHLWyT9etFSPkS0QhUD7nNpoVpX3SROMN3Hl2nsYIIuKp4miNKHEdg3tN7bbpMW+MEW9rDyoXUI10TAcwZ+lQFW0vfpFncvkCIkQvQVbokbDiiPsHnwW++aezwC5JvsbpHMVHLKwHF06NChTgytGIHxjp21DdNVde200Et3CT5TzZVIQtZPkciB6bWk5oRcrLKkUbTs7/pxlobOphUi6CwlxTBO7Vfe6gS+x7wZYesb18bii8KKdSU7sXgKuOSp3doi+DEy0xbenP5x103uu5jsnSLkb9FAfUu5cXoPsizY9NKWCbUqzqpuVJ9XFg8NDXwAzNbGp0mo293KvH4mU6+x2pcNjQKFU5eMVkY2V5MpdqpjWptnEDApeKVuntqpb5iuQXjTamwysx5SMpQzlmsRwnk91JUjmW3Tkp8JeIUjygdAkYO1bdu2zvoeN11lZq/8upBuycS7g9VAQgbuGTJDoUisfAIKRGxDHMVBOG8dhz1xq1at4EYCS2PS0P9c1wTCrRjDMuNBlGnwlA9tREScpM92ulMyF+Wx/frb3/5WOHzPGUsYV23atEm09eV6SAcCMiifx6RoqI6dosnw+cdRpncWosLpRDflk1/HcloSIx/nHeIMsXCAkpwX3pQPuIHTItK1XrRLVBJCndisZtDj8POfO3fu1772NUnoclcsjuWKdmFpxfF9Yl7T1JQPGCBEbdeKUI9vI5M4Mu8x1lfCMjcgXI3/5E6MllynEO1MfKNY03DXO5UlI5gBTRj7LC5lWb/isDZOH1Amd2YZGlgsX4vYyOamGkkmdFn7PX5Bwzh7QF/5hgp7na3gwpU8Zfw84zhhbx3pH1aXGftXGpWnalSMBttCp0MTZUds/qBQr3uRlIIF7AkEun0zOq29fvYAyx6RwdhJ+ApD6dJJZWaf6KiWPBV7Ql6+MSVm1TkDAjYH1+mrx2oLXq7NsJeM4nJ7Jpw8mVVT4KMvG1Zzdj83lIp8EIFUvLYlRPmcByR0FKHfkLgIFtSyCF76l7/8RZLvoWlFRflEe3/729/C5pbTXnPWbnzApgrCjMjSLp6jSOYbN4qCrUnc1RAs3rVrV2SqSGs9KwbKh9USsEB+zaErBTu1cUwAiUt5I21afZ1ZYbfjAMb2a6+9BkfBzKyXudtC+V65mCkfehD7SBhG0g1b8EU1FwUKkPJh4kC8F1HiIpFdXKuDXdUDh9SX8qF1WHCgLypsfYXWHFEf8ZzFE/bFF18UT9jMXrD1rVu3Dm44jhNQ3bbXuTuqBCNkUirrSfmYmIuu7VgS7vsDbrJrEiphyRgQrRdu0xgGLv6r11u4OzJe1O66yPDfapvn1egaLuph5jPPHom+1YVptzC2w9w7zXOHrUA1J0O5q3/4I4Ql2WOXvXRCf+81RK/xQDWRdM4nUXtSAyAnRcxaqC0boq9/yzj6kRm5pJ/aG5vdLdQbbpzIqw6bJ+RtmoY6fD0y6q/Ium6yhHtO1FO29Ik16fgOZUxJbCQcO59LkqRhXIm+Za723khmz0QcI4/WE6wv1RuwCDQY/Zv/MsvKAPMgY8uB0z3WPUT5nEGK9b1nz55whinA9V1UCdakZ599VlKxSrSrqCif2KfiSda5c+dUqllBrfXiadS4cWM4UIlUjeLueDLBSQOBZM6Ngrpj4p4D+yS4AmaQKLJIKB+6A4cjOMpJTJOQo07JYOuGQxykXhTZGtMKU8ls95PWt5z64KQcmhZgUG4h2QLcKcbhX5yUzxlI+AC7MXKBIi21jMd7BqM3qK8UJuXDZMHIX758OZSQ3U6ehbN6pMLfg/K5T3NgyYQIXGZJKYLqeu/rAPYhQ4Y4PiwZr5Bgfdu2bXvsscfiVrD8tMJ9FzF4gPkDDzyAvBepFmRvymflfTNO7wsNfjBUek1Vm2YgfpYrI6Qp/ax8ImVfdZtmjCgyR9BroyBIWytMjW1lkrINs/q8NqOn0CxhsiUjW2h714qiSV+1UiosVpCLVW6egzA2l4dneqyPJ/djDIrzxt7a7K6xMX8L9/9hqO01IvcgawisnV1ujU5saRzbBlbMdV9sp1NpQU6vZ6RZo62aEitrHmWUj1HfOMVOZUKJBqudcGQt76PO6wO+50n5eOFywNJLWzTQOL4FdXaEdXLB+Yjyif511kHktEFuFvcDsg7Xd+fWOIq78cYbke8lXdevoqJ8zlTF+o70OL/+9a+D3ei4RwL87yGEffDgwaTrA9L1PPzwwzD35eiJgrv/85//FAJfGbwaPOVz7wywaUMqJ3eG5Tqc0XHP/rvuugv50BDFkfFWJoPel/yKm4KK6gFJ1Bb+YGKPWCAwekyx4qR87v4VHXf58mV4Ilx//fUOVuLEqnB6sAApn5sawV46ffp0CF3i1NUxQOVobQ/kst5WPnfTsANEFsHrrrsukPsGeBHsee655x6Rpy6o5REbA9GPdTj40a4777xzwoQJQm7d3RfumetD+YRLIzM6nd0fndYu3OEGTvmQcU4IlvhY+UKiMDP0cY9Q9m4MshQZ+VdtU7kRumiJZLpqZIQuITEDz8vH+B4cO7W1sx2+l0j8uMaopeNisUhkad+9XFs4kOdnd8Lb0iB+4EXqO51jU1+IDHko3O2b3IeTZSm0jJYdbwoP/pW6dLhxbKsZuWjGQsx0KbQzGfHzl6jxfTRCG0aZ8pIyslWs7Dku19k8nvJN7eC4rcJ2B49Nn2yEdqoGdeEgE2nokY3QjoMU1c7WLpnQJKJ8bsqHz1gUjh8/jiiIG264wcMlLMClzftSeMYg/x7MjwhUSzd2Gc0pBsqX+DwQf4FfH/yafvKTnyBLW5ah6m6T3X//938jB8DSpUvF0WOqpxG2WeXl5UjOi7sHtbXCkSfyhsOzFwfP7vwc6T4RGzzli1vnMHGOHDny0ksvwZEm3fS4uZjpmNRIRAbBcRwZxE3qdLvS9yGVZYHE+kA0b8GCBYg8bNq0aVKfiFwgltk1i5nyxW0lQVr27t0L4oftJg6MglqRMuuXxG8VIOVzJo6D5Llz5xYvXoyA7WuvvRYe9YWGoRtVX8fOuHm9c+dOCD5jp1EIyyOOSu+4445XXnkFu5dUjCiDZU3wK7zOnj07evRohP1D1iXYQ2GP6YDRAgUamBPgFoRk9L7rvB/lcwOga8qK0ZEed8NqZ/E3lruPpe/jXo7JGaAw9Il33OfoiN+rS14zzh8WtM2K9NNiyuyXuX2vJXfvbK5VvM7cJpkNz2FTtpTLFa6ewkuRkxdDN0/u1JcPY+KcPKJP6HnCcIcPwo5neTlCsZPbvkS0G97K9DaRN/4Q7vXdUNvrYM1jAi2sgZy1droxOuRBbVqJrYDSS1vyqrZ6tLH1XWPvezCdGeePmKFLMpqfbsOaZR+E2U2pNi6fYDXfVI5k62g+AvmEic8xe4pf1Zld7AT0qcie02oQQp6CD7RwyavGsc01JrLP58ab0zVa5CkfJEMymGY5+sp3v/tdmYfN448/nnEFBGF48sknIfKb/zMh3BGK7cB8/PjxMsm1UzXTm/I5Dy24y1dUVGSMVbpfhD+DTPeB3sBYl+7F48rDLrpy5UpwZnC/dCXs3JUE3br//vsRVAABdPk4chAz0DPIo+EZJqnn6T6JF7YUvBo1avSb3/zmzTffhGXP92nhi1i6lC/7O/pWCQVWrFghGXeUcSYDPOxHjRoFXybE52R5CuA7gN07QtGJGABNmjR55plnMKnTcs+WQS//ZZC0GuZTOFFjZGKqiqzH7sMRAVH+F0+na+IUO+OGcbDyLTIEABs+d2ad/HcZ7ggVVhxFde/eHeAgf7Rt7ctrijZfypefBSdj/LGMvP322+BISNouRr7vapDnAr6UL2nbQUXeeOMNBMNjHU4czzIjXLKZSS8F34FbbrkF3isIb8O5Usa9I/lFbAzWrFnTu3dveAOB+6WVtD1VM5O26/Of/zzGSY8ePVavXi0fkZgO5WPakJqORO1lfwt1vFG4ODJXT/ZTWPCYSmda73DpNUz7ZPTT2sbZSAFvKmGWW658MJgeY32jWuGDOr27ee6IoHK2PcrRG3E4H+cwwsrmWNuQKWHLu/qC/kKwhJvCemgsKwPjfoIy4YNe3kOd2UGZ9Fxk+CPh7t9EVr1QKRLKM7dVxOwx+x64X7e7Yq//Xp3VSSvvntSkxpwqkeNuUX918UBt8SsaEsR/MFH/aJa5bZ65fT5EVtxv89BaA9nzdi83ts4ztszV103WV4/SlwzRFg9WkT99fl9lcjvO8axoRpvyMR0X9vexz6tzXvI266mM43Fmy50/mVfnyjfNcwcR8RiIHdJ39OPwD2MRXsW+L4ieF85CDIHNn/GXd7UhheSLgHcB+FyB+2GfjSUeR6TYrmFB9MUqgwJInouD8+bNmw8ePBgbKdwRq57j8ZUx8thZelRGAIifjzzyCJhMlljJfx0WS5la4Tx1+/bt8pf1KAlbCiQckaQYqzw23OBg2Pc/9NBDovmpKgPXfwRPokdg1jh8+DDUWTIwtKJW8IiD0Q+Gpjlz5vTr1w/XFLf2eOHWGGyQKZs1a9aOHTvQj9irBQIFxhJGF9QsZUYphr24acYjUL7OW7duxRmNd63ErE/lTytzLzQEjB14AlU4+YDDYyQ8+uij3iNBBqu4MqgnsoZgUmMTgwR3mNQYQlhM5M8LZJpT52UwtjEygScUaxHvh+NzsU4CVWx0BKq+C3UG8Mp8BVpKHvjgAEXmImgFZq7HdfAMxaW82yj+FSQTPLnOuwxTQKyHoC44PcEsePXVVyEghEQvMoDkogwmfh5WmGCRB4aI/8fIx/DAbh4jH/xBwAjn51ygJH9N1CSzxqIXcEwJeyYi3+DF3alTJ8xlnOkEuzyKNQEPQWCFSGZEv+NgFGmWEHqd5+URnYjlS7QXDzvsctFe+NHIQ+0uKdb8P/7xjxgG6ALMrM2bN2PXgXGS7s4hPconrENm5LK6ZnII+eigX8I9HoVGSwaUT3hLQgsUkXLh3vdHxzytrhobndMP/pzITxBlDo3N1dEvGPsQzldrxONz2PkV1j9B/PhfhYuisBriA7LVHdmoLR3CbFxgd1byhp5M8mRud3Ua9FEej/b7EXJRhNpfByNkVRuQPeGJCptk01C7a8N97ouOfVqZ1YHRJ0YXk3hRxqllMn7Fk7yLvPAOvbTtckJEVEQMWm9hhBTWSHVO19jYFxxPzlorH5MwZdKdLJBvLlLPezlz2lXtzYycFb2NdZONyydda1/gjpzJ14EAreeZLTS5+FbgjxAHpWCvHOzVkiLpe4u6GgC+FQt8YDh39G5y9hVzXyGQq2V/kcDBrPMLZoOJewAEO/5xtVS9n02F6xztbCpQtw0P5O6+F0lVINjRlU0vON9NWiXfBgZy61QXqdu7Z9O0QuvfLOsT94jMBhmZ3Uiub+dd/8SnQPbtzRJ/UYH0KB9/4FjmNqPqnLJwWLj393hcH2xidsaCNA19wjzI7IQlLPWf+LW6w43h7vdGX/5FdNCj0eF/iVUMNZEhncmzWE6gtmuiI0FiMUCRqI43jRNAwf8ilfqm2eqs9srk52Jlf4kNezjc657qDl9nN+JqNEx9tBSfuR4p+Ge7ZuGe90SG/Uad0Q5WQRbaJ/iVcI9MSPzAc7uzN0vvzrJEcNsa188U/qJx9MzRWcEHcWWmDoqQPEZKe8beaqfAqVVItnDNUov+2R6e6rROEskn+E3LexmLB+m7l0FjhuPCtVJrlUWzH4RFfYVsniWJ383mah7dkPSygawdBdj3cXvivNUwrb6LexRl/2RK6+4Ck2B3Zrl4vAXVd/LgyJf03VwmErPsL+70mvtDUCgV1HUCwSoPLYqbRM7iU1/qnweICmSsiq6pX/2SOLry01+5vkuudz5xa2+ddH3gIy3xeZ3xLdKjfLU8SjAsQzNO7VUqBkX63B9qyzw80/LqFNF9IHu2MkojuFOCdwlFUCQDZKneS68Ntb8x9OKdkRG/j0xurS4ZAQOjum2RfnC9ceYA3vqZ/calE+bFE8bF4/q5g+yPp/drRzbr2xcp66Yqi4dFp3eIjfpbZOAD1S/eEep4Q3VpMyuvINeSERUWbpzV7a4PdbkjOviXsYn/UOcgWI4zMaYBw1MgcO7nmfih1moH+x4ngfHWPFcAnmCP3NfUDjhkN3qnWxR2PCd+z05WIfI0MK/O0a00P69Oy4o4v4+OZOtnDzC2bBF1rtSSgyx8GS8TGQ/cjO8o80XJx0OWlc/dmp7WqpplK7zZZu4u7t2PHvcNvEo52t9784rAW5HN5kxyvshMvVyUyQVWuain7zWdhqQ1wX0vW7AFfDvOt0Dempa4COTt1gV1I48hWlD1pMrUFQJJtz2BT+Tcba58cYtrSyBNC+QioubpUj7xLZdbIKuLqZ/ZF1s0NPLqQ6GSRlylUwi6cPJmhfyxXzm142nceSZ34QvK0v252JdNGgV7dH4K99EkgYKgcDDZhZEbvfvdoXbXwF5XxWVCOZFD7vg4ZVFHaUaU4dXAFbreHnnll8r4vzOmxwkeZ3qp0tx5pb/zVc5MKGAxQ8YnIR4zua0yyqXP6SRh5+Y+OLvGxpdws6EwJ7JKciLKuSjEaWBm5N6nxsJXjD0rzCiT/rsyXCZPLp2+E4MKEAKEACFACBAChAAhQAgQAoRAHhDIjPK5K2bH1EFwMlKpbVsUmdgq3PteRMHVOnxaJBAxcpzvCSpYwkRf3CwunELzU8ZyaF2We2baNkOhLgN/UfaTvwXnZOkiYDwMd7sLspxIw6BMaaW98yJEXCzuxJmeiMfzTm6ePrvzSaUAHc7omOdZagoes+d+w+7HPDxHtdBmdBb3Ze6jIkCRc7za98L++kdvmyxyjycMZP3DFW3oRQgQAoQAIUAIEAKEACFACBACxYdAVpSvNmaOWZJ4OnKmmKlpJ7ara9+KvvHnUOdbquHwybM4cFkUi4bZCp9pO4KmoH+M5sERlHM87inKwwJDsCsyEohQPZZigVnz2l0b6n1f9K3W2oaZxsEP9X2r9VUjtfksdx8L2AODgqyLHa0XJ8oSOMGLD/Cb2wO6LDwBfTzfY9yPe3gqY19Q33kJSRdYhS0jJCx+MPT1Yaqk83vr748zTu0yNeS0EGI2IvzxSsNs8Y1yajEhQAgQAoQAIUAIEAKEACFQtAhkT/ni5VJErBh/mUbksvbRu7G3u4SH/zb04jdEdjt4forwOR77l9JjU8a458r414yLrzSzhUPB/TjD7PD1SN8fRkf9NTbzRRV5IM4csPmPkPhEOKJhnDtsbJqtL3mFh97VsinGqcTbUxszoH/tqU7toJTZCegTrXxw6UQ6vilMOJQ7nVpBgFZA4KKB+obp5tn98LG1E9M7Q1qYYelFCBAChAAhQAgQAoQAIUAIEALFiEDWlM/W0LSIhZUdgbM+65/A/DSj6ox+Yqe+pSL2bu9I2VNI8BDu9s3qzrdUtbsO3IzJtNhel2kxPSYAw4177N3xxnDX20I9vh3ucz/jeLNfUtdP1Q9tAKMzw5dB7WwxTyFhIjwebRsYBE4unzT2rdRXj9SQW6+ij8iyEBCd8/HnZH6kM7soY55n9j2Y8iDfkkD5kJ8wNoan4+NGSJFqj1Vv2VB90yzj4rEaNWqJ69j5KSzWbeuWFuPopjYTAoQAIUAIEAKEACFACBACRY9AVpQvGXqOs2etlelKx0LOBpWwcXKXtnOZ9v746MIhyoxOkdF/j474fbjvD8O97wsjGTqUVxyhF64Bw1VhuGGww/WRXveiWGT4Y9FRf4tNLY2Wv6wuf1PdPFfbu8Y4f7hGjfh2a0qrF7jfmT3a7iXa8hEslg9+npZVjWVTYLFz5fgg1FNYpJ9I8m4pZF6Zas9S+2TfuiIFnyPaqTCpFbx7q+92j41rHYOJz2J6Igk73q1E/B7PSt9cndxWOHAyslfRx1g7ydi3yghfBKVOaI5jeq3tBV9MqAAhQAgQAoQAIUAIEAKEACFACDQ8BAKnfGlCxCxuRo2mmEoY8pIgMEbooll52rx82ji+Uxn5z9hrz8Reezb6+tOx15/BT/ahfLBx8YRZfcGIVLFv4buGnuZdPYuzcETdVCLmxWPGzsXaB+O1Ja/C7sdTsfNs6S4KZ2XkY+nOhZgKV9FEOgf8yqkgj7tLmpcPajGcCs7tqU5qEyfWAr4Hcx9LzACjn2B9Y14AM9SXvqqvGWfuXQmIarSoSD5vZ6IPEgO6FiFACBAChAAhQAgQAoQAIUAINAwE6pryMRSFFyhjWpbmCE/8gGxyasVwdWRzIVYpLF3M9XFsib5nLUdfuGpaJq5ARSmdKpmwGcJt0jiy0dg0R1s+HFovTOWFGfeE7Y4lT3f8PznrY16X4l/B9+zUfFf4dvLkCizxOr6oTG2vlLWKl+i0yB5PxMcz8mkLhpmHNzAWCgfOK2x4LggaxpCkVhAChAAhQAgQAoQAIUAIEAKEQHAI1DnlsxgbT+Yr7FUiry9i7Uz94Eex8aWxUa0Y7RnJVSvh+ggWNKmjcXKvVdDhfBYLzB4bTqKYAKnFKW1ayZlppNI4scPYvUz/cJq+erS6eLAd9WdZ+Vz0T/h8Jg8IVFkG9l7q1Pbw52Tem1cG7/EUfHgLE19L9Z1BNZFKpxoCK/sl+B/ps2Tf73QFQoAQIAQIAUKAECAECAFCoAEiUMeUz9HPtNVfGKuywwFNM1KtzX/doj0sdQEz9AlfR2XOAOPSKUESHYHQuKTjGXeXQz+dwETbgdKWe8F/daUmWmlWnTYvwQa4Qd+7Sts8R/9gor7idX3ZcG3pEOaEuWiAvnCAtniQvnig+60tGqQt6I8IxtjoVpYDZwLlw9+5ia9ldOwLxqHNFrFD2J6lzsJSYtj2zYwbSl8kBAgBQoAQIAQIAUKAECAECIEGjkAdU74EdN3WKk4DT+xlZr2RzzFqBCnLUc8xFgQLWFkLbf5rsLnZoptu7pefPkttWIPvZeRyTdU5JqR54Yh54XDNhaPmhaO1Py8e1ddMjYxrzbjryBZqkkR8LaKssbD+tdBWTUGso8uml5/W0V0IAUKAECAECAFCgBAgBAgBQqAhIFBolC+RA5rK4pFMyKSMi1gyx04rui82+nlt6Wiz6rz1HZ5lr8D7hNFENaZvXhwdVyqsl0yppVar00rCDlprRTC+M9isPFfgjaLqEQKEACFACBAChAAhQAgQAoRAwSJQ6JQPHEk/sVt5q7NqMT1h6+OhbtwIpjJbX5UjAeMKcisgzGv9QjVVWzk5Ot7ie4zKspi9+Fg+S6hmTGvz+G4K1SugjqSqEAKEACFACBAChAAhQAgQAvUNgXpA+ZAyQd20iMlaiig+6LjwvAXRMsT1tYLdT31ngHnuKAQ/OfgFauiDHox5/pg6fwTMeqwVLOde89hoHp2YxMrXIja2rbFjpckyyJM4S32bVVRfQoAQIAQIAUKAECAECAFCoGAQKHTKZ5nvlIg69xUm2snkW7hlTKStExY/vN/ubZ4/GnCCvuA6ydRixpFtyoyeCEFURJAeiB+Meyz/BPNTjU/SUNZKWzfbVFjmvUCTTwTXJLoSIUAIEAKEACFACBAChAAhQAjUBwQKnvKJ9AOQqTx9QJnaDUonIjW5ZRkb9RxPcgAe2FyZ3FnftryGJVcQLImH9vHUfbbES+6yGThSLvEimkxQNFKpvD9FGcucOaNMdNSpPI/cE78y7gczJqOv0dGtlEWjTF0R6py8+pSDoT5MJqojIUAIEAKEACFACBAChAAhUHgIFDzl4+n6eJ48Q9+7NjquDZc8YeyICV2KiD6Wvw7ukcjS/oKydJx54ZhlG7MMZIIw2fwpB+xJpIS3bmHRTThkmjVKRN/3oTqzT7TseZUb9IRZMtGTk6Uc5NbLWNlz6uLRRtU58FXh0pmD+hbeMKQaEQKEACFACBAChAAhQAgQAoRAbhAofMpnZ1xn7MfQ1s1hSRq4VycneywWDkF9sJ7h71zipaUyo4e+Z62pq5apD9zJyuDgWM2CxtK+PudnnKMhBu/CUXXp6NjY1oKRRrlBkjtzWrKcLmdOVJ47rOI9+2UzdMk2U4oKE+kLur/oeoQAIUAIEAKEACFACBAChEDRIFDolM9KPG7/x1Rj6poZSM/AtDq5fY+nMhd8CbTKStTOjGlz+hsHNppqlH9Vt9w7c0Sf3BIruqqf3KutGK+Mfh5he6hVhFVMhB0ypso/J0p0tkRadmVJmRm6KOya3LSJH8xPtWhGIzWUECAECAFCgBAgBAgBQoAQIAQCRqDQKR9vrtu7EaFx1eqycRDw5ARP2MeEYyf37eRxcSKpXWxcqbrgNf3gFp7FQePentzfMvgXv3YspB/brS56MzaxPa8DrJGW+Y7ZHpnTJrfj4Y+snle8lTEvaEvLzHCVE7WXi1oG3266IiFACBAChAAhQAgQAoQAIUAIFDYChU/54pMUMNtXuFJbMiZW9oIjfyLC5K7IbC78J9nPVursfupHFcaFYzW6bqe5c1E/HoUnIgYtL1KXGot9e8bFLPsbJ6Gch1rxe8bZg/qWJeo7g2DZi41kFJRTO5FjnRv0LEYqPFGZuU/Y+kTK9Rj43uoZZizMr6szzZmc8NLCHolUO0KAECAECAFCgBAgBAgBQoAQyAEChU/54hstuJqphLX3pytjnudhckyxk3EtRreErc/1ZiovyIDXSpnQXikfpm1faZw5CNLIMzrgUsLux7L52ekQbPon/mAROzcNM01dM6suGKcOaFsWqeVDlYntYmWoCWNxwoczoQ7Nee54YY1EAR5zyKIQW0UndtA2LajRFKHLyUlnDvqZLkkIEAKEACFACBAChAAhQAgQAkWJQL2jfCz1gsXClKi25m1lTGum3cLIHk9rniCOYmXwsxK4g2s1B0NT331FXT5e3/6ecWo/kiiAwzGRF57XgTEunuiBW/x4KJ3wCDV0I3RJP7RZ2zhfXTpWndUX8qHcg9TKDcgSRYxqlVSQE+n4uAWyOQRmYiN5GZZTvoUy5SXt4EeGrlmhe7alsSiHIjWaECAECAFCgBAgBAgBQoAQIASCR6DeUT5mjXPIGEvdcPAj5a1OQsQlKpwq48LkRFzfSPCx56JMPFMkwRNKKsJC2EKZ2FGZ2QdheOriUeq62dq6WZAGZe/V09SFI9WKYTyL+vPcJ5N9l38RnE2QPfbZctQUSSOSWPkYKQXHi45EeglmioyOaa0uLjOqzqMFrFeF36jtMZqbgMPgRw9dkRAgBAgBQoAQIAQIAUKAECAEChyB+kf5nCR4lt8laN+JXfCujI22MzckUD7uVCkkPUU0HQRURIgdJ4rMOieMhPi1laMCysVXxDVRQPyr0N7kBI/TSCHHwn+1IvcYG0wQ5BTlQTiZJRDSnZM7a9tWmNFqmA6tnA5OYCBjfZwE0osQIAQIAUKAECAECAFCgBAgBAiBrBGof5TPTl9gRbwxgiTUMtfMVMa3ExY8i5IJYiZ0MnmOBJ4WT1jh2M+ope/CfuU8TRRmv4IT2rqa1gcUZuX5W/wT/7VWftMqYMnG2OKcNlfkWQQZ39NWTDLOH3eb9URwYm1ehqw7lS5ACBAChAAhQAgQAoQAIUAIEAKEgECgPlK+K/pOCGmyFyRVTu9X5w5WEd3HU/aJpO1WLB+3xSXXVkn0w8zuL9zbU1j/WHSfSBsYHdtKKx+qH/jI1GI0+AgBQoAQIAQIAUKAECAECAFCgBDIDwINgfI5Hp6M/kWrjB0rY5M7cklMRvmE4Y4TPytXe5JYu+w4XoI+p50YENGDTJyzRWx6d307PDlDzJxHORjyM7TpLoQAIUAIEAKEACFACBAChAAh0ACsfLwTLd0Tnt6Am/2UiIZEebP6itwJCsvSzkx83P6WkEEhWL5npQdkGi3KmBJ1/mv6vnXMAunIgNKwIwQIAUKAECAECAFCgBAgBAgBQiBfCNR7K58IhHOMZyJBusiwYFSd03a8p84ZoIwv5WItraCxyQRaAud4cYIxo19QJnXSlow2jmwzY9Ui1Z6VbY9y7uVrZNN9CAFCgBAgBAgBQoAQIAQIAUIACNR/yucy83Huxxgg/4+d98DQ9aM7tQ/eVqa8yDRXRKCdkzNdSLzYoXci3o+ncGBane43XDR5bj1brtPKzSBSq3PLIb4ytkRd+IaxdZlx+bSog5uPihTvNOwIAUKAECAECAFCgBAgBAgBQoAQyBsCDYTy+eNl6Ga40ji+R39/mvLOIFjhlNFCWEUkRmdhfjyJAvvMczY8J2RXxNvO1iByM+CLVi71aFmr2MT2Iqefvut98+KpGiVqW/T8K0UlCAFCgBAgBAgBQoAQIAQIAUKAEMgpAsVA+RIMa7GwcWq/vm+tvnaG+s4gdUI7kUEB7whLsN5C5VovwtCnjAQn5IF5Qv2FJXtoERvbRpkzQF0xwdjxnnF8t1l9scbQrTBC27iX026jixMChAAhQAgQAoQAIUAIEAKEACEgg8D/B8QvoQ1c7vLhAAAAAElFTkSuQmCC"


THIN_SIDE = Side(style="thin", color=COLORS["line"])
MEDIUM_SIDE = Side(style="medium", color=COLORS["black"])
THICK_SIDE = Side(style="thick", color=COLORS["black"])
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _normalize_label(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _to_number(value: object) -> int | float:
    if isinstance(value, (int, float)):
        return value
    digits = re.sub(r"[^0-9.-]", "", str(value or ""))
    if not digits or digits in {"-", ".", "-."}:
        return 0
    number = float(digits)
    return int(number) if number.is_integer() else number


def _is_fully_paid(payment_count: object) -> bool:
    """납입횟수가 1/1회, 180/180회처럼 모두 채워진 계약인지 확인합니다."""
    match = re.fullmatch(
        r"\s*([0-9,]+)\s*/\s*([0-9,]+)\s*(?:회)?\s*",
        str(payment_count or ""),
    )
    if not match:
        return False
    paid_count = int(match.group(1).replace(",", ""))
    total_count = int(match.group(2).replace(",", ""))
    return total_count > 0 and paid_count == total_count


def _format_won_text(value: object) -> str:
    number = _to_number(value)
    if isinstance(number, float) and not number.is_integer():
        return f"{number:,.2f}원"
    return f"{int(number):,}원"


def _group_for(label: str) -> str:
    normalized = _normalize_label(label)
    for group_name, members in GROUP_RULES:
        if normalized in {_normalize_label(item) for item in members}:
            return group_name
    return "기타"


def _extract_customer_name(value: object) -> str:
    text = str(value or "OOO").strip()
    name = re.split(r"[을를]\s*위한", text, maxsplit=1)[0].strip()
    return name or "OOO"


def _extract_age(value: object) -> str:
    match = re.search(r"(\d+)\s*세", str(value or ""))
    return match.group(1) if match else ""


def parse_source_file(main_bytes: bytes) -> dict:
    workbook = openpyxl.load_workbook(BytesIO(main_bytes), data_only=True)
    required = ["계약사항", "상품별보장내용"]
    missing = [name for name in required if name not in workbook.sheetnames]
    if missing:
        raise ValueError("필수 시트 없음:" + ",".join(missing))

    contracts_ws = workbook["계약사항"]
    coverage_ws = workbook["상품별보장내용"]
    customer_name = _extract_customer_name(contracts_ws["B2"].value)
    age = _extract_age(contracts_ws["D2"].value)

    contract_columns = []
    for col in range(6, coverage_ws.max_column + 1):
        if coverage_ws.cell(2, col).value or coverage_ws.cell(3, col).value:
            contract_columns.append(col)

    if not contract_columns:
        raise ValueError("원본 파일에서 보험계약 정보를 찾을 수 없습니다.")

    contracts = []
    for index, col in enumerate(contract_columns):
        contract_row = 9 + index
        contracts.append(
            {
                "company": coverage_ws.cell(2, col).value or "",
                "product": coverage_ws.cell(3, col).value or "",
                "coverage_period": coverage_ws.cell(4, col).value or "",
                "payment_count": coverage_ws.cell(5, col).value or "",
                "payment_cycle": coverage_ws.cell(6, col).value or "",
                "monthly": _to_number(coverage_ws.cell(7, col).value),
                "total": _to_number(contracts_ws.cell(contract_row, 10).value),
                "paid": _to_number(contracts_ws.cell(contract_row, 11).value),
                "remaining": _to_number(contracts_ws.cell(contract_row, 12).value),
            }
        )

    coverages = []
    started = False
    for row in range(9, coverage_ws.max_row + 1):
        raw_label = coverage_ws.cell(row, 2).value
        if raw_label in (None, ""):
            if started:
                break
            continue
        started = True
        label = _normalize_label(raw_label)
        values = [_to_number(coverage_ws.cell(row, col).value) for col in contract_columns]
        coverages.append(
            {
                "label": label,
                "display": DISPLAY_NAMES.get(label, str(raw_label).strip()),
                "group": _group_for(label),
                "values": values,
            }
        )

    if not coverages:
        raise ValueError("원본 파일에서 보장항목을 찾을 수 없습니다.")

    return {
        "customer_name": customer_name,
        "age": age,
        "contracts": contracts,
        "coverages": coverages,
    }


def _set_outline(ws, min_row: int, max_row: int, min_col: int, max_col: int, side: Side) -> None:
    def replace_side(cell, **changes) -> None:
        border = cell.border
        cell.border = Border(
            left=changes.get("left", copy(border.left)),
            right=changes.get("right", copy(border.right)),
            top=changes.get("top", copy(border.top)),
            bottom=changes.get("bottom", copy(border.bottom)),
            diagonal=copy(border.diagonal),
            diagonal_direction=border.diagonal_direction,
            diagonalUp=border.diagonalUp,
            diagonalDown=border.diagonalDown,
            outline=border.outline,
            vertical=copy(border.vertical),
            horizontal=copy(border.horizontal),
        )

    for col in range(min_col, max_col + 1):
        top_cell = ws.cell(min_row, col)
        bottom_cell = ws.cell(max_row, col)
        replace_side(top_cell, top=side)
        replace_side(bottom_cell, bottom=side)
    for row in range(min_row, max_row + 1):
        left_cell = ws.cell(row, min_col)
        right_cell = ws.cell(row, max_col)
        replace_side(left_cell, left=side)
        replace_side(right_cell, right=side)


def _set_vertical_borders(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    side: Side,
) -> None:
    """표 안의 모든 열 경계만 지정한 굵기로 통일합니다."""

    def replace_side(cell, **changes) -> None:
        border = cell.border
        cell.border = Border(
            left=changes.get("left", copy(border.left)),
            right=changes.get("right", copy(border.right)),
            top=copy(border.top),
            bottom=copy(border.bottom),
            diagonal=copy(border.diagonal),
            diagonal_direction=border.diagonal_direction,
            diagonalUp=border.diagonalUp,
            diagonalDown=border.diagonalDown,
            outline=border.outline,
            vertical=copy(border.vertical),
            horizontal=copy(border.horizontal),
        )

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col):
            replace_side(ws.cell(row, col), right=side)
            replace_side(ws.cell(row, col + 1), left=side)


def _extract_logo() -> bytes:
    """외부 파일 없이 코드에 내장된 Hanwha Life Lab 로고를 반환합니다."""
    return base64.b64decode(LOGO_BASE64)


def _configure_print(
    ws,
    contract_count: int,
    coverage_count: int,
    last_row: int,
    last_col: int,
    page_count: int = 1,
) -> None:
    # 다운로드 직후 바로 인쇄할 수 있도록 A3 세로형에서
    # 너비와 높이를 모두 1페이지에 맞춥니다.
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = "portrait"
    ws.page_setup.pageOrder = "overThenDown"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.page_setup.scale = None
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    # A~C열은 반복 인쇄하지 않습니다.
    ws.print_title_cols = None
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.oddFooter.center.text = "페이지 &P / &N"
    ws.oddFooter.center.size = 9
    # 좌우·위쪽은 인쇄 공간을 넓게 쓰고, 아래쪽은 표와 페이지 번호가
    # 겹치지 않도록 여유를 둡니다. 바닥글은 일반 프린터의 비인쇄 영역을
    # 고려해 용지 아래에서 0.20인치 위치에 배치합니다.
    ws.page_margins.left = 0.12
    ws.page_margins.right = 0.12
    ws.page_margins.top = 0.12
    ws.page_margins.bottom = 0.38
    ws.page_margins.header = 0
    ws.page_margins.footer = 0.20
    ws.sheet_view.zoomScale = 100


def _contract_column_width(contract_count: int) -> float:
    """보험 수가 적당할 때 A3 가로폭을 넉넉히 쓰되 지나치게 넓어지지 않게 합니다."""
    widths = {1: 25.0, 2: 25.0, 3: 27.0, 4: 24.0, 5: 20.5, 6: 18.0}
    return widths.get(contract_count, max(14.0, 108.0 / max(contract_count, 1)))


def _populate_analysis_sheet(
    workbook: Workbook,
    title: str,
    data: dict,
    selected: list[dict],
    contract_indices: list[int],
    contracts_per_page: int | None = None,
    page_count: int = 1,
) -> None:
    ws = workbook.create_sheet(title)
    ws.sheet_view.showGridLines = False
    contracts = [data["contracts"][index] for index in contract_indices]
    contract_count = len(contracts)
    last_col = 3 + contract_count
    coverage_start = 11

    output_items = [
        {"label": "일반사망", "display": "일반 사망", "group": "사망", "values": [0] * contract_count},
        *[
            {
                **item,
                "values": [item["values"][index] for index in contract_indices],
            }
            for item in selected
        ],
        {"label": "기타", "display": "기타", "group": "기타", "values": [0] * contract_count},
    ]
    coverage_end = coverage_start + len(output_items) - 1

    normal_font = Font(name="나눔고딕", size=10, color=COLORS["black"])
    bold_font = Font(name="나눔고딕", size=10, bold=True, color=COLORS["black"])
    blue_font = Font(name="나눔고딕", size=11, bold=True, color=COLORS["blue"])
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:C1")
    age_text = f" (보험연령:{data['age']}세)" if data["age"] else ""
    title_customer_name = re.sub(r"님$", "", str(data["customer_name"] or "OOO").strip()) or "OOO"
    ws["A1"] = f"{title_customer_name}님의 보장 분석{age_text}"
    ws["A1"].font = Font(name="나눔고딕", size=13, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="bottom")
    ws.row_dimensions[1].height = 82
    for col in range(1, last_col + 1):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=COLORS["white"])
        ws.cell(1, col).border = Border()
        ws.cell(1, col).alignment = center

    logo = XLImage(BytesIO(_extract_logo()))
    logo.width = 350
    logo.height = 43
    ws.add_image(logo, "A1")

    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    ws["A2"] = "합 계"
    ws["B2"] = "구분"
    ws["C2"] = "보장명"
    ws["A2"].font = Font(name="나눔고딕", size=11, bold=True, color=COLORS["red"])

    for row in range(2, 4):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=COLORS["header"])
            cell.border = THIN_BORDER
            cell.alignment = center
            if cell.coordinate != "A2":
                cell.font = bold_font

    for index, contract in enumerate(contracts, start=4):
        ws.cell(2, index, contract["company"])
        ws.cell(3, index, contract["product"])
        ws.cell(2, index).font = Font(name="나눔고딕", size=10, bold=True, color="1F4E78")
        ws.cell(3, index).font = Font(name="나눔고딕", size=9, bold=True)
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 55

    meta_rows = [
        (4, "보장기간", "coverage_period"),
        (5, "납입횟수", "payment_count"),
        (6, "납입주기", "payment_cycle"),
        (7, "월보험료", "monthly"),
        (8, "납입완료", "paid"),
        (9, "납입예정", "remaining"),
        (10, "총보험료", "total"),
    ]
    for row, label, key in meta_rows:
        if row <= 6:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row, 1, label)
        else:
            last_contract_col = get_column_letter(last_col)
            ws.cell(row, 1, f"=SUM(D{row}:{last_contract_col}{row})")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row, 2, label)

        fill_color = COLORS["header"] if row == 7 or row <= 6 else COLORS["premium"]
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=fill_color if col == 1 or col >= 4 else COLORS["header"])
            cell.border = THIN_BORDER
            cell.alignment = center
            cell.font = bold_font

        for index, contract in enumerate(contracts, start=4):
            cell = ws.cell(row, index)
            if row == 7 and _to_number(contract["monthly"]) == 0:
                cell.value = "확인 필요"
                cell.font = normal_font
            elif row == 7 and _is_fully_paid(contract["payment_count"]):
                cell.value = _format_won_text(contract["monthly"])
                cell.font = blue_font
            else:
                cell.value = contract[key]
            if row >= 7:
                cell.number_format = '#,##0"원"'
                if not (row == 7 and _to_number(contract["monthly"]) == 0):
                    cell.font = blue_font
        if row >= 7:
            ws.cell(row, 1).number_format = '#,##0"원"'
            ws.cell(row, 1).font = blue_font

    # 완납 계약은 보험회사명부터 총보험료까지 해당 보험 열 전체를 녹색으로 표시합니다.
    for index, contract in enumerate(contracts, start=4):
        if _is_fully_paid(contract["payment_count"]):
            completed_fill = PatternFill("solid", fgColor=COLORS["completed"])
            for row in range(2, 11):
                ws.cell(row, index).fill = completed_fill

    group_ranges: list[tuple[int, int]] = []
    group_start = coverage_start
    current_group = output_items[0]["group"]
    for offset, item in enumerate(output_items):
        row = coverage_start + offset
        group = item["group"]
        if group != current_group:
            group_ranges.append((group_start, row - 1))
            group_start = row
            current_group = group

        section_color = COLORS["white"]
        if group == "암\n보장":
            section_color = COLORS["cancer"]
        elif group == "뇌\n보장":
            section_color = COLORS["brain"]
        elif group == "심장\n보장":
            section_color = COLORS["heart"]

        last_contract_col = get_column_letter(last_col)
        ws.cell(row, 1, f"=SUM(D{row}:{last_contract_col}{row})")
        ws.cell(row, 2, group)
        ws.cell(row, 3, item["display"])
        for index, value in enumerate(item["values"], start=4):
            ws.cell(row, index, value)

        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=COLORS["header"] if col == 2 else section_color)
            cell.border = THIN_BORDER
            cell.alignment = center
            cell.font = blue_font if col == 1 else bold_font
            if col == 1 or col >= 4:
                cell.number_format = '#,##0"만원";[Red]-#,##0"만원";;'
        ws.row_dimensions[row].height = 25
    group_ranges.append((group_start, coverage_end))

    for start, end in group_ranges:
        if end > start:
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        ws.cell(start, 2).alignment = center
        _set_outline(ws, start, end, 1, last_col, MEDIUM_SIDE)

    _set_outline(ws, 2, 3, 1, last_col, MEDIUM_SIDE)
    _set_outline(ws, 4, 6, 1, last_col, MEDIUM_SIDE)
    _set_outline(ws, 7, 10, 1, last_col, MEDIUM_SIDE)

    # C열(보장명)과 D열(첫 보험계약) 사이를 굵게 구분합니다.
    # 같은 보험사의 연속된 계약은 한 묶음으로 두고, 보험사가 바뀌는
    # 지점에만 굵은 세로 경계선을 표시합니다.
    company_group_start = 4
    for col in range(5, last_col + 1):
        previous_company = str(ws.cell(2, col - 1).value or "").strip()
        current_company = str(ws.cell(2, col).value or "").strip()
        if current_company != previous_company:
            _set_outline(
                ws,
                2,
                coverage_end,
                company_group_start,
                col - 1,
                MEDIUM_SIDE,
            )
            company_group_start = col
    _set_outline(
        ws,
        2,
        coverage_end,
        company_group_start,
        last_col,
        MEDIUM_SIDE,
    )

    # A열부터 마지막 보험계약 열까지 모든 내부 세로선을 굵게 표시합니다.
    # 가로선은 기존 굵기를 그대로 유지합니다.
    _set_vertical_borders(ws, 2, coverage_end, 1, last_col, MEDIUM_SIDE)

    # 모든 내부 경계선을 적용한 뒤 표 전체 외곽선을 마지막에 다시
    # 설정해 2행부터 시작하는 굵은 테두리가 중간에 끊기지 않게 합니다.
    _set_outline(ws, 2, coverage_end, 1, last_col, THICK_SIDE)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 31
    contract_width = _contract_column_width(contract_count)
    for col in range(4, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = contract_width

    _configure_print(ws, contract_count, len(selected), coverage_end, last_col, page_count)


def build_analysis_file(
    main_bytes: bytes,
    selected_labels: list[str] | None = None,
) -> tuple[bytes, str, str]:
    data = parse_source_file(main_bytes)
    available = {item["label"]: item for item in data["coverages"]}

    if selected_labels is None:
        selected_labels = DEFAULT_COVERAGES
    normalized_selection = {_normalize_label(label) for label in selected_labels}
    selected = [item for item in data["coverages"] if item["label"] in normalized_selection]
    if not selected:
        raise ValueError("출력할 보장항목을 한 개 이상 선택해 주세요.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    all_contract_indices = list(range(len(data["contracts"])))
    _populate_analysis_sheet(
        workbook,
        "보장 분석",
        data,
        selected,
        all_contract_indices,
        contracts_per_page=len(data["contracts"]),
        page_count=1,
    )
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    today = datetime.today().strftime("%Y%m%d")
    filename_name = re.sub(r"님$", "", str(data["customer_name"] or "OOO").strip())
    filename_name = re.sub(r'[\\/:*?"<>|]+', "_", filename_name).strip() or "OOO"
    filename = f"{filename_name}님_보장분석엑셀_{today}.xlsx"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue(), filename, data["customer_name"]


def make_input_signature(main_bytes: bytes, mode: str, selected_labels: list[str]) -> str:
    digest = hashlib.sha256()
    digest.update(main_bytes)
    digest.update(mode.encode("utf-8"))
    digest.update("|".join(selected_labels).encode("utf-8"))
    return digest.hexdigest()


def _render_personal_selector(
    available_labels: list[str],
    default_labels: list[str],
    file_key: str,
) -> list[str]:
    """모든 보장항목을 구분별 카드에 펼쳐 보여주는 개인모드 선택기입니다."""
    prefix = f"analyzer_v2_cov_{file_key}_"
    default_set = set(default_labels)

    for label in available_labels:
        state_key = f"{prefix}{label}"
        if state_key not in st.session_state:
            st.session_state[state_key] = label in default_set

    col_all, col_default, col_clear = st.columns(3)
    if col_all.button("전체 선택", use_container_width=True, key=f"{prefix}all"):
        for label in available_labels:
            st.session_state[f"{prefix}{label}"] = True
    if col_default.button("간편모드 기본값", use_container_width=True, key=f"{prefix}default"):
        for label in available_labels:
            st.session_state[f"{prefix}{label}"] = label in default_set
    if col_clear.button("전체 해제", use_container_width=True, key=f"{prefix}clear"):
        for label in available_labels:
            st.session_state[f"{prefix}{label}"] = False

    grouped: dict[str, list[str]] = {}
    for label in available_labels:
        group = _group_for(label).replace("\n", " ")
        grouped.setdefault(group, []).append(label)

    card_columns = st.columns(3)
    for group_index, (group, labels) in enumerate(grouped.items()):
        selected_count = sum(bool(st.session_state[f"{prefix}{label}"]) for label in labels)
        with card_columns[group_index % 3]:
            with st.container(border=True):
                st.markdown(f"**{group}** · {selected_count}/{len(labels)}")
                for label in labels:
                    st.checkbox(
                        DISPLAY_NAMES.get(label, label),
                        key=f"{prefix}{label}",
                    )

    selected = [label for label in available_labels if st.session_state[f"{prefix}{label}"]]
    st.caption(f"전체 {len(available_labels)}개 중 {len(selected)}개 선택")
    return selected


def run() -> None:
    page_header(
        "고객 상담",
        "보장 분석 도우미",
        "전체 보장분석 원본을 고객 상담용 엑셀로 자동 정리합니다.",
        "▤",
    )

    with st.expander("사용 방법 안내"):
        st.markdown(
            """
            1. 전체 보장내용이 포함된 **컨설팅보장분석.xlsx** 파일을 업로드합니다.
            2. **간편모드**는 추천 기본 보장을 즉시 적용합니다.
            3. **개인모드**는 전체 보장 중 원하는 항목을 직접 선택합니다.
            4. 간편모드는 업로드 즉시 결과가 생성되며, 개인모드는 항목 선택 후 시작 버튼을 누릅니다.

            - 결과물은 하나의 시트에서 A3 용지에 맞춰집니다.
            - 결과 엑셀은 A3 세로형에서 **너비 1페이지·높이 자동 맞춤**을 기본값으로 사용합니다.
            - 필요하면 가로·세로 방향, 출력 배율과 페이지 나누기를 엑셀 인쇄 화면에서 직접 조정할 수 있습니다.
            - 페이지 하단에는 현재 페이지와 전체 페이지 번호가 표시됩니다.
            """
        )
        st.caption("제작 박병선 팀장 최종 · 버전 v2.13.0")

    st.markdown("### ✦ 전체 보장분석 원본")
    uploaded_main = st.file_uploader(
        "전체 보장내용이 포함된 컨설팅보장분석.xlsx 파일을 업로드하세요",
        type=["xlsx"],
        key="analyzer_v2_main_file",
    )

    parsed = None
    parse_error = None
    main_bytes = uploaded_main.getvalue() if uploaded_main else b""
    if uploaded_main:
        try:
            parsed = parse_source_file(main_bytes)
        except Exception as exc:
            parse_error = exc
            st.error(str(exc))

    st.markdown("### ✦ 분석 방식 선택")
    mode = st.radio(
        "분석 방식을 선택하세요",
        ["간편모드", "개인모드"],
        horizontal=True,
        key="analyzer_v2_mode",
    )

    selected_labels: list[str] = []
    if parsed:
        available_labels = [item["label"] for item in parsed["coverages"]]
        default_labels = [label for label in available_labels if label in set(DEFAULT_COVERAGES)]

        if mode == "간편모드":
            selected_labels = default_labels
            st.info(f"기본 보장 {len(selected_labels)}개가 자동으로 적용됩니다.")
            with st.expander("간편모드 적용 항목 보기"):
                st.write([DISPLAY_NAMES.get(label, label) for label in selected_labels])
        else:
            st.markdown("#### 출력할 보장항목")
            st.caption("구분별 카드에서 필요한 보장항목을 바로 체크하거나 해제하세요.")
            selected_labels = _render_personal_selector(
                available_labels,
                default_labels,
                hashlib.sha256(main_bytes).hexdigest()[:10],
            )
    elif not uploaded_main:
        st.caption("원본 파일을 업로드하면 선택 가능한 전체 보장항목이 표시됩니다.")

    ready = bool(parsed) and bool(selected_labels) and parse_error is None
    signature = make_input_signature(main_bytes, mode, selected_labels) if ready else None

    should_generate = False
    if mode == "간편모드" and ready:
        current_result = st.session_state.get("analyzer_v2_result")
        current_error = st.session_state.get("analyzer_v2_error")
        should_generate = not (
            (current_result and current_result.get("signature") == signature)
            or (current_error and current_error.get("signature") == signature)
        )
    else:
        st.markdown("### ✦ 보장 분석 실행")
        should_generate = st.button(
            "보장 분석 시작",
            type="primary",
            disabled=not ready,
            use_container_width=True,
            key="analyzer_v2_run",
        )

    if should_generate:
        st.session_state.pop("analyzer_v2_result", None)
        st.session_state.pop("analyzer_v2_error", None)
        try:
            with st.spinner("고객 상담용 보장분석 엑셀을 만들고 있습니다..."):
                result_bytes, filename, customer_name = build_analysis_file(
                    main_bytes,
                    selected_labels,
                )
            st.session_state["analyzer_v2_result"] = {
                "signature": signature,
                "bytes": result_bytes,
                "filename": filename,
                "customer_name": customer_name,
                "mode": mode,
                "coverage_count": len(selected_labels),
            }
        except Exception as exc:
            st.session_state["analyzer_v2_error"] = {
                "signature": signature,
                "message": str(exc),
                "detail": repr(exc),
            }

    error = st.session_state.get("analyzer_v2_error")
    if error and error.get("signature") == signature:
        st.error(error["message"])
        with st.expander("오류 상세 보기"):
            st.code(error["detail"])

    result = st.session_state.get("analyzer_v2_result")
    if result and result.get("signature") == signature:
        st.divider()
        st.markdown("### ✦ 분석 결과 및 다운로드")
        st.success("보장 분석이 완료되었습니다.")
        col1, col2, col3 = st.columns(3)
        col1.metric("고객명", result["customer_name"])
        col2.metric("분석 모드", result["mode"])
        col3.metric("보장항목", f"{result['coverage_count']}개")
        st.download_button(
            "결과 엑셀 다운로드",
            data=result["bytes"],
            file_name=result["filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            key="analyzer_v2_download",
        )


if __name__ == "__main__":
    run()
