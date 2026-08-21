import streamlit as st
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Tuple
from .ui_components import page_header

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # 엑셀 다운로드 기능은 openpyxl 설치 후 활성화됩니다.
    Workbook = None

EOK = 10_000  # 내부 계산 단위: 만원
UI_VERSION = "2026-07-consulting-report-v5-spouse-funeral"


# -----------------------------------------------------------------------------
# 금액 표시·입력 보조 함수
# -----------------------------------------------------------------------------
def won_text(amount_manwon: float) -> str:
    won = max(0, round(float(amount_manwon) * 10_000))
    if won == 0:
        return "0원"
    eok, remainder = divmod(won, 100_000_000)
    man, one = divmod(remainder, 10_000)
    parts = []
    if eok:
        parts.append(f"{eok:,}억")
    if man:
        parts.append(f"{man:,}만")
    if one:
        parts.append(f"{one:,}")
    return " ".join(parts) + "원"


def parse_manwon(value: str) -> tuple[float, str | None]:
    """사용자 입력 문자열을 만원 단위 숫자로 변환한다."""
    raw = str(value or "").strip()
    if not raw:
        return 0.0, None

    cleaned = raw.replace("만원", "").replace(",", "").replace(" ", "")
    if not cleaned:
        return 0.0, None

    try:
        amount = float(cleaned)
    except ValueError:
        return 0.0, "숫자, 쉼표, 공백, '만원'만 입력할 수 있습니다."

    if amount < 0:
        return 0.0, "금액은 0 이상으로 입력해 주세요."

    return amount, None


def format_manwon_input(amount: float) -> str:
    amount = max(0.0, float(amount))
    if amount.is_integer():
        return f"{int(amount):,}"
    return f"{amount:,.2f}".rstrip("0").rstrip(".")


def amount_input(
    label: str,
    *,
    value: float = 0,
    key: str,
    help_text: str | None = None,
    show_default_notice: bool = False,
) -> float:
    """만원 단위 금액 입력: 우측 단위, 쉼표 정리, 원화 해석을 함께 표시한다."""
    display_key = f"{key}_display"
    error_key = f"{key}_error"

    if display_key not in st.session_state:
        st.session_state[display_key] = format_manwon_input(float(value))
    if error_key not in st.session_state:
        st.session_state[error_key] = None

    def normalize() -> None:
        parsed, error = parse_manwon(st.session_state.get(display_key, ""))
        st.session_state[error_key] = error
        if error is None:
            st.session_state[display_key] = format_manwon_input(parsed)

    input_col, unit_col = st.columns([10, 1.35], vertical_alignment="bottom")
    with input_col:
        st.text_input(label, key=display_key, on_change=normalize, help=help_text)
    with unit_col:
        st.markdown(
            "<div style='padding:0 0 0.55rem 0;font-weight:600;'>만원</div>",
            unsafe_allow_html=True,
        )

    amount, current_error = parse_manwon(st.session_state.get(display_key, ""))
    error = st.session_state.get(error_key) or current_error
    if error:
        st.error(error, icon="⚠️")
        return 0.0

    st.caption(f"입력 금액 해석: **{won_text(amount)}**")
    if show_default_notice:
        st.caption("기본값이 입력되어 있습니다. 실제 금액을 확인해 수정하세요.")
    return amount


# -----------------------------------------------------------------------------
# 기존 상속세 계산 로직 — 변경하지 않음
# -----------------------------------------------------------------------------
def tax_rate_and_deduction(tax_base: float) -> Tuple[float, float]:
    if tax_base <= 1 * EOK:
        return 0.10, 0
    if tax_base <= 5 * EOK:
        return 0.20, 1_000
    if tax_base <= 10 * EOK:
        return 0.30, 6_000
    if tax_base <= 30 * EOK:
        return 0.40, 16_000
    return 0.50, 46_000


def financial_asset_deduction(net_financial_assets: float) -> float:
    x = max(0.0, net_financial_assets)
    if x <= 2_000:
        return x
    if x <= 10_000:
        return 2_000
    if x <= 100_000:
        return x * 0.20
    return 20_000


def spouse_statutory_share(group: str, count: int) -> float:
    if group == "배우자 단독":
        return 1.0
    count = max(1, int(count))
    return 1.5 / (count + 1.5)


@dataclass
class Result:
    gross_estate: float
    taxable_estate: float
    personal_or_lump: float
    spouse_deduction: float
    financial_deduction: float
    home_deduction: float
    other_deduction: float
    deduction_before_limit: float
    deduction_limit: float
    allowed_deduction: float
    tax_base: float
    rate: float
    progressive_deduction: float
    calculated_tax: float
    generation_skip_surcharge: float
    tax_credits: float
    filing_credit: float
    estimated_tax_due: float


def calculate(**v) -> Result:
    gross = max(0, v["gross_estate"]) + max(0, v["deemed_estate"])
    expenses = max(0, v["public_dues"]) + max(0, v["funeral_expense"]) + max(0, v["liabilities"])
    prior_gifts = max(0, v["prior_gifts_heirs"]) + max(0, v["prior_gifts_non_heirs"])
    taxable_estate = max(0, gross - max(0, v["non_taxable"]) - expenses + prior_gifts)

    personal = (
        20_000
        + max(0, v["children_count"]) * 5_000
        + max(0, v["minor_deduction"])
        + max(0, v["elderly_count"]) * 5_000
        + max(0, v["disability_deduction"])
    )
    personal_or_lump = max(50_000, personal) if v["lump_mode"] else personal
    if v["spouse_exists"] and v["spouse_solo"]:
        personal_or_lump = personal

    spouse_deduction = 0.0
    if v["spouse_exists"]:
        if v["spouse_actual_inheritance"] < 50_000:
            spouse_deduction = 50_000
        else:
            spouse_limit_base = max(
                0,
                gross + max(0, v["prior_gifts_heirs"])
                - max(0, v["non_heir_bequest"])
                - max(0, v["non_taxable"])
                - max(0, v["public_dues"])
                - max(0, v["liabilities"]),
            )
            spouse_limit = min(
                300_000,
                max(
                    0,
                    spouse_limit_base * min(1.0, max(0.0, v["spouse_share"]))
                    - max(0, v["spouse_prior_gift_tax_base"]),
                ),
            )
            spouse_deduction = min(max(0, v["spouse_actual_inheritance"]), spouse_limit)

    financial = financial_asset_deduction(v["net_financial_assets"])
    home = min(max(0, v["cohabiting_home_value"]), 60_000)
    deduction_before_limit = personal_or_lump + spouse_deduction + financial + home + max(0, v["other_deduction"])
    deduction_limit = max(
        0,
        taxable_estate
        - max(0, v["non_heir_bequest"])
        - max(0, v["inheritance_waiver_next_rank"])
        - max(0, v["prior_gift_tax_base_for_limit"]),
    )
    allowed = min(deduction_before_limit, deduction_limit)
    tax_base = max(0, taxable_estate - allowed - max(0, v["appraisal_fee"]))
    rate, progressive = tax_rate_and_deduction(tax_base)
    calculated_tax = max(0, tax_base * rate - progressive)

    gen_amount = min(max(0, v["generation_skip_amount"]), taxable_estate)
    gen_ratio = gen_amount / taxable_estate if taxable_estate else 0
    gen_rate = 0.40 if v["generation_skip_minor_over_2b"] else 0.30
    gen_surcharge = calculated_tax * gen_ratio * gen_rate

    before_credit = calculated_tax + gen_surcharge
    tax_credits = min(before_credit, max(0, v["gift_tax_credit"]) + max(0, v["other_tax_credit"]))
    after_credit = max(0, before_credit - tax_credits)
    filing_credit = after_credit * 0.03 if v["apply_filing_credit"] else 0

    return Result(
        gross, taxable_estate, personal_or_lump, spouse_deduction, financial, home,
        max(0, v["other_deduction"]), deduction_before_limit, deduction_limit, allowed,
        tax_base, rate, progressive, calculated_tax, gen_surcharge, tax_credits,
        filing_credit, max(0, after_credit - filing_credit)
    )


# -----------------------------------------------------------------------------
# UI 상태·예시 관리
# -----------------------------------------------------------------------------
AMOUNT_DEFAULTS: Dict[str, float] = {
    "it_gross": 100_000,
    "it_deemed": 0,
    "it_nontax": 0,
    "it_dues": 0,
    "it_funeral_actual": 0,
    "it_burial_actual": 0,
    "it_liab": 0,
    "it_gift_h": 0,
    "it_gift_n": 0,
    "it_minor": 0,
    "it_disability": 0,
    "it_spouse_amount": 50_000,
    "it_spouse_prior_base": 0,
    "it_fin": 0,
    "it_home": 0,
    "it_other_ded": 0,
    "it_appraisal": 0,
    "it_bequest": 0,
    "it_waiver": 0,
    "it_prior_base": 0,
    "it_gen": 0,
    "it_gift_credit": 0,
    "it_other_credit": 0,
    "it_cash": 0,
    "it_death_benefit": 0,
    "it_other_liquidity": 0,
}

WIDGET_DEFAULTS: Dict[str, Any] = {
    "it_mode": "일괄공제와 인적공제 중 큰 금액",
    "it_children": 1,
    "it_elderly": 0,
    "it_spouse_exists": True,
    "it_group": "직계비속",
    "it_count": 1,
    "it_gen_minor": False,
    "it_filing": True,
}

EXAMPLES: Dict[str, Dict[str, Any]] = {
    "기본형 · 10억원": {
        "description": "배우자 없이 10억원을 상속받는 기본 사례입니다.",
        "amounts": {"it_gross": 100_000, "it_cash": 3_000},
        "widgets": {"it_spouse_exists": False, "it_children": 1},
    },
    "배우자 공동상속 · 20억원": {
        "description": "배우자와 자녀 2명이 공동상속하고 금융재산과 납부재원이 있는 사례입니다.",
        "amounts": {
            "it_gross": 200_000,
            "it_spouse_amount": 80_000,
            "it_fin": 30_000,
            "it_cash": 10_000,
            "it_death_benefit": 10_000,
        },
        "widgets": {
            "it_spouse_exists": True,
            "it_children": 2,
            "it_group": "직계비속",
            "it_count": 2,
        },
    },
    "부동산 중심 · 30억원": {
        "description": "재산은 크지만 즉시 사용할 현금이 부족한 사례입니다.",
        "amounts": {
            "it_gross": 300_000,
            "it_fin": 10_000,
            "it_cash": 5_000,
            "it_death_benefit": 0,
        },
        "widgets": {"it_spouse_exists": False, "it_children": 1},
    },
    "보험금 준비형 · 30억원": {
        "description": "부동산 중심 사례와 같은 조건에서 사망보험금 5억원을 준비한 사례입니다.",
        "amounts": {
            "it_gross": 300_000,
            "it_fin": 10_000,
            "it_cash": 5_000,
            "it_death_benefit": 50_000,
        },
        "widgets": {"it_spouse_exists": False, "it_children": 1},
    },
}


def _set_amount_state(key: str, amount: float) -> None:
    st.session_state[f"{key}_display"] = format_manwon_input(amount)
    st.session_state[f"{key}_error"] = None


def reset_all_inputs() -> None:
    """모든 계산·납부재원 입력값을 최초 기본값으로 되돌린다."""
    for key, amount in AMOUNT_DEFAULTS.items():
        _set_amount_state(key, amount)
    for key, value in WIDGET_DEFAULTS.items():
        st.session_state[key] = value
    st.session_state["it_active_example"] = ""
    st.session_state["it_flash"] = "입력값을 최초 상태로 초기화했습니다."


def load_selected_example() -> None:
    """기존 값을 모두 초기화한 뒤 선택한 예시를 적용한다."""
    selected = st.session_state.get("it_example_select", "")
    if selected not in EXAMPLES:
        st.session_state["it_flash"] = "먼저 불러올 예시를 선택해 주세요."
        return

    for key, amount in AMOUNT_DEFAULTS.items():
        _set_amount_state(key, amount)
    for key, value in WIDGET_DEFAULTS.items():
        st.session_state[key] = value

    example = EXAMPLES[selected]
    for key, amount in example["amounts"].items():
        _set_amount_state(key, amount)
    for key, value in example["widgets"].items():
        st.session_state[key] = value

    st.session_state["it_active_example"] = selected
    st.session_state["it_flash"] = f"‘{selected}’ 예시를 불러왔습니다."


def render_sidebar() -> None:
    # 함수명은 기존 호출과의 호환을 위해 유지하고, 표시는 본문으로 이동합니다.
    st.caption("제작 박병선 팀장 · 버전 v2.1.3")

    with st.expander("계산 흐름"):
        st.markdown(
            """
            **① 총상속재산** — 상속재산 + 추정·간주상속재산

            **② 상속세 과세가액** — 비과세 재산·공과금·장례비용·채무 차감 후 사전증여재산 합산

            **③ 상속공제** — 일괄·인적·배우자·금융재산·동거주택공제 등 반영

            **④ 과세표준** — 과세가액 − 실제 적용 공제액

            **⑤ 예상 납부세액** — 누진세율·세대생략 할증·세액공제 반영
            """
        )

    with st.expander("상속공제 계산 안내"):
        st.markdown(
            """
            - 일괄공제와 기초·인적공제 중 적용 가능한 유리한 금액을 반영합니다.
            - 배우자 단독상속 등 일부 경우에는 일괄공제 적용이 제한될 수 있습니다.
            - 입력한 공제 합계는 상속공제 종합한도에 따라 제한될 수 있습니다.
            """
        )

    with st.expander("배우자·금융재산공제 안내"):
        st.markdown(
            """
            - 배우자공제는 실제 상속금액과 법정상속지분, 공제한도에 따라 달라집니다.
            - 금융재산공제는 금융재산에서 금융채무를 뺀 순금융재산을 기준으로 계산합니다.
            - 순금융재산과 실제 상속세 납부에 쓸 수 있는 현금성 자산은 서로 다른 개념입니다.
            """
        )

    with st.expander("상속세율표"):
        st.markdown(
            """
            | 과세표준 | 세율 | 누진공제 |
            |---|---:|---:|
            | 1억원 이하 | 10% | 없음 |
            | 1억원 초과~5억원 이하 | 20% | 1,000만원 |
            | 5억원 초과~10억원 이하 | 30% | 6,000만원 |
            | 10억원 초과~30억원 이하 | 40% | 1억 6,000만원 |
            | 30억원 초과 | 50% | 4억 6,000만원 |
            """
        )

    with st.expander("결과 해석 및 유의사항"):
        st.markdown(
            """
            - 부동산·비상장주식·보험금 등은 평가방법과 계약관계에 따라 과세가액이 달라질 수 있습니다.
            - 사전증여재산의 합산 여부와 배우자공제 요건에 따라 실제 신고 결과가 달라질 수 있습니다.
            - 본 결과는 상담용 예상치이며 실제 신고 전 별도 검토가 필요합니다.
            """
        )

    st.info(
        "상속세가 예상된다면 세액뿐 아니라 상속 직후 바로 사용할 수 있는 현금성 납부재원도 함께 확인해야 합니다."
    )
    st.caption(f"UI 버전: {UI_VERSION}")


def render_quick_actions() -> None:
    st.markdown("### 빠른 시작")
    q1, q2, q3 = st.columns([2.2, 1, 1])
    with q1:
        selected = st.selectbox(
            "예시 선택",
            ["예시를 선택하세요"] + list(EXAMPLES.keys()),
            key="it_example_select",
            label_visibility="collapsed",
        )
    with q2:
        st.button("예시 불러오기", use_container_width=True, on_click=load_selected_example)
    with q3:
        st.button("전체 초기화", use_container_width=True, on_click=reset_all_inputs)

    if selected in EXAMPLES:
        st.caption(EXAMPLES[selected]["description"] + " 현재 입력값은 불러오기 버튼을 누를 때 변경됩니다.")

    flash = st.session_state.pop("it_flash", None)
    if flash:
        st.success(flash)

    active = st.session_state.get("it_active_example", "")
    if active:
        st.info(f"현재 적용된 예시: **{active}** · 실제 상담 시 고객 상황에 맞게 값을 수정하세요.")


def render_result_interpretation(tax_due: float, liquid_funds: float) -> None:
    if tax_due <= 0:
        st.info(
            "**현재 예상세액 없음**  \n현재 입력 기준으로 예상 납부세액이 발생하지 않습니다. 재산평가, 사전증여재산과 공제요건에 따라 실제 결과는 달라질 수 있습니다."
        )
        return

    gap = liquid_funds - tax_due
    if gap < 0:
        st.error(
            f"**납부재원 점검 필요**  \n현재 예상 상속세보다 즉시 사용할 수 있는 납부재원이 **{won_text(abs(gap))} 부족**합니다. "
            "부동산 등 비유동성 자산 비중이 높다면 자산 매각이나 대출이 필요할 수 있으며, 사망보험금 등 별도의 현금성 재원을 준비하면 급매 위험을 줄이는 데 도움이 될 수 있습니다."
        )
    else:
        st.success(
            f"**납부재원 확보**  \n현재 입력 기준으로 예상 상속세를 납부한 뒤 **{won_text(gap)}의 자금 여유**가 있습니다. "
            "장례비용, 생활비, 채무 정리 등 상속세 외 지출도 함께 고려하세요."
        )



# -----------------------------------------------------------------------------
# 결과 카드·엑셀 상담 보고서
# -----------------------------------------------------------------------------
NAVY = "17365D"
BLUE = "DCE6F1"
LIGHT_BLUE = "EEF4FA"
GREEN = "2E7D32"
LIGHT_GREEN = "EAF4EA"
RED = "B42318"
LIGHT_RED = "FDECEC"
ORANGE = "B54708"
LIGHT_ORANGE = "FFF4E5"
GRAY = "667085"
LIGHT_GRAY = "F2F4F7"
WHITE = "FFFFFF"
BORDER_GRAY = "D0D5DD"


def result_card(title: str, value: str, caption: str, tone: str = "blue") -> None:
    tones = {
        "blue": ("#EEF4FA", "#17365D", "#B8CCE4"),
        "green": ("#EAF4EA", "#2E7D32", "#A9D3AB"),
        "red": ("#FDECEC", "#B42318", "#F0B4AE"),
        "gray": ("#F2F4F7", "#344054", "#D0D5DD"),
    }
    bg, fg, border = tones.get(tone, tones["blue"])
    st.markdown(
        f"""
        <div style="background:{bg};border:1px solid {border};border-radius:14px;
                    padding:18px 18px 16px 18px;min-height:148px;">
          <div style="font-size:0.93rem;color:#667085;font-weight:700;">{title}</div>
          <div style="font-size:1.75rem;color:{fg};font-weight:800;line-height:1.25;margin-top:10px;">{value}</div>
          <div style="font-size:0.84rem;color:#667085;margin-top:10px;line-height:1.45;">{caption}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def flow_card(title: str, value: str, subtitle: str = "") -> None:
    st.markdown(
        f"""
        <div style="border:1px solid #D0D5DD;border-radius:12px;padding:13px 12px;background:#FFFFFF;text-align:center;min-height:112px;">
          <div style="font-size:0.82rem;color:#667085;font-weight:700;">{title}</div>
          <div style="font-size:1.22rem;color:#17365D;font-weight:800;margin-top:8px;">{value}</div>
          <div style="font-size:0.75rem;color:#98A2B3;margin-top:6px;">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _excel_title(ws, title: str, subtitle: str, end_col: int = 8) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=end_col)
    cell = ws.cell(2, 2, title)
    cell.font = Font(name="맑은 고딕", size=21, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=end_col)
    sub = ws.cell(4, 2, subtitle)
    sub.font = Font(name="맑은 고딕", size=9, color=GRAY)
    sub.alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 27
    ws.row_dimensions[3].height = 10


def _style_range(ws, cell_range: str, *, fill=None, font=None, border=None, alignment=None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if border is not None:
                cell.border = border
            if alignment is not None:
                cell.alignment = alignment


def _write_section_title(ws, row: int, title: str, start_col: int = 2, end_col: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row, start_col, title)
    c.font = Font(name="맑은 고딕", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 23


def _write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], widths: list[float]) -> int:
    thin = Side(style="thin", color=BORDER_GRAY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for index, header in enumerate(headers, start=2):
        c = ws.cell(start_row, index, header)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for r_offset, row_data in enumerate(rows, start=1):
        for c_offset, value in enumerate(row_data, start=2):
            c = ws.cell(start_row + r_offset, c_offset, value)
            c.font = Font(name="맑은 고딕", size=10, color="344054")
            c.fill = PatternFill("solid", fgColor=WHITE if r_offset % 2 else "F8FAFC")
            c.alignment = Alignment(
                horizontal="right" if c_offset == 4 and len(headers) >= 3 else "left",
                vertical="center",
                wrap_text=True,
            )
            c.border = border
        ws.row_dimensions[start_row + r_offset].height = 22
    for idx, width in enumerate(widths, start=2):
        ws.column_dimensions[get_column_letter(idx)].width = width
    return start_row + len(rows)


def build_excel_report(
    result: Result,
    input_data: Dict[str, Any],
    available_cash: float,
    death_benefit: float,
    other_liquidity: float,
) -> bytes:
    """현재 계산 결과를 서식이 적용된 4개 시트 엑셀 보고서로 만든다."""
    if Workbook is None:
        raise RuntimeError("openpyxl이 설치되어 있지 않습니다. 'pip install openpyxl'을 실행해 주세요.")

    liquid_funds = max(0, available_cash) + max(0, death_benefit) + max(0, other_liquidity)
    tax_due = result.estimated_tax_due
    gap = liquid_funds - tax_due
    ratio = liquid_funds / tax_due if tax_due > 0 else None
    restricted = max(0, result.deduction_before_limit - result.allowed_deduction)

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color=BORDER_GRAY)
    medium = Side(style="medium", color=NAVY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 1. 한눈에 보는 결과
    ws = wb.create_sheet("한눈에 보는 결과")
    _excel_title(ws, "상속세 예상 계산 결과", "상담 및 사전 검토용")
    for col, width in {"A": 3, "B": 18, "C": 18, "D": 4, "E": 18, "F": 18, "G": 4, "H": 18, "I": 18}.items():
        ws.column_dimensions[col].width = width

    cards = [
        (2, 5, 3, "예상 상속세", won_text(tax_due), LIGHT_BLUE, NAVY),
        (5, 5, 6, "준비된 납부재원", won_text(liquid_funds), LIGHT_GREEN, GREEN),
        (8, 5, 9, "예상 부족액" if gap < 0 else "예상 여유액", won_text(abs(gap)), LIGHT_RED if gap < 0 else LIGHT_GREEN, RED if gap < 0 else GREEN),
    ]
    for start_col, start_row, end_col, label, value, fill_color, font_color in cards:
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=end_col)
        label_cell = ws.cell(start_row, start_col, label)
        label_cell.font = Font(name="맑은 고딕", size=10, bold=True, color=GRAY)
        label_cell.fill = PatternFill("solid", fgColor=fill_color)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell = ws.cell(start_row + 1, start_col, value)
        value_cell.font = Font(name="맑은 고딕", size=19 if start_col < 8 else 21, bold=True, color=font_color)
        value_cell.fill = PatternFill("solid", fgColor=fill_color)
        value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _style_range(ws, f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row+2}", border=Border(left=medium, right=medium, top=medium, bottom=medium))
    for r in range(5, 8):
        ws.row_dimensions[r].height = 27

    ws.merge_cells("B9:I10")
    status = (
        f"예상 상속세 대비 납부재원이 {won_text(abs(gap))} 부족합니다."
        if gap < 0 else
        f"예상 상속세 납부 후 {won_text(gap)}의 자금 여유가 예상됩니다."
    )
    if tax_due <= 0:
        status = "현재 입력 기준 예상 상속세가 발생하지 않습니다."
    c = ws["B9"]
    c.value = status
    c.font = Font(name="맑은 고딕", size=13, bold=True, color=RED if gap < 0 else GREEN)
    c.fill = PatternFill("solid", fgColor=LIGHT_RED if gap < 0 else LIGHT_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.row_dimensions[9].height = 27
    ws.row_dimensions[10].height = 14

    _write_section_title(ws, 12, "계산 흐름", 2, 9)
    flow = [
        ("총상속재산", won_text(result.gross_estate)),
        ("공제 전 과세대상 재산", won_text(result.taxable_estate)),
        ("최종 적용 공제액", won_text(result.allowed_deduction)),
        ("세율이 적용되는 금액", won_text(result.tax_base)),
        ("예상 상속세", won_text(tax_due)),
    ]
    columns = [(2,3), (4,5), (6,7), (8,9), (10,11)]
    for idx, (label, value) in enumerate(flow):
        if idx >= 4:
            sc, ec = 8, 9
            row = 17
        else:
            sc, ec = columns[idx]
            row = 13
        ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        ws.merge_cells(start_row=row+1, start_column=sc, end_row=row+2, end_column=ec)
        ws.cell(row, sc, label).font = Font(name="맑은 고딕", size=9, bold=True, color=GRAY)
        ws.cell(row, sc).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row+1, sc, value).font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
        ws.cell(row+1, sc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _style_range(ws, f"{get_column_letter(sc)}{row}:{get_column_letter(ec)}{row+2}", fill=PatternFill("solid", fgColor=WHITE), border=border)
    ws.merge_cells("B17:G19")
    ws["B17"] = f"납부재원 충족률\n{ratio:.1%}" if ratio is not None else "납부재원 충족률\n해당 없음"
    ws["B17"].font = Font(name="맑은 고딕", size=15, bold=True, color=GREEN if gap >= 0 else ORANGE)
    ws["B17"].fill = PatternFill("solid", fgColor=LIGHT_GREEN if gap >= 0 else LIGHT_ORANGE)
    ws["B17"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["B17"].border = border

    _write_section_title(ws, 21, "납부재원 구성", 2, 9)
    funding_rows = [
        ["현금·예금", won_text(available_cash)],
        ["사망보험금", won_text(death_benefit)],
        ["기타 즉시 사용 가능 자금", won_text(other_liquidity)],
        ["총 납부재원", won_text(liquid_funds)],
    ]
    end = _write_table(ws, 22, ["납부재원 구분", "금액"], funding_rows, [34, 28])
    # 납부재원 구성 표 전체 가운데 정렬
    for row in range(22, end + 1):
        for col in range(2, 4):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[end][1:3]:
        cell.font = Font(name="맑은 고딕", size=11, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.freeze_panes = "B5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "B2:I27"
    ws.sheet_view.zoomScale = 90

    # 2. 상속공제 상세
    ws = wb.create_sheet("상속공제 상세")
    _excel_title(ws, "상속공제 상세", "현재 입력 기준 계산 반영 내역이며 실제 요건 충족 여부는 별도 확인이 필요합니다.")
    _write_section_title(ws, 6, "공제 항목별 구성", 2, 8)
    deduction_items = [
        ("일괄·인적공제", result.personal_or_lump, "선택한 공제 방식에 따른 계산금액"),
        ("배우자공제", result.spouse_deduction, "실제 상속금액과 법정상속지분 반영"),
        ("금융재산공제", result.financial_deduction, "순금융재산 기준"),
        ("동거주택공제", result.home_deduction, "입력한 공제대상 주택가액 기준"),
        ("기타 공제", result.other_deduction, "가업·영농 등 입력금액"),
    ]
    deduction_rows = [[name, "계산 반영" if amount > 0 else "입력 없음", won_text(amount), desc] for name, amount, desc in deduction_items]
    end = _write_table(ws, 7, ["공제 항목", "현재 계산 상태", "계산 금액", "설명"], deduction_rows, [24, 17, 22, 43])
    # 공제 항목별 구성 표 전체 가운데 정렬
    for row in range(7, end + 1):
        for col in range(2, 6):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(8, end + 1):
        status_cell = ws.cell(row, 3)
        if status_cell.value == "계산 반영":
            status_cell.font = Font(name="맑은 고딕", size=10, bold=True, color=GREEN)
            status_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        else:
            status_cell.font = Font(name="맑은 고딕", size=10, color=GRAY)
            status_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    summary_row = end + 3
    _write_section_title(ws, summary_row, "공제한도 적용 결과", 2, 8)
    summaries = [
        ("한도 적용 전 공제 합계", result.deduction_before_limit, LIGHT_BLUE, NAVY),
        ("실제 적용 가능한 공제 한도", result.deduction_limit, LIGHT_GRAY, "344054"),
        ("최종 적용 공제액", result.allowed_deduction, LIGHT_GREEN, GREEN),
        ("한도로 인해 제한된 금액", restricted, LIGHT_RED if restricted > 0 else LIGHT_GREEN, RED if restricted > 0 else GREEN),
    ]
    for i, (label, amount, fill_color, font_color) in enumerate(summaries):
        row = summary_row + 2 + i * 2
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        ws.cell(row, 2, label).font = Font(name="맑은 고딕", size=10, bold=True, color=GRAY)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row, 5, won_text(amount)).font = Font(name="맑은 고딕", size=14 if i >= 2 else 12, bold=True, color=font_color)
        ws.cell(row, 5).fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")
        _style_range(ws, f"B{row}:H{row}", border=border)
        ws.row_dimensions[row].height = 26
    note_row = summary_row + 11
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row+1, end_column=8)
    note = (
        f"계산된 공제 합계는 {won_text(result.deduction_before_limit)}이나, 공제 종합한도에 따라 "
        f"{won_text(result.allowed_deduction)}이 최종 적용되었습니다."
        if restricted > 0 else
        f"계산된 공제 합계 {won_text(result.deduction_before_limit)}이 한도 내에서 전액 적용되었습니다."
    )
    ws.cell(note_row, 2, note).font = Font(name="맑은 고딕", size=10, bold=True, color=ORANGE if restricted > 0 else GREEN)
    ws.cell(note_row, 2).fill = PatternFill("solid", fgColor=LIGHT_ORANGE if restricted > 0 else LIGHT_GREEN)
    ws.cell(note_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(note_row, 2).border = border
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"B2:H{note_row+1}"

    # 3. 세액 계산 상세
    ws = wb.create_sheet("세액 계산 상세")
    _excel_title(ws, "세액 계산 상세", "재산에서 예상 납부세액까지의 계산 흐름")
    _write_section_title(ws, 6, "세액 계산 단계", 2, 7)
    expenses = max(0, input_data["public_dues"]) + max(0, input_data["funeral_expense"]) + max(0, input_data["liabilities"])
    gifts = max(0, input_data["prior_gifts_heirs"]) + max(0, input_data["prior_gifts_non_heirs"])
    tax_rows = [
        ["총상속재산", "기준", won_text(result.gross_estate), "상속재산 + 추정·간주상속재산"],
        ["비과세 재산", "차감", won_text(input_data["non_taxable"]), "과세가액에서 제외"],
        ["공과금·장례비용·채무", "차감", won_text(expenses), "과세가액에서 차감"],
        ["사전증여재산", "가산", won_text(gifts), "상속인 10년·상속인 외 5년 내 증여"],
        ["상속세 과세가액", "중간결과", won_text(result.taxable_estate), "공제 전 과세대상 재산"],
        ["최종 적용 공제액", "차감", won_text(result.allowed_deduction), "공제 종합한도 반영"],
        ["감정평가수수료", "차감", won_text(input_data["appraisal_fee"]), "과세표준 계산 시 차감"],
        ["과세표준", "중간결과", won_text(result.tax_base), "세율이 적용되는 금액"],
        ["적용세율", "세액계산", f"{result.rate:.0%}", "과세표준 구간별 누진세율"],
        ["누진공제", "차감", won_text(result.progressive_deduction), "세율 적용 후 차감"],
        ["산출세액", "중간결과", won_text(result.calculated_tax), "과세표준 × 세율 − 누진공제"],
        ["세대생략 할증", "가산", won_text(result.generation_skip_surcharge), "해당되는 경우 가산"],
        ["증여세액·기타 세액공제", "차감", won_text(result.tax_credits), "산출세액에서 차감"],
        ["신고세액공제", "차감", won_text(result.filing_credit), "기한 내 신고 가정"],
        ["예상 납부세액", "최종결과", won_text(result.estimated_tax_due), "현재 입력 기준 최종 예상치"],
    ]
    end = _write_table(ws, 7, ["계산 단계", "구분", "금액", "설명"], tax_rows, [28, 16, 25, 45])
    # 계산 단계·구분·설명 열 가운데 정렬 (금액 열은 가독성을 위해 오른쪽 정렬 유지)
    for row in range(7, end + 1):
        for col in (2, 3, 5):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    color_map = {
        "가산": (LIGHT_ORANGE, ORANGE),
        "차감": (LIGHT_BLUE, NAVY),
        "중간결과": (LIGHT_GRAY, "344054"),
        "세액계산": (LIGHT_GREEN, GREEN),
        "최종결과": (NAVY, WHITE),
    }
    for row in range(8, end + 1):
        kind = ws.cell(row, 3).value
        if kind in color_map:
            fill_color, font_color = color_map[kind]
            for col in range(2, 6):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=fill_color)
                ws.cell(row, col).font = Font(name="맑은 고딕", size=11 if kind == "최종결과" else 10, bold=kind in {"중간결과", "최종결과"}, color=font_color)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"B2:E{end}"

    # 4. 입력 내역
    ws = wb.create_sheet("입력 내역")
    _excel_title(ws, "입력 내역", "계산 시 사용된 원본 입력값 기록")
    sections = [
        ("기본 재산", [
            ("상속재산가액", input_data["gross_estate"]),
            ("추정·간주상속재산", input_data["deemed_estate"]),
            ("비과세·과세가액 불산입액", input_data["non_taxable"]),
            ("공과금", input_data["public_dues"]),
            ("실제 일반 장례비용", input_data["funeral_actual"]),
            ("일반 장례비용 공제액", input_data["general_funeral_deduction"]),
            ("봉안시설·자연장지 실제 비용", input_data["burial_actual"]),
            ("봉안시설·자연장지 공제액", input_data["burial_deduction"]),
            ("최종 장례비용 공제액", input_data["funeral_expense"]),
            ("피상속인 채무", input_data["liabilities"]),
            ("10년 이내 상속인 증여재산", input_data["prior_gifts_heirs"]),
            ("5년 이내 상속인 외 증여재산", input_data["prior_gifts_non_heirs"]),
        ]),
        ("상속인·공제", [
            ("공제 방식", "일괄·인적 비교" if input_data["lump_mode"] else "기초공제 + 인적공제"),
            ("자녀공제 대상 자녀 수", input_data["children_count"]),
            ("65세 이상 연로자 수", input_data["elderly_count"]),
            ("배우자 생존 여부", "예" if input_data["spouse_exists"] else "아니오"),
            ("배우자 실제 상속금액", input_data["spouse_actual_inheritance"]),
            ("배우자 법정상속지분", input_data["spouse_share"]),
            ("배우자 사전증여재산 과세표준", input_data["spouse_prior_gift_tax_base"]),
            ("순금융재산가액", input_data["net_financial_assets"]),
            ("공제대상 동거주택가액", input_data["cohabiting_home_value"]),
            ("기타 공제", input_data["other_deduction"]),
        ]),
        ("고급 입력", [
            ("감정평가수수료 공제액", input_data["appraisal_fee"]),
            ("상속인 아닌 자에 대한 유증 등", input_data["non_heir_bequest"]),
            ("상속포기로 다음 순위가 받은 재산", input_data["inheritance_waiver_next_rank"]),
            ("공제한도 차감 사전증여 과세표준", input_data["prior_gift_tax_base_for_limit"]),
            ("세대를 건너뛴 상속재산가액", input_data["generation_skip_amount"]),
            ("증여세액공제", input_data["gift_tax_credit"]),
            ("기타 세액공제", input_data["other_tax_credit"]),
            ("신고세액공제 적용", "예" if input_data["apply_filing_credit"] else "아니오"),
        ]),
        ("납부재원", [
            ("현금·예금", available_cash),
            ("사망보험금", death_benefit),
            ("기타 즉시 사용 가능 자금", other_liquidity),
            ("총 납부재원", liquid_funds),
        ]),
    ]
    current_row = 6
    money_labels = {label for _, items in sections for label, value in items if isinstance(value, (int, float)) and "수" not in label and "지분" not in label}
    for section, items in sections:
        _write_section_title(ws, current_row, section, 2, 6)
        current_row += 1
        for label, value in items:
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
            ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
            ws.cell(current_row, 2, label).font = Font(name="맑은 고딕", size=10, bold=True, color="344054")
            ws.cell(current_row, 2).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            ws.cell(current_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            display = won_text(value) if label in money_labels and isinstance(value, (int, float)) else value
            ws.cell(current_row, 4, display).font = Font(name="맑은 고딕", size=10, color=NAVY)
            ws.cell(current_row, 4).fill = PatternFill("solid", fgColor=WHITE)
            ws.cell(current_row, 4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            _style_range(ws, f"B{current_row}:F{current_row}", border=border)
            current_row += 1
        current_row += 1
    for col, width in {"A":3, "B":22, "C":12, "D":18, "E":12, "F":18}.items():
        ws.column_dimensions[col].width = width
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"B2:F{current_row}"

    for sheet in wb.worksheets:
        # 모든 다운로드 시트의 기본 인쇄 용지를 A4로 고정
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.4
        sheet.page_margins.bottom = 0.4
        sheet.oddFooter.center.text = "상담용 예상치이며 실제 신고 전 별도 검토가 필요합니다."
        sheet.oddFooter.center.size = 8
        sheet.oddFooter.center.color = GRAY

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# -----------------------------------------------------------------------------
# Streamlit 화면
# -----------------------------------------------------------------------------
def run():
    page_header("고객 상담", "상속세 예상 계산기", "상속재산과 공제 항목을 입력하여 예상 상속세와 납부재원 부족액을 확인합니다.", "IT")
    st.info("모든 금액은 **만원 단위**로 입력합니다. 입력값은 자동으로 억·만원 단위로 해석해 표시합니다.")

    render_sidebar()
    render_quick_actions()
    st.divider()

    tab1, tab2, tab3 = st.tabs(["① 기본 재산", "② 공제·상속인", "③ 고급 입력"])

    with tab1:
        st.caption("예상세액 계산에 필요한 기본 재산과 차감 항목을 입력하세요.")
        c1, c2 = st.columns(2)
        with c1:
            gross_estate = amount_input(
                "상속재산가액 · 필수",
                value=100_000,
                key="it_gross",
                help_text="상속개시일 현재 피상속인이 보유한 부동산, 예금, 주식 등 상속재산 평가액입니다.",
            )
            deemed_estate = amount_input(
                "추정·간주상속재산 · 선택",
                value=0,
                key="it_deemed",
                help_text="세법상 상속재산으로 추정하거나 간주하는 재산가액입니다.",
            )
            non_taxable = amount_input(
                "비과세·과세가액 불산입액 · 선택",
                value=0,
                key="it_nontax",
            )
        with c2:
            public_dues = amount_input("공과금 · 선택", value=0, key="it_dues")
            funeral_actual = amount_input(
                "실제 일반 장례비용",
                value=0,
                key="it_funeral_actual",
                help_text="피상속인의 사망일부터 장례일까지 직접 소요된 일반 장례비용입니다. 공제액은 최소 500만원, 최대 1,000만원으로 자동 계산합니다.",
            )
            burial_actual = amount_input(
                "봉안시설·자연장지 실제 비용",
                value=0,
                key="it_burial_actual",
                help_text="봉안시설 또는 자연장지 사용에 실제 소요된 비용입니다. 최대 500만원까지 자동 반영합니다.",
            )
            liabilities = amount_input(
                "피상속인 채무 · 중요",
                value=0,
                key="it_liab",
                help_text="상속개시일 현재 피상속인이 부담하는 확정 채무를 입력합니다.",
            )
        general_funeral_deduction = min(max(funeral_actual, 500), 1_000)
        burial_deduction = min(max(0, burial_actual), 500)
        funeral_expense = general_funeral_deduction + burial_deduction
        st.info(
            f"자동 계산된 장례비용 공제액: **{won_text(funeral_expense)}**  \n"
            f"일반 장례비용 {won_text(general_funeral_deduction)}"
            f" + 봉안시설·자연장지 {won_text(burial_deduction)}"
        )
        c3, c4 = st.columns(2)
        with c3:
            prior_gifts_heirs = amount_input(
                "10년 이내 상속인 증여재산 · 중요",
                value=0,
                key="it_gift_h",
            )
        with c4:
            prior_gifts_non_heirs = amount_input(
                "5년 이내 상속인 외 증여재산 · 중요",
                value=0,
                key="it_gift_n",
            )

    with tab2:
        st.caption("배우자와 상속인의 구성 및 적용 가능한 공제 항목을 입력하세요.")
        lump_mode = st.radio(
            "공제 방식",
            ["일괄공제와 인적공제 중 큰 금액", "기초공제 + 인적공제"],
            horizontal=True,
            key="it_mode",
            help="기본적으로 적용 가능한 방식 중 유리한 금액을 선택합니다.",
        ) == "일괄공제와 인적공제 중 큰 금액"

        a1, a2 = st.columns(2)
        with a1:
            children_count = st.number_input(
                "자녀공제 대상 자녀 수",
                min_value=0,
                value=1,
                step=1,
                key="it_children",
                help="자녀공제 계산에 반영할 자녀 수입니다.",
            )
            elderly_count = st.number_input("65세 이상 연로자 수", min_value=0, value=0, step=1, key="it_elderly")
        with a2:
            minor_deduction = amount_input("미성년자공제 합계액", value=0, key="it_minor")
            disability_deduction = amount_input("장애인공제 합계액", value=0, key="it_disability")

        spouse_exists = st.toggle("배우자가 생존해 있음", value=True, key="it_spouse_exists")
        spouse_actual_inheritance, spouse_share = 0.0, 0.0
        if spouse_exists:
            b1, b2 = st.columns(2)
            with b1:
                group = st.selectbox(
                    "배우자와 공동상속하는 상속인",
                    ["직계비속", "직계존속", "배우자 단독"],
                    key="it_group",
                )
                count = 0 if group == "배우자 단독" else st.number_input(
                    f"배우자를 제외한 {group} 상속인 수",
                    min_value=1,
                    value=1,
                    step=1,
                    key="it_count",
                    help=f"배우자와 함께 상속받는 {group} 인원만 입력하세요. 배우자는 포함하지 않습니다.",
                )
            with b2:
                spouse_actual_inheritance = amount_input(
                    "배우자가 실제 상속받는 금액 · 중요",
                    value=50_000,
                    key="it_spouse_amount",
                    help_text="배우자가 실제로 상속받는 재산가액입니다.",
                    show_default_notice=True,
                )
                calculated_share = float(spouse_statutory_share(group, count))
                spouse_share = calculated_share
                st.metric("자동 계산된 배우자 법정상속지분", f"{spouse_share:.2%}")
                if group == "배우자 단독":
                    st.caption("배우자 단독상속으로 일괄공제를 제외하고 기초공제와 인적공제를 반영합니다.")

        d1, d2 = st.columns(2)
        with d1:
            net_financial_assets = amount_input(
                "순금융재산가액 · 중요",
                value=0,
                key="it_fin",
                help_text="금융재산에서 금융채무를 차감한 금액으로 금융재산공제 계산에 사용합니다.",
            )
            cohabiting_home_value = amount_input("공제대상 동거주택가액 · 선택", value=0, key="it_home")
        with d2:
            other_deduction = amount_input("가업·영농 등 기타 공제 · 선택", value=0, key="it_other_ded")
            appraisal_fee = amount_input("감정평가수수료 공제액 · 선택", value=0, key="it_appraisal")

    with tab3:
        st.caption("세대생략, 유증, 세액공제 등 해당되는 경우에만 입력하세요.")
        e1, e2, e3 = st.columns(3)
        with e1:
            non_heir_bequest = amount_input("상속인 아닌 자에 대한 유증 등", value=0, key="it_bequest")
        with e2:
            inheritance_waiver_next_rank = amount_input("상속포기로 다음 순위가 받은 재산", value=0, key="it_waiver")
        with e3:
            prior_gift_tax_base_for_limit = amount_input("공제한도 차감 사전증여 과세표준", value=0, key="it_prior_base")
            spouse_prior_gift_tax_base = amount_input(
                "배우자의 사전증여재산 과세표준",
                value=0,
                key="it_spouse_prior_base",
                help_text="배우자가 피상속인으로부터 사전에 증여받은 재산의 증여세 과세표준입니다. 해당 사항이 없으면 0으로 둡니다.",
            )
        f1, f2 = st.columns(2)
        with f1:
            generation_skip_amount = amount_input("세대를 건너뛴 상속재산가액", value=0, key="it_gen")
            generation_skip_minor_over_2b = st.checkbox(
                "미성년자가 세대생략으로 20억원 초과 상속",
                key="it_gen_minor",
            )
        with f2:
            gift_tax_credit = amount_input("사전증여 관련 증여세액공제", value=0, key="it_gift_credit")
            other_tax_credit = amount_input("기타 세액공제", value=0, key="it_other_credit")
            apply_filing_credit = st.checkbox(
                "기한 내 신고세액공제 3% 적용",
                value=True,
                key="it_filing",
                help="기본값으로 적용되어 있습니다. 실제 적용 요건을 확인하세요.",
            )

    # 기존 calculate() 함수에 기존 입력값만 전달한다.
    result = calculate(
        gross_estate=gross_estate,
        deemed_estate=deemed_estate,
        non_taxable=non_taxable,
        public_dues=public_dues,
        funeral_expense=funeral_expense,
        liabilities=liabilities,
        prior_gifts_heirs=prior_gifts_heirs,
        prior_gifts_non_heirs=prior_gifts_non_heirs,
        lump_mode=lump_mode,
        children_count=children_count,
        elderly_count=elderly_count,
        minor_deduction=minor_deduction,
        disability_deduction=disability_deduction,
        spouse_exists=spouse_exists,
        spouse_solo=spouse_exists and group == "배우자 단독",
        spouse_actual_inheritance=spouse_actual_inheritance,
        spouse_share=spouse_share,
        spouse_prior_gift_tax_base=spouse_prior_gift_tax_base,
        net_financial_assets=net_financial_assets,
        cohabiting_home_value=cohabiting_home_value,
        other_deduction=other_deduction,
        appraisal_fee=appraisal_fee,
        non_heir_bequest=non_heir_bequest,
        inheritance_waiver_next_rank=inheritance_waiver_next_rank,
        prior_gift_tax_base_for_limit=prior_gift_tax_base_for_limit,
        generation_skip_amount=generation_skip_amount,
        generation_skip_minor_over_2b=generation_skip_minor_over_2b,
        gift_tax_credit=gift_tax_credit,
        other_tax_credit=other_tax_credit,
        apply_filing_credit=apply_filing_credit,
    )

    st.divider()
    st.subheader("💧 상속세 납부재원 점검")
    st.caption(
        "순금융재산은 금융재산공제 계산용입니다. 아래에는 상속 직후 실제로 상속세 납부에 사용할 수 있는 금액을 입력하세요."
    )
    l1, l2, l3 = st.columns(3)
    with l1:
        available_cash = amount_input(
            "즉시 사용 가능한 현금·예금",
            value=0,
            key="it_cash",
            help_text="상속 직후 인출·사용 가능한 현금과 예금을 입력합니다.",
        )
    with l2:
        death_benefit = amount_input(
            "상속으로 지급되는 사망보험금",
            value=0,
            key="it_death_benefit",
            help_text="실제로 상속세 납부재원으로 사용할 수 있는 사망보험금을 입력합니다.",
        )
    with l3:
        other_liquidity = amount_input(
            "기타 즉시 사용 가능한 자금",
            value=0,
            key="it_other_liquidity",
        )

    liquid_funds = max(0, available_cash) + max(0, death_benefit) + max(0, other_liquidity)
    tax_due = result.estimated_tax_due
    funding_gap = liquid_funds - tax_due

    st.divider()
    st.subheader("한눈에 보는 결과")
    st.caption("먼저 결론을 확인하고, 필요한 경우 아래 상세 계산을 펼쳐보세요.")

    c1, c2, c3 = st.columns(3)
    with c1:
        result_card("예상 상속세", won_text(tax_due), "현재 입력값을 기준으로 계산한 예상 납부세액입니다.", "blue")
    with c2:
        result_card("준비된 납부재원", won_text(liquid_funds), "현금·예금, 사망보험금과 기타 즉시 사용 가능 자금의 합계입니다.", "green")
    with c3:
        if tax_due <= 0:
            result_card("납부재원 상태", "예상세액 없음", "현재 입력 기준 납부할 상속세가 계산되지 않았습니다.", "gray")
        elif funding_gap < 0:
            result_card("예상 부족액", won_text(abs(funding_gap)), "예상 상속세 대비 즉시 사용할 수 있는 납부재원이 부족합니다.", "red")
        else:
            result_card("예상 여유액", won_text(funding_gap), "예상 상속세를 납부한 뒤 남는 현금성 재원입니다.", "green")

    if tax_due > 0:
        funding_ratio = liquid_funds / tax_due
        shortage_ratio = max(0.0, 1.0 - funding_ratio)
        st.markdown(f"**납부재원 충족률 {funding_ratio:.1%}**")
        st.progress(min(1.0, max(0.0, funding_ratio)))
        if funding_gap < 0:
            st.markdown(
                f"""
                <div style="
                    margin-top:0.65rem;
                    padding:0.85rem 1rem;
                    border-radius:0.75rem;
                    background:#FEF3F2;
                    color:#B42318;
                    font-size:1.32rem;
                    font-weight:800;
                    text-align:center;
                    line-height:1.45;
                ">
                    예상 상속세의 약 {shortage_ratio:.1%}가 부족합니다.
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                """
                <div style="
                    margin-top:0.65rem;
                    padding:0.85rem 1rem;
                    border-radius:0.75rem;
                    background:#ECFDF3;
                    color:#027A48;
                    font-size:1.18rem;
                    font-weight:750;
                    text-align:center;
                ">
                    현재 입력 기준으로 예상 상속세 납부재원이 충족됩니다.
                </div>
                """,
                unsafe_allow_html=True,
            )
    else:
        funding_ratio = None
        st.caption("예상 상속세가 없어 납부재원 충족률을 계산하지 않습니다.")

    render_result_interpretation(tax_due, liquid_funds)

    st.markdown("### 계산 흐름")
    flow_cols = st.columns([1, 0.12, 1, 0.12, 1, 0.12, 1])
    with flow_cols[0]:
        flow_card("총상속재산", won_text(result.gross_estate), "상속재산 + 간주재산")
    with flow_cols[1]:
        st.markdown("<div style='text-align:center;padding-top:38px;font-size:1.5rem;color:#98A2B3;'>→</div>", unsafe_allow_html=True)
    with flow_cols[2]:
        flow_card("공제 전 과세대상 재산", won_text(result.taxable_estate), "상속세 과세가액")
    with flow_cols[3]:
        st.markdown("<div style='text-align:center;padding-top:38px;font-size:1.5rem;color:#98A2B3;'>→</div>", unsafe_allow_html=True)
    with flow_cols[4]:
        flow_card("최종 적용 공제액", won_text(result.allowed_deduction), "공제 종합한도 반영")
    with flow_cols[5]:
        st.markdown("<div style='text-align:center;padding-top:38px;font-size:1.5rem;color:#98A2B3;'>→</div>", unsafe_allow_html=True)
    with flow_cols[6]:
        flow_card("세율이 적용되는 금액", won_text(result.tax_base), "과세표준")

    if result.deduction_before_limit > result.allowed_deduction:
        restricted = result.deduction_before_limit - result.allowed_deduction
        st.warning(
            f"공제 합계 **{won_text(result.deduction_before_limit)}** 중 공제 종합한도에 따라 "
            f"**{won_text(result.allowed_deduction)}**이 적용되었으며, **{won_text(restricted)}**은 제한되었습니다."
        )
    else:
        restricted = 0.0
        st.success(f"계산된 공제 합계 **{won_text(result.deduction_before_limit)}**이 한도 내에서 전액 적용되었습니다.")

    with st.expander("상속공제 상세 · 어떤 공제가 얼마 반영됐는지 보기"):
        st.markdown("#### 공제 항목별 구성")
        deduction_items = [
            ("일괄·인적공제", result.personal_or_lump, "선택한 공제 방식에 따른 계산금액"),
            ("배우자공제", result.spouse_deduction, "실제 상속금액과 법정상속지분 반영"),
            ("금융재산공제", result.financial_deduction, "순금융재산 기준"),
            ("동거주택공제", result.home_deduction, "입력한 공제대상 주택가액 기준"),
            ("기타 공제", result.other_deduction, "가업·영농 등 입력금액"),
        ]
        st.dataframe(
            {
                "공제 항목": [x[0] for x in deduction_items],
                "현재 계산 상태": ["계산 반영" if x[1] > 0 else "입력 없음" for x in deduction_items],
                "계산 금액": [won_text(x[1]) for x in deduction_items],
                "쉽게 보는 설명": [x[2] for x in deduction_items],
            },
            hide_index=True,
            use_container_width=True,
        )

        st.markdown("#### 공제한도 적용 결과")
        d1, d2, d3, d4 = st.columns(4)
        d1.metric("한도 적용 전 공제 합계", won_text(result.deduction_before_limit))
        d2.metric("실제 적용 가능한 공제 한도", won_text(result.deduction_limit))
        d3.metric("최종 적용 공제액", won_text(result.allowed_deduction))
        d4.metric("제한된 금액", won_text(restricted))
        if restricted > 0:
            st.info(
                f"계산된 공제 합계는 {won_text(result.deduction_before_limit)}이나, 공제 종합한도에 따라 "
                f"{won_text(result.allowed_deduction)}이 최종 적용되었습니다."
            )
        else:
            st.info("현재 계산된 공제 합계가 공제 종합한도 내에서 전액 적용되었습니다.")

    with st.expander("세액 계산 상세 · 재산에서 납부세액까지 보기"):
        expenses_total = public_dues + funeral_expense + liabilities
        gifts_total = prior_gifts_heirs + prior_gifts_non_heirs
        st.dataframe(
            {
                "계산 단계": [
                    "총상속재산", "비과세 재산", "공과금·장례비용·채무", "사전증여재산",
                    "상속세 과세가액", "최종 적용 공제액", "감정평가수수료", "과세표준",
                    "적용세율", "누진공제", "산출세액", "세대생략 할증",
                    "증여세액·기타 세액공제", "신고세액공제", "예상 납부세액",
                ],
                "구분": [
                    "기준", "차감", "차감", "가산", "중간결과", "차감", "차감", "중간결과",
                    "세액계산", "차감", "중간결과", "가산", "차감", "차감", "최종결과",
                ],
                "금액": [
                    won_text(result.gross_estate), won_text(non_taxable), won_text(expenses_total), won_text(gifts_total),
                    won_text(result.taxable_estate), won_text(result.allowed_deduction), won_text(appraisal_fee), won_text(result.tax_base),
                    f"{result.rate:.0%}", won_text(result.progressive_deduction), won_text(result.calculated_tax),
                    won_text(result.generation_skip_surcharge), won_text(result.tax_credits), won_text(result.filing_credit), won_text(tax_due),
                ],
                "쉽게 보는 설명": [
                    "상속재산과 추정·간주상속재산 합계", "과세대상에서 제외", "과세가액에서 차감", "일정 기간 내 증여재산 합산",
                    "공제 전 과세대상 재산", "공제 종합한도 반영", "과세표준 계산 시 차감", "세율이 적용되는 금액",
                    "과세표준 구간별 누진세율", "세율 적용 후 차감", "과세표준 × 세율 − 누진공제", "해당되는 경우 가산",
                    "산출세액에서 차감", "기한 내 신고 가정", "현재 입력 기준 최종 예상치",
                ],
            },
            hide_index=True,
            use_container_width=True,
        )

    with st.expander("납부재원 상세 · 준비된 현금성 자금 보기"):
        st.dataframe(
            {
                "납부재원 구분": ["현금·예금", "사망보험금", "기타 즉시 사용 가능 자금", "총 납부재원", "예상 상속세", "부족액" if funding_gap < 0 else "여유액"],
                "금액": [won_text(available_cash), won_text(death_benefit), won_text(other_liquidity), won_text(liquid_funds), won_text(tax_due), won_text(abs(funding_gap))],
            },
            hide_index=True,
            use_container_width=True,
        )

    input_data = {
        "gross_estate": gross_estate,
        "deemed_estate": deemed_estate,
        "non_taxable": non_taxable,
        "public_dues": public_dues,
        "funeral_actual": funeral_actual,
        "general_funeral_deduction": general_funeral_deduction,
        "burial_actual": burial_actual,
        "burial_deduction": burial_deduction,
        "funeral_expense": funeral_expense,
        "liabilities": liabilities,
        "prior_gifts_heirs": prior_gifts_heirs,
        "prior_gifts_non_heirs": prior_gifts_non_heirs,
        "lump_mode": lump_mode,
        "children_count": children_count,
        "elderly_count": elderly_count,
        "minor_deduction": minor_deduction,
        "disability_deduction": disability_deduction,
        "spouse_exists": spouse_exists,
        "spouse_actual_inheritance": spouse_actual_inheritance,
        "spouse_share": spouse_share,
        "spouse_prior_gift_tax_base": spouse_prior_gift_tax_base,
        "net_financial_assets": net_financial_assets,
        "cohabiting_home_value": cohabiting_home_value,
        "other_deduction": other_deduction,
        "appraisal_fee": appraisal_fee,
        "non_heir_bequest": non_heir_bequest,
        "inheritance_waiver_next_rank": inheritance_waiver_next_rank,
        "prior_gift_tax_base_for_limit": prior_gift_tax_base_for_limit,
        "generation_skip_amount": generation_skip_amount,
        "generation_skip_minor_over_2b": generation_skip_minor_over_2b,
        "gift_tax_credit": gift_tax_credit,
        "other_tax_credit": other_tax_credit,
        "apply_filing_credit": apply_filing_credit,
    }

    st.markdown("### 결과 보고서 다운로드")
    st.caption("서식이 적용된 엑셀 파일에는 요약, 상속공제 상세, 세액 계산 상세, 입력 내역의 4개 시트가 포함됩니다.")
    if Workbook is None:
        st.error("엑셀 다운로드를 사용하려면 터미널에서 `pip install openpyxl`을 실행해 주세요.")
    else:
        try:
            excel_bytes = build_excel_report(result, input_data, available_cash, death_benefit, other_liquidity)
            file_name = f"상속세_예상계산_상담보고서_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
            st.download_button(
                "📥 서식 적용 엑셀 보고서 다운로드",
                data=excel_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as exc:
            st.error(f"엑셀 보고서를 생성하지 못했습니다: {exc}")

    st.info(
        "상담용 예상치입니다. 실제 신고 시 상속관계, 재산평가, 사전증여 내역과 공제 요건을 별도로 확인해야 합니다."
    )


if __name__ == "__main__":
    run()
